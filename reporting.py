"""
============================================================
 reporting.py — SHARED PDF/Excel export helpers
============================================================
The Super Admin and Trainer dashboards both generate server-side
PDF (reportlab) and Excel (openpyxl) exports. These helpers were
extracted from superadmin.py so neither role module duplicates the
export architecture — both call this single module.

Everything here is pure/database-agnostic except
attendance_counts_for_date and overall_report_rows, which take the
Mongo db (and users) as explicit parameters.
"""

import io
import re
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from quiz_common import iso_utc


def safe_filename(name):
    """Sanitize a name for use in a download filename (drops any
    characters that would be invalid in a Windows/Linux filename)."""
    cleaned = re.sub(r'[^A-Za-z0-9 _\-]+', '', str(name or 'report')).strip().replace(' ', '_')
    return cleaned[:60] or 'report'


def excel_bytes(columns, rows, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(columns)
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F9E93")
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(["" if v is None else v for v in r])
    for i, col_name in enumerate(columns, start=1):
        longest = len(str(col_name))
        for r in rows:
            longest = max(longest, len(str(r[i - 1])) if r[i - 1] is not None else 0)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(longest + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def pdf_col_widths(doc, columns, data):
    """Allocate the table's full horizontal width (landscape page) to
    columns in proportion to their content. Guarantees the widths sum to
    exactly the usable width; long values are wrapped by the Paragraph
    cells, never clipped or overlapped."""
    usable = doc.width
    n = len(columns)
    if not n:
        return []
    raw = []
    for ci in range(n):
        longest = len(str(columns[ci]))
        for row in data[1:]:
            val = row[ci] if ci < len(row) else ""
            longest = max(longest, len(str(val)))
        raw.append(max(6, min(longest, 70)))
    total = sum(raw) or 1
    return [usable * (w / total) for w in raw]


def pdf_bytes(title, subtitle, columns, rows):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("RTitle", parent=styles["Title"], fontSize=15, spaceAfter=4)
    sub_style = ParagraphStyle(
        "RSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=12)
    cell_style = ParagraphStyle(
        "RCell", parent=styles["Normal"], fontSize=7.5, leading=9.5)
    head_style = ParagraphStyle(
        "RHead", parent=cell_style, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER)
    data = [columns] + [["" if v is None else str(v) for v in row] for row in rows]
    col_widths = pdf_col_widths(doc, columns, data)
    body = [[Paragraph(c, head_style) for c in data[0]]]
    body.extend([Paragraph(c, cell_style) for c in row] for row in data[1:])
    table = Table(body, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F9E93")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEE2")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8CDB8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story = [Paragraph(title, title_style), Paragraph(subtitle, sub_style), table]
    doc.build(story)
    buf.seek(0)
    return buf


def fmt_value(v):
    return round(v, 1) if isinstance(v, (int, float)) else v


def pdf_sections(title, subtitle, sections):
    """PDF with multiple labelled table sections (profile + performance +
    interventions + activity in the student report). Each section is a
    (heading, columns, rows) tuple rendered as its own table, on a
    landscape page so wide tables use the full horizontal width."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("RTitle", parent=styles["Title"], fontSize=15, spaceAfter=4)
    sub_style = ParagraphStyle(
        "RSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=10)
    head_style = ParagraphStyle(
        "RSec", parent=styles["Heading2"], fontSize=11, spaceBefore=12, spaceAfter=6,
        textColor=colors.HexColor("#1F9E93"), keepWithNext=1)
    cell_style = ParagraphStyle(
        "RCell", parent=styles["Normal"], fontSize=7.5, leading=9.5)
    cell_head_style = ParagraphStyle(
        "RHead", parent=cell_style, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER)
    story = [Paragraph(title, title_style), Paragraph(subtitle, sub_style)]
    for heading, columns, rows in sections:
        story.append(Paragraph(heading, head_style))
        data = [columns] + [["" if v is None else str(v) for v in row] for row in rows]
        col_widths = pdf_col_widths(doc, columns, data)
        body = [[Paragraph(c, cell_head_style) for c in data[0]]]
        body.extend([Paragraph(c, cell_style) for c in row] for row in data[1:])
        table = Table(body, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F9E93")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEE2")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8CDB8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
    doc.build(story)
    buf.seek(0)
    return buf


def attendance_counts_for_date(db, date_str):
    """Present / absent / total attendance records for one date from
    db.attendance — used to size participation on intervention sessions."""
    if not date_str:
        return {"total": 0, "present": 0, "absent": 0}
    present = db.attendance.count_documents({"date": date_str, "status": "present"})
    absent = db.attendance.count_documents({"date": date_str, "status": "absent"})
    return {"total": present + absent, "present": present, "absent": absent}


def overall_report_rows(db, users, student_oid=None, quiz_ids=None, college=None, colleges=None):
    """Shared row builder for the Overall Report JSON + all of its
    PDF/Excel exports. Every row is computed live from db.quiz_attempts
    + db.users.finalEmployabilityScore. Optional quiz_ids (list of
    assessment ObjectIds), college (exact college name) and colleges
    (list of exact college names) scope the query so the Assessment
    Report export never has to load the whole platform attempt set."""
    query = {"status": "submitted"}
    if student_oid:
        query["studentId"] = student_oid
    if quiz_ids:
        query["quizId"] = {"$in": list(quiz_ids)}
    if college:
        query["college"] = college
    if colleges:
        query["college"] = {"$in": list(colleges)}
    attempts = list(db.quiz_attempts.find(query).sort("submittedAt", -1))
    user_scores = {}
    if student_oid:
        stu = users.find_one({"_id": student_oid}, {"finalEmployabilityScore": 1})
        if stu:
            user_scores[student_oid] = stu.get("finalEmployabilityScore")
    else:
        ids = {a.get("studentId") for a in attempts if a.get("studentId")}
        for s in users.find({"_id": {"$in": list(ids)}}, {"finalEmployabilityScore": 1}):
            user_scores[s["_id"]] = s.get("finalEmployabilityScore")
    rows = []
    for a in attempts:
        overall = a.get("overall") or {}
        rows.append({
            "attemptId": str(a["_id"]),
            "studentId": str(a["studentId"]) if a.get("studentId") else None,
            "studentName": a.get("studentName"),
            "rollNumber": a.get("studentRollNumber"),
            "department": a.get("department"),
            "college": a.get("college"),
            "assessmentId": str(a["quizId"]) if a.get("quizId") else None,
            "assessmentName": a.get("quizTitle") or "Quiz",
            "quizMarks": overall.get("percentage"),
            "quizMarksObtained": overall.get("marksObtained") or overall.get("obtainedMarks"),
            "quizTotalMarks": overall.get("totalMarks") or overall.get("total"),
            "passFail": overall.get("passFail"),
            "interviewMarks": a.get("interviewMarks"),
            "averageMarks": a.get("finalAverage"),
            "finalOverallMarks": user_scores.get(a.get("studentId")) if a.get("studentId") else None,
            "submittedAt": iso_utc(a.get("submittedAt")),
        })
    return rows
