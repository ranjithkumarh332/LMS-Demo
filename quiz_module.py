"""
============================================================
 quiz_module.py — "Create Quiz" backend (Assessment > Create Quiz)
============================================================
Dedicated, self-contained backend for the Create-Quiz feature used by
both the Trainer Dashboard and the Super Admin Dashboard's "Quiz
Management" wizard. Deliberately kept separate from quiz_common.py's
cohort/placement-rule assessment engine (db.assessments) — that
system powers baseline/mid/final cohort exams drawn randomly from a
shared question bank, which is a different feature with different
rules (random sampling, cohort scoring, placement). This module
covers manually authored quizzes: one document per quiz, storing its
own question list inline.

Collection: db.quizzes — the single source of truth for every quiz
created from either dashboard. Nothing about a quiz is ever
hardcoded in the frontend; every read here comes straight from Mongo.

Mounted twice from app.py, once per dashboard, via init_quiz(db, scope):
  - init_quiz(db, scope="trainer")    -> /api/trainer  (own-college scoped)
  - init_quiz(db, scope="super_admin")-> /api/admin     (full platform view)

Both scopes share 100% of the validation, status-computation and
editing-restriction logic below, so the two dashboards can never
disagree about whether a quiz is editable/deletable or what its
current status is.
"""

import hashlib
import logging
import re
from datetime import datetime, timezone
from io import BytesIO

from flask import Blueprint, request, send_file, jsonify
from flask_jwt_extended import get_jwt_identity, get_jwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from quiz_common import (
    ok, error, role_required, now, to_object_id, serialize, log_activity,
    VALID_COHORT_TARGETS,
)
from question_bank import (
    sync_questions_to_bank,
    bank_availability_summary,
    validate_question_bank_config,
    draw_questions_from_bank,
)

# Dedicated logger for this module (Part 1/11: "Add Extensive Logging").
# Uses the standard `logging` module rather than print() so output goes
# through whatever handler/level the host app (app.py) configures, and
# so it can be filtered/searched in production log aggregators.
logger = logging.getLogger("quiz_module")

# ------------------------------------------------------------
# Defaults / constants
# ------------------------------------------------------------
# Seeded into db.quiz_sections exactly once if the collection is empty,
# so the "Section Distribution" / category picker has real DB-backed
# options from the very first run instead of a hardcoded frontend list.
_DEFAULT_QUIZ_SECTIONS = [
    "Communication", "Programming", "Reasoning", "Professionalism", "Interview Readiness",
]

VALID_QUIZ_STATE = {"draft", "published"}


# ------------------------------------------------------------
# Small helpers
# ------------------------------------------------------------
def parse_dt(value):
    """Parse an ISO-ish datetime string (as produced by <input type=datetime-local>
    or a full ISO string) into an aware UTC datetime. Returns None if unparsable."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    s = str(value).strip()
    if not s:
        return None
    try:
        # datetime-local inputs look like "2026-08-01T09:30"
        if "T" in s and len(s) <= 16:
            s = s + ":00"
        s = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except ValueError:
        return None


def _naive_utc(dt):
    """pymongo round-trips datetimes as naive UTC (Mongo has no tz concept),
    even though we insert timezone-aware ones — so always normalize to
    naive-UTC before comparing, regardless of which side a value came from."""
    if dt is None or not isinstance(dt, datetime):
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def compute_status(doc):
    """Dynamic status, computed fresh from current server time on every
    read — never stored/cached, so it can never go stale.
      draft      -> creator hasn't published it yet
      scheduled  -> published, current time is before start
      active     -> published, current time is within [start, end] ("Live")
      completed  -> published (or was active) and current time is after end
      cancelled  -> manually cancelled (overrides the time-based cycle)
      archived   -> manually archived by Super Admin (overrides everything)
    """
    if doc.get("archived"):
        return "archived"
    if doc.get("cancelled"):
        return "cancelled"
    if doc.get("state") != "published":
        return "draft"
    start = _naive_utc(doc.get("startDateTime"))
    end = _naive_utc(doc.get("endDateTime"))
    n = _naive_utc(now())
    if start and n < start:
        return "scheduled"
    if end and n > end:
        return "completed"
    return "active"


def editable_now(doc):
    """Business rule: fully editable before the quiz starts (draft or
    scheduled). Locked the instant the start time passes, permanently."""
    return compute_status(doc) in ("draft", "scheduled")


def deletable_now(doc):
    """A quiz can only be removed while still a Draft — once scheduled it
    can be Cancelled instead, but not deleted outright."""
    return compute_status(doc) == "draft"


def cancellable_now(doc):
    """Only a Draft or Scheduled quiz can be cancelled (Live/Completed quizzes
    run their course; use Archive to retire a Completed one instead)."""
    return compute_status(doc) in ("draft", "scheduled")


def archivable_now(doc):
    """Only a Completed or Cancelled quiz can be archived."""
    return compute_status(doc) in ("completed", "cancelled")


def edit_block_message(doc):
    status = compute_status(doc)
    if status == "active":
        return "This quiz has already started and can no longer be edited."
    if status == "completed":
        return "This quiz has been completed and is now read-only."
    if status == "cancelled":
        return "This quiz has been cancelled and can no longer be edited."
    if status == "archived":
        return "This quiz has been archived and is now read-only."
    return None


def serialize_quiz(doc):
    out = serialize(doc)
    status = compute_status(doc)
    out["status"] = status
    out["canEdit"] = editable_now(doc)
    out["canDelete"] = deletable_now(doc)
    out["canCancel"] = cancellable_now(doc)
    out["canArchive"] = archivable_now(doc)
    questions = doc.get("questions") or []
    out["totalQuestions"] = len(questions)
    out["totalMarks"] = round(sum(float(q.get("marks") or 0) for q in questions), 2)
    return out


def _actor(db):
    """Resolve {id, role, name} for whoever is making the request, from the JWT."""
    claims = get_jwt()
    uid = get_jwt_identity()
    role = claims.get("role")
    user_doc = None
    oid = to_object_id(uid)
    if oid:
        user_doc = db.users.find_one({"_id": oid})
    name = (user_doc or {}).get("fullName") or ("Super Admin" if role == "super_admin" else "Unknown")
    college = (user_doc or {}).get("college")
    return {"id": uid, "role": role, "name": name, "college": college}


# ------------------------------------------------------------
# Validation — every mandatory field from the spec, checked server-side
# regardless of what the client already validated (defense in depth).
# Accepts both the Trainer wizard's payload shape and the Super Admin
# wizard's payload shape (field names differ slightly between the two
# existing, unmodified UIs) and normalizes them into one document.
# ------------------------------------------------------------
def _first(data, *keys, default=None):
    for k in keys:
        if k in data and data[k] not in (None, ""):
            return data[k]
    return default


def _resolve_college_names(db, colleges):
    """Authoritative college names for a list of college ids — always
    looked up server-side rather than trusted from the client. See the
    long comment in validate_and_normalize for why this matters: a
    quiz's collegeNames must be real names (student eligibility compares
    against the student's college name), never the ids themselves.
    """
    if not colleges:
        return []
    id_oids = [oid for oid in (to_object_id(c) for c in colleges) if oid]
    name_by_id = {
        str(doc["_id"]): doc.get("college_name")
        for doc in db.colleges.find({"_id": {"$in": id_oids}})
    }
    return [name_by_id.get(c, c) for c in colleges]


def normalize_quiz_college_names(db, quiz):
    """Self-heals a legacy quiz document whose collegeNames array
    actually holds college ids — the exact shape produced by the old
    'college_names = college_names or colleges' fallback, before
    validate_and_normalize was fixed to always resolve real names
    server-side. Called on every read path (list/detail/student
    eligibility) so existing quizzes correct themselves automatically
    the first time they're touched — no manual edit-and-republish
    required, and no separate migration step to run beforehand.

    A real college name is never also a valid ObjectId, so that's the
    detection signal: if every entry in collegeNames parses as an
    ObjectId, it's ids-mislabeled-as-names, not real names.
    """
    names = quiz.get("collegeNames") or []
    ids = quiz.get("colleges") or []
    if not names or not ids:
        return quiz
    if not all(to_object_id(n) for n in names):
        return quiz  # already real names (or already empty/mixed) — nothing to heal
    fixed = _resolve_college_names(db, [str(c) for c in ids])
    db.quizzes.update_one({"_id": quiz["_id"]}, {"$set": {"collegeNames": fixed}})
    quiz["collegeNames"] = fixed
    return quiz


_VALID_DIFFICULTIES = {"easy", "medium", "hard"}


def _validate_section_config(raw_config, clean_questions):
    """Validates the granular per-section, per-difficulty Available/Display
    grid produced by the Create Quiz wizard's Section Distribution step
    (Part 4/5 of the spec) and cross-checks it against the question pool
    that was actually entered/uploaded.

    Rules enforced:
      - Display can never exceed Available for any (section, difficulty) cell.
      - Available for a given (section, difficulty) cell must EXACTLY match
        the number of questions actually entered/uploaded for that cell —
        no more, no less (spec: "must contain exactly ... no more, no
        less"). This also implicitly enforces total Questions Available /
        Questions Displayed / manual-entry-count / difficulty-distribution
        consistency, since every one of those is just a sum over this grid.
      - Every authored/uploaded question must belong to a configured cell —
        a stray question tagged with an unconfigured section/difficulty is
        flagged rather than silently dropped or silently counted.

    Returns (clean_section_config, section_distribution, difficulty_distribution,
             questions_available, questions_displayed, errors).
    """
    errors = []
    if not isinstance(raw_config, dict) or not raw_config:
        return {}, {}, {}, 0, 0, ["Section Distribution is required — configure at least one section."]

    def _int(v):
        try:
            n = int(v)
        except (TypeError, ValueError):
            n = 0
        return max(0, n)

    clean_config = {}
    section_distribution = {}
    difficulty_totals_display = {"easy": 0, "medium": 0, "hard": 0}
    total_available = 0
    total_display = 0

    for raw_name, cfg in raw_config.items():
        name = str(raw_name).strip()
        if not name or not isinstance(cfg, dict):
            continue
        h_a, m_a, e_a = _int(cfg.get("hardAvailable")), _int(cfg.get("mediumAvailable")), _int(cfg.get("easyAvailable"))
        h_d, m_d, e_d = _int(cfg.get("hardDisplay")), _int(cfg.get("mediumDisplay")), _int(cfg.get("easyDisplay"))
        for label, avail, disp in (("Hard", h_a, h_d), ("Medium", m_a, m_d), ("Easy", e_a, e_d)):
            if disp > avail:
                errors.append(f'"{name}" section: {label} Display ({disp}) cannot exceed {label} Available ({avail}).')
        avail_total = h_a + m_a + e_a
        disp_total = h_d + m_d + e_d
        clean_config[name] = {
            "hardAvailable": h_a, "mediumAvailable": m_a, "easyAvailable": e_a,
            "hardDisplay": h_d, "mediumDisplay": m_d, "easyDisplay": e_d,
            "questionsAvailable": avail_total, "questionsToDisplay": disp_total,
        }
        total_available += avail_total
        total_display += disp_total
        difficulty_totals_display["hard"] += h_d
        difficulty_totals_display["medium"] += m_d
        difficulty_totals_display["easy"] += e_d
        if disp_total > 0:
            section_distribution[name] = disp_total

    if total_available == 0:
        errors.append("Configure at least one section with Questions Available.")
    if total_display == 0:
        errors.append("Configure at least one section with Questions to Display.")

    # Cross-check against the pool actually authored/uploaded — every
    # configured (section, difficulty) cell must match the real count
    # exactly, and every real question must belong to a configured cell.
    pool_counts = {}
    for q in clean_questions:
        key = (q.get("section") or "", (q.get("difficulty") or "").strip().lower())
        pool_counts[key] = pool_counts.get(key, 0) + 1

    configured_keys = set()
    for name, cfg in clean_config.items():
        for level, avail_key in (("hard", "hardAvailable"), ("medium", "mediumAvailable"), ("easy", "easyAvailable")):
            configured = cfg[avail_key]
            actual = pool_counts.get((name, level), 0)
            configured_keys.add((name, level))
            if (configured or actual) and configured != actual:
                errors.append(
                    f'"{name}" section — {level.capitalize()}: configured {configured} available but '
                    f'{actual} question(s) were actually entered/uploaded.'
                )

    for (sect, level), count in pool_counts.items():
        if (sect, level) not in configured_keys and count:
            label = level.capitalize() if level in _VALID_DIFFICULTIES else (level or "no difficulty")
            errors.append(
                f'{count} question(s) found for section "{sect}" ({label}) which has no configured Available count.'
            )

    difficulty_distribution = {k: v for k, v in difficulty_totals_display.items() if v}
    return clean_config, section_distribution, difficulty_distribution, total_available, total_display, errors


def validate_and_normalize(data, db, actor, existing_id=None):
    errors = []

    title = str(_first(data, "title", "assessment_name", "name", default="")).strip()
    if not title:
        errors.append("Quiz title is required.")
    else:
        # Duplicate assessment name check — case-insensitive, excludes the
        # quiz currently being edited (existing_id) so re-saving a quiz
        # under its own unchanged name never false-positives.
        dup_query = {"title": {"$regex": f"^{re.escape(title)}$", "$options": "i"}}
        if existing_id is not None:
            dup_query["_id"] = {"$ne": existing_id}
        if db.quizzes.find_one(dup_query):
            errors.append(f'An assessment named "{title}" already exists. Choose a different name.')

    description = str(_first(data, "description", "instructions", default="")).strip()

    category = str(_first(data, "category", "assessment_type", "quizType", "type", default="")).strip()
    if not category:
        errors.append("Assessment category is required.")

    # quizType drives which pool the questions come from: "manual"/"bulk"
    # (Trainer or Super Admin authors/uploads the pool themselves, exactly
    # as before) or "question_bank" (Part 4/6 — Super Admin only; the pool
    # is drawn from the permanent QuestionBank instead of being authored
    # here). Resolved up front so every branch below can read it.
    quiz_type = str(_first(data, "quizType", "assessment_type", default="manual")).strip().lower() or "manual"
    is_question_bank_mode = quiz_type == "question_bank"
    if is_question_bank_mode and actor.get("role") != "super_admin":
        errors.append("Question Bank quiz creation is only available to Super Admin.")

    cohort_target = str(_first(data, "cohortTarget", "applicable_cohort", "cohort", default="all")).strip()
    if cohort_target and cohort_target not in VALID_COHORT_TARGETS:
        errors.append(
            f'Invalid cohort selection "{cohort_target}" — must be one of: '
            f'{", ".join(sorted(VALID_COHORT_TARGETS))}.'
        )

    colleges = _first(data, "colleges", "applicable_colleges", default=None)
    if colleges is None and _first(data, "college", default=None):
        colleges = [str(_first(data, "college"))]
    colleges = [str(c) for c in (colleges or [])]

    # Always resolve the authoritative college names server-side from the
    # ids themselves — never trust a client-sent collegeNames array. This
    # was the actual root cause of college-scoped quizzes never appearing
    # to any eligible student: the Create Quiz wizard's payload only ever
    # sent college IDs (applicable_colleges), and the old fallback quietly
    # stored those IDs in place of names when collegeNames was absent.
    # student.py's eligibility check compares a quiz's collegeNames
    # against the student's college NAME, so IDs stored there could never
    # match — silently hiding every quiz that targeted specific colleges
    # (quizzes left at "All Colleges" were unaffected, since that path
    # skips the college check entirely).
    college_names = _resolve_college_names(db, colleges)

    start_dt = parse_dt(_first(data, "startDateTime", "start_datetime", "start", "date"))
    end_dt = parse_dt(_first(data, "endDateTime", "end_datetime", "end"))
    if not start_dt:
        errors.append("Quiz start date & time is required.")
    if not end_dt:
        errors.append("Quiz end date & time is required.")
    if start_dt and end_dt and end_dt <= start_dt:
        errors.append("End date & time must be after the start date & time.")

    duration_raw = _first(data, "durationMinutes", "duration", default=None)
    try:
        duration = int(duration_raw)
    except (TypeError, ValueError):
        duration = None
    if not duration or duration < 1:
        errors.append("Duration (in minutes) is required and must be a positive number.")

    # NOTE: Passing Marks has been removed entirely (field, validation, storage) per spec.

    visibility = str(_first(data, "visibility", default="")).strip()
    if not visibility:
        visibility = "colleges" if colleges else "all"

    # --- Questions Available / Questions Displayed ------------------------
    # "Available" = the FULL pool of questions actually authored/uploaded
    # (every one of them is stored). "Displayed" = how many of those a given
    # student is randomly served on their attempt (<= Available). Available
    # is always derived from the real question count, never trusted from the
    # client, so the two can never drift out of sync with what's stored.
    questions_raw = _first(data, "questions", "question_pool", default=[]) or []
    if not is_question_bank_mode and (not isinstance(questions_raw, list) or not questions_raw):
        errors.append("At least one question is required. Enter or upload the full pool of Questions Available.")

    # The Create Quiz wizard sends a granular per-section, per-difficulty
    # Available/Display grid ("sectionConfig") instead of the older flat
    # sectionDistribution/difficultyDistribution/questionsDisplayed fields.
    # Both shapes are supported here: sectionConfig (current wizard) takes
    # priority when present; the flat fields remain supported for any other
    # caller (e.g. a legacy quiz document being re-saved without ever
    # opening the redesigned wizard).
    raw_section_config = data.get("sectionConfig")
    using_section_config = isinstance(raw_section_config, dict) and bool(raw_section_config)

    section_distribution = None
    difficulty_distribution = None
    questions_displayed = None

    if not using_section_config:
        displayed_raw = _first(data, "questionsDisplayed", "questions_displayed", default=None)
        try:
            questions_displayed = int(displayed_raw)
        except (TypeError, ValueError):
            questions_displayed = None
        if not questions_displayed or questions_displayed < 1:
            errors.append("Questions Displayed is required and must be a positive number.")
        elif isinstance(questions_raw, list) and questions_raw and questions_displayed > len(questions_raw):
            errors.append(
                f"Questions Displayed ({questions_displayed}) cannot exceed Questions Available "
                f"({len(questions_raw)} question(s) actually entered/uploaded)."
            )

        # --- Section distribution (flat / legacy) ---------------------
        section_distribution = _first(data, "sectionDistribution", "section_distribution", default=None)
        if isinstance(section_distribution, dict) and section_distribution:
            clean_dist = {}
            for k, v in section_distribution.items():
                try:
                    clean_dist[str(k)] = int(v)
                except (TypeError, ValueError):
                    clean_dist[str(k)] = 0
            section_distribution = clean_dist
            dist_sum = sum(section_distribution.values())
            if questions_displayed and dist_sum != questions_displayed:
                errors.append(
                    f"Section distribution ({dist_sum}) must sum to Questions Displayed ({questions_displayed})."
                )
        else:
            section_distribution = None

        # --- Difficulty distribution (flat / legacy, optional) ---------
        difficulty_distribution = _first(data, "difficultyDistribution", "difficulty_distribution", default=None)
        if isinstance(difficulty_distribution, dict) and any(
            difficulty_distribution.get(k) not in (None, "") for k in ("easy", "medium", "hard")
        ):
            clean_diff = {}
            for k in ("easy", "medium", "hard"):
                try:
                    clean_diff[k] = int(difficulty_distribution.get(k) or 0)
                except (TypeError, ValueError):
                    clean_diff[k] = 0
            difficulty_distribution = clean_diff
            diff_sum = sum(difficulty_distribution.values())
            if questions_displayed and diff_sum != questions_displayed:
                errors.append(
                    f"Difficulty distribution ({diff_sum}) must sum to Questions Displayed ({questions_displayed})."
                )
        else:
            difficulty_distribution = None

    clean_questions = []
    for i, q in enumerate(questions_raw if isinstance(questions_raw, list) else []):
        n = i + 1
        if not isinstance(q, dict):
            errors.append(f"Question {n} is invalid.")
            continue
        text = str(q.get("text") or "").strip()
        if not text:
            errors.append(f"Question {n}: question text is required.")
        raw_options = q.get("options") or []
        options = [str(o).strip() if o is not None else "" for o in raw_options]
        filled_options = [o for o in options if o]
        if len(filled_options) < 2:
            errors.append(f"Question {n}: at least 2 answer options are required.")

        correct = q.get("correct")
        if correct is None:
            correct = q.get("correctAnswers")
        if correct is None:
            correct = []
        # Accept either index-based ([0,2]) or letter-based (["A","C"]) correct answers.
        normalized_correct = []
        letter_map = {"A": 0, "B": 1, "C": 2, "D": 3}
        for c in (correct if isinstance(correct, list) else [correct]):
            if isinstance(c, bool):
                continue
            if isinstance(c, (int, float)):
                normalized_correct.append(int(c))
            elif isinstance(c, str) and c.strip().upper() in letter_map:
                normalized_correct.append(letter_map[c.strip().upper()])
        if not normalized_correct:
            errors.append(f"Question {n}: at least one correct answer must be selected.")
        else:
            for idx in normalized_correct:
                if idx < 0 or idx >= len(options) or not options[idx]:
                    errors.append(f"Question {n}: correct answer references an empty/invalid option.")
                    break

        q_type = str(q.get("type") or "single_choice").strip().lower()
        if q_type not in ("single_choice", "multiple_choice"):
            q_type = "single_choice"
        if q_type == "single_choice" and len(normalized_correct) > 1:
            errors.append(f"Question {n}: Single Choice questions must have exactly one correct answer.")

        marks_raw = q.get("marks")
        try:
            marks = float(marks_raw)
        except (TypeError, ValueError):
            marks = None
        if marks is None or marks <= 0:
            errors.append(f"Question {n}: marks must be a positive number.")

        section = str(q.get("section") or "").strip()
        if (using_section_config or section_distribution) and not section:
            errors.append(f"Question {n}: a section is required.")

        # Difficulty Level may ONLY be Easy / Medium / Hard — never any
        # other free-text value, and it's mandatory whenever the quiz uses
        # the granular Section Distribution grid (or the legacy flat
        # difficulty distribution).
        difficulty_raw = str(q.get("difficulty") or "").strip()
        difficulty_clean = difficulty_raw.capitalize() if difficulty_raw.lower() in _VALID_DIFFICULTIES else None
        if using_section_config or difficulty_distribution:
            if not difficulty_raw:
                errors.append(f"Question {n}: a difficulty level is required.")
            elif difficulty_raw.lower() not in _VALID_DIFFICULTIES:
                errors.append(f'Question {n}: difficulty must be Easy, Medium, or Hard (got "{difficulty_raw}").')
        elif difficulty_raw and difficulty_raw.lower() not in _VALID_DIFFICULTIES:
            errors.append(f'Question {n}: difficulty must be Easy, Medium, or Hard (got "{difficulty_raw}").')

        clean_questions.append({
            "text": text,
            "options": options,
            "correct": sorted(set(normalized_correct)),
            "type": q_type,
            "section": section or None,
            "difficulty": difficulty_clean or (difficulty_raw or None),
            "marks": marks if marks is not None else 0,
            "explanation": (q.get("explanation") or "").strip(),
        })

    # Duplicate detection (Part 3: "Duplicate question detected") — Bulk
    # Upload already catches this at validate-file time; Manual Entry never
    # did, so the exact same check (section + question text, case/whitespace
    # insensitive) is applied here too, catching duplicates regardless of
    # how the questions were entered.
    seen_question_keys = {}
    for i, q in enumerate(clean_questions, start=1):
        key = f"{(q['section'] or '').strip().lower()}|{' '.join(q['text'].strip().lower().split())}"
        if not q["text"]:
            continue
        if key in seen_question_keys:
            errors.append(f"Question {i}: duplicate of Question {seen_question_keys[key]} (same section and text).")
        else:
            seen_question_keys[key] = i

    clean_section_config = {}
    if is_question_bank_mode:
        (clean_section_config, section_distribution, difficulty_distribution,
         sc_available, questions_displayed, sc_errors) = validate_question_bank_config(db, raw_section_config)
        errors.extend(sc_errors)
    elif using_section_config:
        (clean_section_config, section_distribution, difficulty_distribution,
         sc_available, questions_displayed, sc_errors) = _validate_section_config(raw_section_config, clean_questions)
        errors.extend(sc_errors)
    else:
        # Cross-validate the legacy flat distributions against what's
        # actually in the pool — a random draw can only ever be as good as
        # what was actually stored.
        if section_distribution and clean_questions:
            pool_by_section = {}
            for q in clean_questions:
                pool_by_section[q["section"]] = pool_by_section.get(q["section"], 0) + 1
            for sect, required in section_distribution.items():
                available = pool_by_section.get(sect, 0)
                if available < required:
                    errors.append(
                        f'"{sect}" section requires at least {required} question(s) but only {available} were entered/uploaded.'
                    )

        if difficulty_distribution and clean_questions:
            pool_by_diff = {}
            for q in clean_questions:
                key = (q["difficulty"] or "").strip().lower()
                pool_by_diff[key] = pool_by_diff.get(key, 0) + 1
            for level in ("easy", "medium", "hard"):
                required = difficulty_distribution.get(level, 0)
                available = pool_by_diff.get(level, 0)
                if required and available < required:
                    errors.append(
                        f'"{level.capitalize()}" difficulty requires at least {required} question(s) but only {available} were entered/uploaded.'
                    )

    if errors:
        return None, errors

    normalized = {
        "title": title,
        "description": description,
        "category": category,
        "cohortTarget": cohort_target or "all",
        "colleges": [str(c) for c in colleges],
        "collegeNames": [str(c) for c in college_names],
        "startDateTime": start_dt,
        "endDateTime": end_dt,
        "durationMinutes": duration,
        "visibility": visibility,
        # Available is always the *actual* stored pool size — never trusted
        # from the client — so it can never drift from what's in Mongo.
        # Question Bank mode is the one exception: Available there means
        # "how many currently sit in the live QuestionBank", not "how many
        # were authored for this quiz" — sc_available already holds that
        # live count (see validate_question_bank_config above).
        "questionsAvailable": sc_available if is_question_bank_mode else len(clean_questions),
        "questionsDisplayed": questions_displayed,
        "difficultyDistribution": difficulty_distribution,
        "sectionDistribution": section_distribution,
        "sectionConfig": clean_section_config,
        "quizType": quiz_type,
        # Question Bank mode: the actual question set is drawn fresh from
        # db.question_bank at publish time (Part 7) by the create/update
        # endpoint, using clean_section_config above — never here, since
        # validate_and_normalize() runs for drafts too and a draft must
        # not consume/lock in a random draw before the admin is ready to
        # publish. Manual Entry / Bulk Upload quizzes already have their
        # real, authored question list in clean_questions.
        "questions": [] if is_question_bank_mode else clean_questions,
    }
    return normalized, []


def select_random_questions(doc):
    """Stratified random draw of `questionsDisplayed` questions out of the
    full `questions` pool (Questions Available), respecting Section
    Distribution and, if set, Difficulty Distribution. Called once per
    student attempt (NOT at quiz-creation time), so every student gets a
    different, independently-randomized combination from the same pool.
    Returns the drawn question list (each question keeps its original
    pool index under "_poolIndex" so answers can be graded against the
    original document later).
    """
    import random

    questions = list(enumerate(doc.get("questions") or []))
    displayed = doc.get("questionsDisplayed") or len(questions)
    section_config = doc.get("sectionConfig")
    section_dist = doc.get("sectionDistribution")
    difficulty_dist = doc.get("difficultyDistribution")

    def _tag(idx_q, key_fn):
        idx, q = idx_q
        out = dict(q)
        out["_poolIndex"] = idx
        return out

    pool = [_tag(iq, None) for iq in questions]

    if section_config:
        # Combined per-(section, difficulty) stratified draw — the exact
        # Hard/Medium/Easy Display counts configured for each section are
        # drawn from the matching slice of the pool, independently
        # randomized per student attempt.
        by_key = {}
        for q in pool:
            key = (q.get("section"), (q.get("difficulty") or "").strip().lower())
            by_key.setdefault(key, []).append(q)
        chosen = []
        for sect, cfg in section_config.items():
            for level, count_key in (("hard", "hardDisplay"), ("medium", "mediumDisplay"), ("easy", "easyDisplay")):
                count = cfg.get(count_key) or 0
                if count <= 0:
                    continue
                bucket = list(by_key.get((sect, level), []))
                random.shuffle(bucket)
                chosen.extend(bucket[:count])
        if len(chosen) < displayed:
            chosen_idx = {q["_poolIndex"] for q in chosen}
            remaining = [q for q in pool if q["_poolIndex"] not in chosen_idx]
            random.shuffle(remaining)
            chosen.extend(remaining[: displayed - len(chosen)])
        random.shuffle(chosen)
        return chosen[:displayed]

    if section_dist:
        by_section = {}
        for q in pool:
            by_section.setdefault(q.get("section"), []).append(q)
        chosen = []
        for sect, count in section_dist.items():
            bucket = by_section.get(sect, [])
            random.shuffle(bucket)
            chosen.extend(bucket[:count])
        # Any shortfall (shouldn't happen thanks to create-time validation)
        # is topped up from the remaining pool so the quiz never under-delivers.
        if len(chosen) < displayed:
            chosen_idx = {q["_poolIndex"] for q in chosen}
            remaining = [q for q in pool if q["_poolIndex"] not in chosen_idx]
            random.shuffle(remaining)
            chosen.extend(remaining[: displayed - len(chosen)])
        random.shuffle(chosen)
        return chosen[:displayed]

    if difficulty_dist:
        by_diff = {}
        for q in pool:
            by_diff.setdefault((q.get("difficulty") or "").strip().lower(), []).append(q)
        chosen = []
        for level, count in difficulty_dist.items():
            bucket = by_diff.get(level, [])
            random.shuffle(bucket)
            chosen.extend(bucket[:count])
        if len(chosen) < displayed:
            chosen_idx = {q["_poolIndex"] for q in chosen}
            remaining = [q for q in pool if q["_poolIndex"] not in chosen_idx]
            random.shuffle(remaining)
            chosen.extend(remaining[: displayed - len(chosen)])
        random.shuffle(chosen)
        return chosen[:displayed]

    random.shuffle(pool)
    return pool[:displayed]


# ------------------------------------------------------------
# Blueprint factory
# ------------------------------------------------------------
def init_quiz(db, scope):
    """scope: 'trainer' or 'super_admin'.

    Role hierarchy: Super Admin -> Trainer -> Student, with Trainer
    reporting directly to Super Admin (NOT under College Admin). Per
    project requirements, within the Assessment Management module the
    Trainer and Super Admin have IDENTICAL permissions — full platform
    visibility and control over every quiz, not just ones the trainer
    personally created. College Admin and Student have no access to
    this module at all (enforced by which roles are accepted by
    role_required below).
    """
    bp = Blueprint(f"quiz_{scope}", __name__)
    quizzes = db.quizzes
    quiz_sections = db.quiz_sections
    role = "trainer" if scope == "trainer" else "super_admin"

    def _visible_query(actor):
        # Both Super Admin and Trainer get full platform visibility —
        # Assessment Management grants identical access to both roles.
        return {}

    def _find_owned_or_404(quiz_id, actor, require_ownership=True):
        oid = to_object_id(quiz_id)
        if not oid:
            return None, error("Invalid quiz id.", 404)
        doc = quizzes.find_one({"_id": oid})
        if not doc:
            return None, error("Quiz not found.", 404)
        # NOTE: ownership is intentionally NOT enforced. Trainer has the
        # same full access as Super Admin within Assessment Management,
        # so a Trainer can manage (edit/publish/schedule/cancel/archive/
        # delete) any quiz, not only ones they personally created.
        return doc, None

    # ----------------------------------------------------
    # GET /assessment/sections — category/section picker, DB-backed
    # (seeded once so it is never a hardcoded frontend array).
    # ----------------------------------------------------
    def _ensure_sections():
        if quiz_sections.count_documents({}) == 0:
            quiz_sections.insert_many([{"name": n, "createdAt": now()} for n in _DEFAULT_QUIZ_SECTIONS])
        return list(quiz_sections.find({}).sort("name", 1))

    @bp.route("/assessment/sections", methods=["GET"])
    @role_required("trainer", "super_admin")
    def list_sections():
        docs = _ensure_sections()
        return ok({"sections": [{"id": str(d["_id"]), "name": d["name"]} for d in docs]})

    # ----------------------------------------------------
    # DOWNLOAD TEMPLATE (Part 7) — .xlsx with the exact columns the
    # bulk-upload-validate parser below expects, dropdown data
    # validation for Question Type/Section/Difficulty, sample rows
    # covering Easy/Medium/Hard, and a live list of the platform's
    # actual DB-backed sections (never hardcoded).
    # ----------------------------------------------------
    @bp.route("/quizzes/bulk-upload-template", methods=["GET"])
    @role_required(role)
    def bulk_upload_template():
        section_names = [d["name"] for d in _ensure_sections()]

        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"
        headers = [
            "Question Text", "Option A", "Option B", "Option C", "Option D",
            "Correct Answer(s)", "Question Type", "Section", "Difficulty", "Marks", "Explanation",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        sec1 = section_names[0] if section_names else "Communication"
        sec2 = section_names[1] if len(section_names) > 1 else sec1
        ws.append([
            "What does HTML stand for?", "Hyper Text Markup Language", "High Tech Modern Language",
            "Home Tool Markup Language", "Hyperlinks and Text Markup Language", "A", "Single Choice",
            sec1, "Easy", 1, "HTML stands for Hyper Text Markup Language.",
        ])
        ws.append([
            "Which of these are valid loop constructs? (select all that apply)", "for", "while",
            "repeat", "do-while", "A,B,D", "Multiple Choice", sec2, "Medium", 2,
            "for/while/do-while are valid loop keywords; 'repeat' is not.",
        ])
        ws.append([
            "Which sorting algorithm has the best average-case time complexity?", "Bubble Sort",
            "Quick Sort", "Selection Sort", "Insertion Sort", "B", "Single Choice", sec1, "Hard", 3, "",
        ])

        diff_dv = DataValidation(type="list", formula1='"Easy,Medium,Hard"', allow_blank=False,
                                  showErrorMessage=True, errorTitle="Invalid Difficulty",
                                  error="Choose Easy, Medium, or Hard.")
        ws.add_data_validation(diff_dv)
        diff_dv.add("I2:I1000")

        type_dv = DataValidation(type="list", formula1='"Single Choice,Multiple Choice"', allow_blank=False,
                                  showErrorMessage=True, errorTitle="Invalid Question Type",
                                  error="Choose Single Choice or Multiple Choice.")
        ws.add_data_validation(type_dv)
        type_dv.add("G2:G1000")

        if section_names:
            sec_list = ",".join(section_names)
            if len(sec_list) < 255:  # Excel's inline-list formula has a 255-char limit
                sec_dv = DataValidation(type="list", formula1=f'"{sec_list}"', allow_blank=False,
                                         showErrorMessage=True, errorTitle="Invalid Section",
                                         error="Choose one of the configured sections.")
                ws.add_data_validation(sec_dv)
                sec_dv.add("H2:H1000")

        for i, width in enumerate([44, 20, 20, 20, 20, 16, 16, 20, 12, 8, 34], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws2 = wb.create_sheet("Instructions")
        ws2["A1"] = "Create Quiz — Bulk Upload Template"
        ws2["A1"].font = Font(bold=True, size=14)
        instructions = [
            "",
            "Fill one row per question on the 'Questions' sheet. Do not remove, rename, or reorder the header row.",
            "Question Text — required.",
            "Option A / Option B — required. Option C / Option D — optional.",
            "Correct Answer(s) — the letter(s) of the correct option(s), e.g. 'A' or 'A,C' for a "
            "Multiple Choice question. Must point to a filled-in option.",
            "Question Type — 'Single Choice' (exactly one correct answer) or 'Multiple Choice' "
            "(one or more correct answers).",
            "Section — must exactly match one of this platform's configured sections: "
            + (", ".join(section_names) if section_names else "(no sections configured yet)"),
            "Difficulty — must be exactly 'Easy', 'Medium', or 'Hard'.",
            "Marks — a positive number.",
            "Explanation — optional, shown to students after the quiz closes.",
            "Duplicate questions (identical Question Text) are rejected.",
            "After uploading, click 'Validate Upload' to check the file — row-by-row errors are shown "
            "if anything needs fixing — before you can continue to Review & Publish.",
        ]
        for i, text in enumerate(instructions, start=2):
            ws2.cell(row=i, column=1, value=text)
        ws2.column_dimensions["A"].width = 110

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="quiz_bulk_upload_template.xlsx",
        )

    # ----------------------------------------------------
    # BULK UPLOAD VALIDATE (Part 6) — full server-side parse + validation
    # of an uploaded .xlsx. Nothing is trusted or parsed in the browser.
    # Returns every valid question (ready to slot straight into the same
    # question_pool the manual-entry flow produces) plus a row-numbered
    # error list for anything invalid, and a per-section Hard/Medium/Easy
    # summary of the valid rows so the wizard can auto-populate Available.
    # ----------------------------------------------------
    @bp.route("/quizzes/bulk-upload-validate", methods=["POST"])
    @role_required(role)
    def bulk_upload_validate():
        file = request.files.get("file")
        if not file or not file.filename:
            return error("Please choose a file to upload.", 400)

        filename = file.filename.lower()
        if filename.endswith(".xls"):
            return error(
                "Legacy .xls files aren't supported — download the template (which is .xlsx), "
                "fill it in, and re-upload.", 422,
            )
        if not filename.endswith((".xlsx", ".xlsm")):
            return error("Unsupported file type — please upload an .xlsx file.", 422)

        logger.info("bulk_upload_validate: start — file=%r actor_role=%s", file.filename, role)

        try:
            wb = load_workbook(BytesIO(file.read()), data_only=True)
        except Exception:
            logger.exception("bulk_upload_validate: workbook could not be opened (file=%r)", file.filename)
            return error("Could not read this file — make sure it's a valid, uncorrupted .xlsx file.", 422)

        logger.info("bulk_upload_validate: workbook opened — sheets=%s", wb.sheetnames)

        valid_sections = {d["name"].strip().lower(): d["name"] for d in _ensure_sections()}
        letter_map = {"A": 0, "B": 1, "C": 2, "D": 3}

        def _col(header, *names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        # -----------------------------------------------------------
        # PART 1 FIX — read EVERY worksheet in the workbook, not just a
        # sheet literally named "Questions" or whichever sheet happens to
        # be first. This is the root cause of "only ~2 questions detected"
        # on a multi-section workbook: if a trainer's file has one sheet
        # per section (e.g. "Section 1" .. "Section 5", exactly like the
        # sibling parser in quiz_common.py's parse_master_workbook does
        # for the placement-exam question bank), the old code silently
        # looked at a single sheet and ignored the rest.
        #
        # Two supported layouts, auto-detected per sheet (never assumed
        # for the whole workbook, since a trainer could mix both):
        #   (a) One "flat" sheet with a Section column — every row can
        #       belong to a different section (this is what the
        #       Download Template produces).
        #   (b) One sheet per section, sheet TITLE = section name, no
        #       Section column required on that sheet (matches the
        #       platform's other bulk-import convention).
        # A sheet that has neither a recognizable header row nor a
        # matching section name (e.g. the template's "Instructions"
        # sheet) is skipped — logged, not treated as an error — so it
        # never blocks the rest of the workbook from being read.
        # -----------------------------------------------------------
        questions = []
        row_errors = []
        seen_texts = {}          # normalized question text -> "Sheet!Row" first seen at
        section_summary = {}
        total_rows = 0
        sheets_parsed = []
        sheets_skipped = []

        for ws in wb.worksheets:
            sheet_name = ws.title
            logger.info("bulk_upload_validate: scanning sheet %r", sheet_name)
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                logger.info("bulk_upload_validate: sheet %r is empty — skipped", sheet_name)
                sheets_skipped.append(sheet_name)
                continue

            header = [str(h or "").strip().lower() for h in rows[0]]
            col = {
                "text": _col(header, "question text", "question"),
                "a": _col(header, "option a"), "b": _col(header, "option b"),
                "c": _col(header, "option c"), "d": _col(header, "option d"),
                "correct": _col(header, "correct answer(s)", "correct answers", "correct answer", "correct"),
                "type": _col(header, "question type", "type"),
                "section": _col(header, "section"),
                "difficulty": _col(header, "difficulty", "difficulty level"),
                "marks": _col(header, "marks"),
                "explanation": _col(header, "explanation"),
            }

            # A sheet needs at minimum Question Text + Option A + Option B
            # + Correct Answer(s) + Difficulty to be parseable at all.
            # Section is the one column that MAY be implied by the sheet
            # title instead (layout (b) above).
            core_required = {
                "Question Text": col["text"], "Option A": col["a"], "Option B": col["b"],
                "Correct Answer(s)": col["correct"], "Difficulty": col["difficulty"],
            }
            missing_core = [name for name, idx in core_required.items() if idx is None]
            if missing_core:
                logger.info(
                    "bulk_upload_validate: sheet %r has no recognizable question header "
                    "(missing %s) — treated as a non-data sheet and skipped.",
                    sheet_name, missing_core,
                )
                sheets_skipped.append(sheet_name)
                continue

            # Resolve this sheet's section handling.
            sheet_section_name = None
            if col["section"] is None:
                # Layout (b): no Section column on this sheet — fall back
                # to the sheet title itself, same convention already used
                # by parse_master_workbook() in quiz_common.py.
                title_key = sheet_name.strip().lower()
                if title_key in valid_sections:
                    sheet_section_name = valid_sections[title_key]
                    logger.info(
                        "bulk_upload_validate: sheet %r has no Section column — "
                        "using sheet title as section (%s).", sheet_name, sheet_section_name,
                    )
                else:
                    row_errors.append({
                        "row": 1, "sheet": sheet_name,
                        "issue": (
                            f'Sheet "{sheet_name}" has no Section column and its title does not match '
                            f'a configured section ({", ".join(valid_sections.values())}). '
                            f"Add a Section column, or rename this sheet to match a configured section."
                        ),
                    })
                    sheets_skipped.append(sheet_name)
                    continue

            sheets_parsed.append(sheet_name)

            def cell(row, key, _col=col):
                idx = _col[key]
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            for i, row in enumerate(rows[1:], start=2):
                # PART 1 FIX: only a genuinely, entirely empty row is
                # skipped. Every other row — however partially filled or
                # invalid — is counted and validated below, and a bad
                # row never stops the loop from reaching the rest of the
                # sheet or the sheets that follow it.
                if row is None or all(v in (None, "") for v in row):
                    logger.debug("bulk_upload_validate: sheet %r row %s — fully empty, skipped.", sheet_name, i)
                    continue
                total_rows += 1
                issues = []

                text = str(cell(row, "text") or "").strip()
                options = [
                    str(cell(row, "a") or "").strip(), str(cell(row, "b") or "").strip(),
                    str(cell(row, "c") or "").strip(), str(cell(row, "d") or "").strip(),
                ]
                correct_raw = str(cell(row, "correct") or "").strip()
                type_raw = str(cell(row, "type") or "").strip().lower()
                section_raw = str(cell(row, "section") or "").strip() if col["section"] is not None else ""
                difficulty_raw = str(cell(row, "difficulty") or "").strip()
                marks_raw = cell(row, "marks")
                explanation = str(cell(row, "explanation") or "").strip()

                if not text:
                    issues.append("Question Text is required.")
                filled_options = [o for o in options if o]
                if len(filled_options) < 2:
                    issues.append("At least 2 answer options (A and B) are required.")

                q_type = "multiple_choice" if "multi" in type_raw else "single_choice"

                correct_indices = []
                if not correct_raw:
                    issues.append("Correct Answer(s) is required.")
                else:
                    for token in correct_raw.replace(";", ",").replace("/", ",").split(","):
                        token = token.strip().upper()
                        if not token:
                            continue
                        if token not in letter_map:
                            issues.append(f'Correct Answer "{token}" is invalid — use A, B, C, or D.')
                            continue
                        idx = letter_map[token]
                        if idx >= len(options) or not options[idx]:
                            issues.append(f'Correct Answer "{token}" references an empty option.')
                            continue
                        correct_indices.append(idx)
                    correct_indices = sorted(set(correct_indices))
                    if correct_raw and not correct_indices and not any("Correct Answer" in x for x in issues):
                        issues.append("No valid Correct Answer(s) found.")
                    if q_type == "single_choice" and len(correct_indices) > 1:
                        issues.append("Single Choice questions must have exactly one correct answer.")

                # Section: either read from this row (layout a) or
                # inherited from the sheet title, resolved above (layout b).
                if sheet_section_name is not None:
                    resolved_section = sheet_section_name
                elif not section_raw:
                    issues.append("Section is required.")
                    resolved_section = None
                elif section_raw.lower() not in valid_sections:
                    issues.append(f'Unknown section "{section_raw}" — must exactly match a configured section.')
                    resolved_section = None
                else:
                    resolved_section = valid_sections[section_raw.lower()]

                difficulty_clean = None
                if not difficulty_raw:
                    issues.append("Difficulty is required.")
                elif difficulty_raw.lower() not in _VALID_DIFFICULTIES:
                    issues.append(f'Difficulty must be Easy, Medium, or Hard (got "{difficulty_raw}").')
                else:
                    difficulty_clean = difficulty_raw.capitalize()

                marks = None
                try:
                    marks = float(marks_raw) if marks_raw not in (None, "") else 1.0
                    if marks <= 0:
                        issues.append("Marks must be a positive number.")
                except (TypeError, ValueError):
                    issues.append("Marks must be a positive number.")

                norm_text = f"{(resolved_section or '').lower()}|{text.strip().lower()}"
                if text:
                    if norm_text in seen_texts:
                        issues.append(f"Duplicate question — identical to {seen_texts[norm_text]}.")
                    else:
                        seen_texts[norm_text] = f'Sheet "{sheet_name}", Row {i}'

                logger.debug(
                    "bulk_upload_validate: sheet=%r row=%s section=%r question=%.60r issues=%s",
                    sheet_name, i, resolved_section, text, issues,
                )

                if issues:
                    row_errors.append({"row": i, "sheet": sheet_name, "issue": " ".join(issues)})
                    continue

                questions.append({
                    "text": text, "options": options, "correct": correct_indices, "type": q_type,
                    "section": resolved_section, "difficulty": difficulty_clean, "marks": marks,
                    "explanation": explanation,
                })
                bucket = section_summary.setdefault(resolved_section, {"hard": 0, "medium": 0, "easy": 0})
                bucket[difficulty_clean.lower()] += 1

        if not sheets_parsed:
            logger.warning(
                "bulk_upload_validate: no parseable sheet found — sheets in file=%s, skipped=%s",
                wb.sheetnames, sheets_skipped,
            )
            return error(
                "No question data could be found in this workbook. Make sure at least one sheet has "
                "the required header row (Question Text, Option A, Option B, Correct Answer(s), "
                "Difficulty), or is named after a configured section. Please use Download Template.", 422,
            )

        logger.info(
            "bulk_upload_validate: done — sheets_parsed=%s sheets_skipped=%s total_rows=%s valid=%s invalid=%s",
            sheets_parsed, sheets_skipped, total_rows, len(questions), len(row_errors),
        )

        return ok({
            "questions": questions,
            "errors": row_errors,
            "sectionSummary": section_summary,
            "totalRows": total_rows,
            "sheetsParsed": sheets_parsed,
            "sheetsSkipped": sheets_skipped,
        })

    # ----------------------------------------------------
    # Random-draw preview (Part 9) — exercises the same stratified
    # selection a student attempt will use, without exposing correct
    # answers to the eventual student flow. Useful here for trainers/
    # admins to sanity-check their Section/Difficulty Distribution.
    # A student-facing attempt endpoint (returning this shape with
    # correct answers stripped) belongs in student.py and can reuse
    # quiz_module.select_random_questions() directly.
    # ----------------------------------------------------
    @bp.route("/quizzes/<quiz_id>/random-preview", methods=["GET"])
    @role_required(role)
    def random_preview(quiz_id):
        actor = _actor(db)
        doc, err = _find_owned_or_404(quiz_id, actor, require_ownership=False)
        if err:
            return err
        drawn = select_random_questions(doc)
        return ok({
            "questionsDisplayed": doc.get("questionsDisplayed"),
            "questionsAvailable": doc.get("questionsAvailable"),
            "drawn": [{"text": q["text"], "section": q.get("section"), "difficulty": q.get("difficulty")} for q in drawn],
        })

    def _validation_error_response(errors):
        """Part 3: 'Never silently fail. Return meaningful backend
        validation errors.' Returns BOTH a single human-readable message
        (backward compatible with any caller that only reads .message)
        AND the full, un-truncated list of every error found, so the
        frontend can show every problem at once instead of forcing the
        admin to fix-and-resubmit one error at a time."""
        summary = errors[0] if len(errors) == 1 else (
            "; ".join(errors[:5]) + (f" (+{len(errors) - 5} more)" if len(errors) > 5 else "")
        )
        resp, status = error(summary, 422)
        body = resp.get_json()
        body["errors"] = errors
        return jsonify(body), status

    def _finalize_question_bank_pool(normalized):
        """Part 7 — called only when publishing a Question-Bank-sourced
        quiz. Draws a fresh, independent random set from db.question_bank
        for this specific quiz right now (never at draft-save time — see
        the comment in validate_and_normalize). Re-checks the draw against
        what was required per cell as a safety net against a race (another
        admin publishing off the same bank between this request's
        validation and this exact moment) — if the bank came up short,
        the quiz is NOT saved and a descriptive error is returned,
        mirroring Part 8's example wording.

        Returns (questions, error_response_or_None).
        """
        section_config = normalized.get("sectionConfig") or {}
        drawn = draw_questions_from_bank(db, section_config)
        drawn_counts = {}
        for q in drawn:
            key = (q["section"], q["difficulty"].lower())
            drawn_counts[key] = drawn_counts.get(key, 0) + 1
        shortfalls = []
        for sect, cfg in section_config.items():
            for level, count_key in (("hard", "hardDisplay"), ("medium", "mediumDisplay"), ("easy", "easyDisplay")):
                required = cfg.get(count_key) or 0
                if required <= 0:
                    continue
                got = drawn_counts.get((sect, level), 0)
                if got < required:
                    shortfalls.append(
                        f'Only {got} {level.capitalize()} question(s) available in "{sect}" at publish time. '
                        f"Requested: {required}. Please upload more {level.capitalize()} questions."
                    )
        if shortfalls:
            logger.warning("question_bank publish aborted — shortfalls: %s", shortfalls)
            return None, _validation_error_response(shortfalls)
        return drawn, None

    # ----------------------------------------------------
    # CREATE
    # ----------------------------------------------------
    @bp.route("/quizzes", methods=["POST"])
    @role_required(role)
    def create_quiz():
        data = request.get_json(silent=True) or {}
        actor = _actor(db)
        normalized, errors = validate_and_normalize(data, db, actor)
        if errors:
            logger.info("create_quiz: validation failed for actor=%s errors=%s", actor.get("id"), errors)
            return _validation_error_response(errors)

        requested_state = str(data.get("status") or data.get("state") or "published").strip().lower()
        state = "draft" if requested_state in ("draft", "save_draft") else "published"

        if normalized["quizType"] == "question_bank" and state == "published":
            drawn, err_resp = _finalize_question_bank_pool(normalized)
            if err_resp:
                return err_resp
            normalized["questions"] = drawn
            normalized["questionsAvailable"] = len(drawn)

        doc = {
            **normalized,
            "state": state,
            "cancelled": False,
            "archived": False,
            "createdBy": actor,
            "createdOn": now(),
            "updatedAt": now(),
            "updatedBy": actor,
        }
        result = quizzes.insert_one(doc)
        doc["_id"] = result.inserted_id
        log_activity(
            db, actor["id"], actor["role"], "quiz_created" if state == "published" else "quiz_draft_saved",
            f'Created quiz "{doc["title"]}"' + (" (draft)" if state == "draft" else ""),
            college=actor.get("college"), meta={"quizId": str(result.inserted_id)},
        )

        # Part 5 — every validated Trainer-authored question (Manual Entry
        # or Bulk Upload) is permanently mirrored into the QuestionBank.
        # Super Admin's own Manual/Bulk quizzes do the same (the bank is
        # platform-wide, not Trainer-only in practice, since Trainer and
        # Super Admin have identical permissions in this module — see the
        # module docstring) — only Question-Bank-*sourced* quizzes are
        # skipped, since those questions are already in the bank.
        if normalized["quizType"] in ("manual", "bulk"):
            sync_questions_to_bank(db, actor, normalized["questions"], source_quiz_title=doc["title"], source_quiz_id=doc["_id"])

        return ok({"quiz": serialize_quiz(doc)}, message="Quiz created successfully.", status=201)

    # ----------------------------------------------------
    # DRAFT AUTOSAVE — upserts one draft per wizard session. If "id" is
    # supplied (a previously-returned draft id) it updates that same
    # document; otherwise a new draft is created and its id returned so
    # subsequent autosaves target the same record instead of piling up
    # duplicate drafts.
    # ----------------------------------------------------
    @bp.route("/quizzes/draft", methods=["PUT"])
    @role_required(role)
    def save_draft():
        data = request.get_json(silent=True) or {}
        actor = _actor(db)
        draft_id = data.get("id") or data.get("draftId")

        # Drafts are exempt from full validation (they're incomplete by
        # definition) — only require *something* worth saving.
        title = str(_first(data, "title", "assessment_name", "name", default="")).strip()

        colleges = _first(data, "colleges", "applicable_colleges", default=[]) or []
        payload = {
            "title": title,
            "description": str(_first(data, "description", "instructions", default="")).strip(),
            "category": str(_first(data, "category", "assessment_type", "quizType", "type", default="")).strip(),
            "cohortTarget": str(_first(data, "cohortTarget", "applicable_cohort", "cohort", default="all")).strip(),
            "colleges": [str(c) for c in colleges],
            "collegeNames": _resolve_college_names(db, [str(c) for c in colleges]),
            "startDateTime": parse_dt(_first(data, "startDateTime", "start_datetime", "start")),
            "endDateTime": parse_dt(_first(data, "endDateTime", "end_datetime", "end")),
            "durationMinutes": _first(data, "durationMinutes", "duration", default=None),
            "visibility": _first(data, "visibility", default=("colleges" if colleges else "all")),
            "questionsAvailable": _first(data, "questionsAvailable", "questions_available", default=None),
            "questionsDisplayed": _first(data, "questionsDisplayed", "questions_displayed", default=None),
            "difficultyDistribution": _first(data, "difficultyDistribution", "difficulty_distribution", default=None),
            "sectionDistribution": _first(data, "sectionDistribution", "section_distribution", default=None),
            "sectionConfig": data.get("sectionConfig") or {},
            "quizType": str(_first(data, "quizType", "assessment_type", default="manual")).strip().lower() or "manual",
            "questions": data.get("questions") or data.get("question_pool") or [],
            "state": "draft",
            "updatedAt": now(),
            "updatedBy": actor,
        }

        oid = to_object_id(draft_id) if draft_id else None
        if oid:
            existing = quizzes.find_one({"_id": oid})
            # Ownership is not restricted — Trainer has the same full
            # access as Super Admin within Assessment Management.
            if not existing:
                oid = None  # fall through to creating a fresh draft
        if oid:
            quizzes.update_one({"_id": oid}, {"$set": payload})
            saved_id = oid
        else:
            payload["createdBy"] = actor
            payload["createdOn"] = now()
            saved_id = quizzes.insert_one(payload).inserted_id

        return ok({"id": str(saved_id)}, message="Draft saved.")

    # ----------------------------------------------------
    # LIST
    # ----------------------------------------------------
    @bp.route("/quizzes", methods=["GET"])
    @role_required("trainer", "super_admin")
    def list_quizzes():
        actor = _actor(db)
        query = _visible_query(actor)

        # Cohort filter — "entry_level" and "all" mean exactly what they
        # do everywhere else in this platform; anything else must match
        # the quiz's own cohortTarget exactly. Resolved as a real Mongo
        # query clause, not a client-side array filter.
        cohort = (request.args.get("cohort") or "all").strip()
        if cohort and cohort != "all":
            query["cohortTarget"] = cohort

        # Free-text search — quiz title, case-insensitive partial match.
        search = (request.args.get("search") or "").strip()
        if search:
            query["title"] = {"$regex": re.escape(search), "$options": "i"}

        cursor = quizzes.find(query).sort("createdOn", -1)
        docs = [normalize_quiz_college_names(db, d) for d in cursor]

        # Status is a computed value (draft/scheduled/active/completed/
        # cancelled/archived derived from state + dates by compute_status),
        # not a stored field — so this filter is applied here, in the
        # backend, right before serializing, rather than as a raw Mongo
        # match. Still fully server-side: the browser only ever receives
        # rows that already satisfy every requested filter.
        status = (request.args.get("status") or "all").strip().lower()
        if status and status != "all":
            if status == "published":
                docs = [d for d in docs if compute_status(d) in ("scheduled", "active")]
            else:
                docs = [d for d in docs if compute_status(d) == status]

        return ok({"quizzes": [serialize_quiz(d) for d in docs]})

    # ----------------------------------------------------
    # GET ONE
    # ----------------------------------------------------
    @bp.route("/quizzes/<quiz_id>", methods=["GET"])
    @role_required("trainer", "super_admin")
    def get_quiz(quiz_id):
        actor = _actor(db)
        doc, err = _find_owned_or_404(quiz_id, actor, require_ownership=False)
        if err:
            return err
        return ok({"quiz": serialize_quiz(normalize_quiz_college_names(db, doc))})

    # ----------------------------------------------------
    # UPDATE — full edit, blocked once the quiz has started
    # ----------------------------------------------------
    @bp.route("/quizzes/<quiz_id>", methods=["PUT"])
    @role_required(role)
    def update_quiz(quiz_id):
        actor = _actor(db)
        doc, err = _find_owned_or_404(quiz_id, actor)
        if err:
            return err
        if not editable_now(doc):
            return error(edit_block_message(doc), 409)

        data = request.get_json(silent=True) or {}
        normalized, errors = validate_and_normalize(data, db, actor, existing_id=doc["_id"])
        if errors:
            logger.info("update_quiz: validation failed for quiz=%s errors=%s", quiz_id, errors)
            return _validation_error_response(errors)

        requested_state = str(data.get("status") or data.get("state") or doc.get("state") or "published").strip().lower()
        state = "draft" if requested_state in ("draft", "save_draft") else "published"

        if normalized["quizType"] == "question_bank" and state == "published":
            drawn, err_resp = _finalize_question_bank_pool(normalized)
            if err_resp:
                return err_resp
            normalized["questions"] = drawn
            normalized["questionsAvailable"] = len(drawn)

        update = {**normalized, "state": state, "updatedAt": now(), "updatedBy": actor}
        quizzes.update_one({"_id": doc["_id"]}, {"$set": update})
        updated = quizzes.find_one({"_id": doc["_id"]})
        log_activity(
            db, actor["id"], actor["role"], "quiz_updated",
            f'Updated quiz "{updated["title"]}"',
            college=actor.get("college"), meta={"quizId": str(doc["_id"])},
        )
        if normalized["quizType"] in ("manual", "bulk"):
            sync_questions_to_bank(db, actor, normalized["questions"], source_quiz_title=updated["title"], source_quiz_id=doc["_id"])
        return ok({"quiz": serialize_quiz(updated)}, message="Quiz updated successfully.")

    # ----------------------------------------------------
    # PUBLISH / UNPUBLISH / CANCEL / ARCHIVE — one endpoint, branching on
    # the requested target status, each with its own eligibility rule.
    # ----------------------------------------------------
    @bp.route("/quizzes/<quiz_id>/status", methods=["PATCH"])
    @role_required(role)
    def toggle_status(quiz_id):
        actor = _actor(db)
        doc, err = _find_owned_or_404(quiz_id, actor)
        if err:
            return err

        data = request.get_json(silent=True) or {}
        requested = str(data.get("status") or data.get("state") or "").strip().lower()
        if requested not in ("draft", "published", "active", "scheduled", "cancelled", "archived"):
            return error("status must be one of: draft, published, cancelled, archived.")

        if requested == "cancelled":
            if not cancellable_now(doc):
                return error("Only a Draft or Scheduled quiz can be cancelled.", 409)
            update = {"cancelled": True, "updatedAt": now(), "updatedBy": actor}
            action, verb = "quiz_cancelled", "cancelled"
        elif requested == "archived":
            if not archivable_now(doc):
                return error("Only a Completed or Cancelled quiz can be archived.", 409)
            update = {"archived": True, "updatedAt": now(), "updatedBy": actor}
            action, verb = "quiz_archived", "archived"
        else:
            if not editable_now(doc):
                return error(edit_block_message(doc), 409)
            new_state = "published" if requested in ("published", "active", "scheduled") else "draft"
            update = {"state": new_state, "updatedAt": now(), "updatedBy": actor}
            action = "quiz_published" if new_state == "published" else "quiz_unpublished"
            verb = "published" if new_state == "published" else "moved to draft"

            # Part 7 — a Draft created in Question Bank mode hasn't drawn
            # its questions yet (see validate_and_normalize); publishing it
            # from the quiz list (rather than the wizard's own Publish
            # button) still needs to trigger that draw, with the same
            # pre-publish shortfall check as everywhere else.
            if new_state == "published" and doc.get("quizType") == "question_bank":
                drawn, err_resp = _finalize_question_bank_pool(doc)
                if err_resp:
                    return err_resp
                update["questions"] = drawn
                update["questionsAvailable"] = len(drawn)

        quizzes.update_one({"_id": doc["_id"]}, {"$set": update})
        updated = quizzes.find_one({"_id": doc["_id"]})
        log_activity(
            db, actor["id"], actor["role"], action, f'Quiz "{updated["title"]}" {verb}',
            college=actor.get("college"), meta={"quizId": str(doc["_id"])},
        )
        return ok({"quiz": serialize_quiz(updated)}, message=f"Quiz {verb}.")

    # ----------------------------------------------------
    # DUPLICATE — clones a quiz as a fresh Draft (new dates required before
    # it can be published again), available once the source is Completed.
    # ----------------------------------------------------
    @bp.route("/quizzes/<quiz_id>/duplicate", methods=["POST"])
    @role_required(role)
    def duplicate_quiz(quiz_id):
        actor = _actor(db)
        doc, err = _find_owned_or_404(quiz_id, actor, require_ownership=False)
        if err:
            return err
        doc = normalize_quiz_college_names(db, doc)
        clone = {k: v for k, v in doc.items() if k != "_id"}
        clone.update({
            "title": f'{doc.get("title","Quiz")} (Copy)',
            "state": "draft", "cancelled": False, "archived": False,
            "createdBy": actor, "createdOn": now(), "updatedAt": now(), "updatedBy": actor,
        })
        new_id = quizzes.insert_one(clone).inserted_id
        clone["_id"] = new_id
        log_activity(
            db, actor["id"], actor["role"], "quiz_duplicated",
            f'Duplicated quiz "{doc.get("title")}"', college=actor.get("college"),
            meta={"sourceQuizId": str(doc["_id"]), "quizId": str(new_id)},
        )
        return ok({"quiz": serialize_quiz(clone)}, message="Quiz duplicated as a new draft.", status=201)

    # ----------------------------------------------------
    # DELETE — only while still a Draft (Scheduled quizzes must be
    # Cancelled instead, never deleted outright)
    # ----------------------------------------------------
    @bp.route("/quizzes/<quiz_id>", methods=["DELETE"])
    @role_required(role)
    def delete_quiz(quiz_id):
        actor = _actor(db)
        doc, err = _find_owned_or_404(quiz_id, actor)
        if err:
            return err
        if not deletable_now(doc):
            return error("Only a Draft quiz can be deleted. Scheduled quizzes must be cancelled instead.", 409)
        quizzes.delete_one({"_id": doc["_id"]})
        log_activity(
            db, actor["id"], actor["role"], "quiz_deleted",
            f'Deleted quiz "{doc.get("title")}"',
            college=actor.get("college"), meta={"quizId": str(doc["_id"])},
        )
        return ok(message="Quiz deleted successfully.")

    return bp
