"""
============================================================
 quiz_common.py — Shared Quiz Engine
============================================================
Everything that is used by MORE THAN ONE role (super_admin,
trainer, student) lives here so the three role modules
(superadmin.py, trainer.py, student.py) never duplicate logic
and can never drift out of sync with each other. This is the
single source of truth for:

  - Cohort rules (Entry Level -> A/B/C), used identically by
    the Super Admin Dashboard and the Trainer Dashboard.
  - Master Excel workbook parsing -> question bank persistence.
  - The random question engine (per-section random sampling).
  - Score calculation (per-section + overall).
  - Small JSON helpers / auth decorator, mirroring login.py's
    style so all modules feel like one codebase.

Collections used (created lazily by Mongo on first insert):
  - db.questions            one question bank, shared across cohorts
  - db.assessments          assessment definitions (admin/trainer created)
  - db.assessment_attempts  one doc per student attempt (in-progress or submitted)
  - db.activity_log         one doc per logged action, powers every
                            "Recent Activity" panel (student/trainer/admin).
                            See log_activity() / get_recent_activity() below.
============================================================
"""

import os
import random
import re
from datetime import datetime, timezone, timedelta

from bson import ObjectId
from bson.errors import InvalidId
from flask import jsonify
from flask_jwt_extended import jwt_required, get_jwt, get_jwt_identity
from pymongo import UpdateOne

# ------------------------------------------------------------
# Cohort configuration
# ------------------------------------------------------------
# A student with no cohort assigned yet is "Entry Level". Cohort
# generation is a TWO-STAGE process (see check_and_generate_cohort
# below): a real cohort (A/B/C) is only assigned once BOTH the
# assessment score AND the manual interview score exist, combined
# into a Final Employability Score using the Placement Rules that
# the Super Admin configures in db.placement_rules. There are NO
# hardcoded score ranges or weights anywhere in this file — every
# threshold is read from the database, every time, so a Super Admin
# edit takes effect immediately with zero code changes.
VALID_COHORTS = {"A", "B", "C"}
VALID_COHORT_TARGETS = {"A", "B", "C", "all", "entry_level"}
ENTRY_LEVEL = "entry_level"

# Fallback seed ONLY used the very first time the platform boots and
# db.placement_rules is completely empty, so the system has something
# sane to operate with before a Super Admin configures real rules.
# This is a one-time DB seed, not a code-level threshold: once seeded,
# every subsequent read/write goes through db.placement_rules and this
# constant is never consulted again.
#
# Three-threshold model (Super Admin > Placement Rules):
#   placementReadyThreshold — Marks >= this  -> Cohort A (Placement Ready)
#   nearReadyThreshold      — Marks >= this  -> Cohort B (Near Ready)
#   highRiskThreshold       — Marks <  this  -> Cohort C (High Risk)
# highRiskThreshold is kept equal to nearReadyThreshold (it's the same
# boundary described from the other side — "Near Ready starts here" /
# "High Risk ends here") but is stored and edited as its own field so
# the Super Admin UI can show/edit it directly, per spec.
# assessmentWeight / interviewWeight are plain percentages (0-100) that
# must sum to exactly 100; compute_final_employability_score() divides
# by 100 itself.
_DEFAULT_PLACEMENT_RULES_SEED = {
    "placementReadyThreshold": 75,
    "nearReadyThreshold": 50,
    "highRiskThreshold": 50,
    "assessmentWeight": 60,
    "interviewWeight": 40,
}

SECTION_ALIASES = {
    # sheet-name normalisation so "communication", "Communication ",
    # "COMMUNICATION" etc. all map to one canonical section key.
}


def now():
    return datetime.now(timezone.utc)


def iso_utc(dt):
    """Same fix as serialize()'s datetime branch, exposed standalone for
    the handful of call sites that build their response dict by hand
    instead of going through serialize() — student.py's quiz-attempt
    timestamps (startedAt/submittedAt), cohort timestamps, and the other
    role modules' equivalents. Mongo always hands back naive datetimes
    even though every write path here stores UTC; treating a naive value
    as UTC before formatting (instead of leaving it ambiguous, which
    browsers then parse as local time) is what actually fixes displayed
    times being off by the viewer's UTC offset."""
    if not dt:
        return None
    return (dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)).isoformat()


# India Standard Time — Asia/Kolkata is a fixed UTC+5:30 (no DST), so an
# explicit fixed offset is exact and never depends on the host OS tzdata.
IST_TZ = timezone(timedelta(hours=5, minutes=30))


def ist_from_utc(dt):
    """Convert a stored UTC datetime (aware, or naive-as-UTC as PyMongo
    hands back) into IST. The single place timestamps are shifted to
    Asia/Kolkata — the frontend never does its own hour arithmetic, so a
    timestamp is never converted twice."""
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST_TZ)


def fmt_ist(dt, fmt="%d/%m/%Y %H:%M:%S"):
    """IST display string for a stored UTC datetime, or None. Default
    format is the platform-wide Super Admin standard: 24-hour railway
    format, DD/MM/YYYY HH:mm:ss (e.g. 11/08/2026 14:35:27) — never
    AM/PM, never UTC. Every "Generated On" timestamp in the Super
    Admin reports/PDFs/marksheets, and every *_AtIST field returned to
    the frontend, funnels through this one function so the format can
    never drift between call sites."""
    ist = ist_from_utc(dt)
    return ist.strftime(fmt) if ist else None


def error(message, status=400):
    return jsonify(success=False, message=message), status


def ok(payload=None, message=None, status=200):
    body = {"success": True}
    if message:
        body["message"] = message
    if payload:
        body.update(payload)
    return jsonify(body), status


def role_required(*allowed_roles):
    """Same contract as login.py's role_required, duplicated locally so
    quiz_common has no import-time dependency on login.py."""
    def decorator(fn):
        @jwt_required()
        def wrapper(*args, **kwargs):
            claims = get_jwt()
            if claims.get("role") not in allowed_roles:
                return error("You are not authorized to perform this action.", 403)
            return fn(*args, **kwargs)
        wrapper.__name__ = fn.__name__
        return wrapper
    return decorator


def to_object_id(id_str):
    try:
        return ObjectId(id_str)
    except (InvalidId, TypeError):
        return None


def serialize(doc, extra_id_fields=None):
    """Turn a Mongo doc into JSON-safe dict: _id -> id, datetimes -> isoformat.

    Timezone fix (Student Dashboard assessment timing bug): pymongo always
    round-trips datetimes as NAIVE (Mongo has no tz concept), even though
    every write path in this codebase inserts timezone-AWARE UTC values
    (see parse_dt() in quiz_module.py). A naive datetime's .isoformat()
    has no 'Z'/offset suffix — e.g. "2026-08-01T09:30:00" — and browsers
    parse a timezone-less ISO datetime-time string as LOCAL time, not
    UTC. That silently re-interpreted every stored UTC wall-clock time as
    if it were already in the viewer's timezone, which is exactly why the
    Start/End *time* (but usually not the date, unless the shift crossed
    midnight) came out wrong everywhere a quiz's startDateTime/endDateTime
    got serialized this way. Re-attaching UTC before formatting makes the
    output an unambiguous, correctly-offset ISO string for every caller —
    Student, Trainer and Super Admin dashboards alike, since they all
    funnel through this one function."""
    if doc is None:
        return None
    out = {}
    for k, v in doc.items():
        if k == "_id":
            out["id"] = str(v)
        elif isinstance(v, datetime):
            out[k] = (v if v.tzinfo else v.replace(tzinfo=timezone.utc)).isoformat()
        elif isinstance(v, ObjectId):
            out[k] = str(v)
        else:
            out[k] = v
    return out


# ------------------------------------------------------------
# Cohort helpers — SHARED between Super Admin and Trainer dashboards.
# Both dashboards must call these functions rather than reimplementing
# the rule, so behaviour can never diverge between the two.
# ------------------------------------------------------------
def student_cohort_label(user_doc):
    """Returns 'A' / 'B' / 'C' / 'entry_level' for a student user doc."""
    cohort = (user_doc or {}).get("cohort")
    return cohort if cohort in VALID_COHORTS else ENTRY_LEVEL


# ------------------------------------------------------------
# Attendance — SINGLE SOURCE OF TRUTH for per-student attendance.
# The Super Admin Mark/View screens write db.attendance (one doc per
# student per class date, status present/absent, markedBy = the admin).
# The percentage is NEVER stored anywhere: every consumer (Student
# Dashboard, College Admin student list/profile, Super Admin records)
# derives it live from this helper so all three can never drift apart.
# ------------------------------------------------------------
def attendance_summary(db, student_id, include_history=True):
    """Attendance for ONE student, computed live from db.attendance.

    Returns:
        { total, present, absent, percentage, lastUpdated, history,
          thisMonth, monthly }
    total = present + absent for THIS student only (a student with no
    records gets honest all-zero stats rather than inheriting someone
    else's numbers). history is newest-first and only built when
    include_history is True (college list view just needs the counts).
    thisMonth holds the current calendar month's present/absent/total
    and monthly holds a chronological month-wise present/absent
    breakdown (both derived live from each record's stored date, so the
    Monthly Attendance Graph and This-Month stats can never drift from
    the same attendance records the totals come from)."""
    student_oid = student_id if isinstance(student_id, ObjectId) else to_object_id(student_id)
    if not student_oid:
        return {"total": 0, "present": 0, "absent": 0, "percentage": 0,
                "lastUpdated": None, "history": [], "thisMonth": {}, "monthly": []}
    records = list(db.attendance.find({"studentId": student_oid}))
    present = sum(1 for r in records if r.get("status") == "present")
    absent = sum(1 for r in records if r.get("status") == "absent")
    total = present + absent
    percentage = round(present / total * 100, 2) if total else 0
    last_ts = None
    for r in records:
        ts = r.get("updatedAt") or r.get("markedAt")
        if ts and (last_ts is None or ts > last_ts):
            last_ts = ts
    history = []
    if include_history:
        history = [{
            "date": r.get("date"),
            "status": r.get("status"),
            "markedAt": iso_utc(r.get("markedAt")),
            "markedBy": (r.get("markedBy") or {}).get("name"),
        } for r in records]
        history.sort(key=lambda h: (h.get("date") or ""), reverse=True)

    # Month-wise breakdown — one bucket per YYYY-MM of a record's date
    # (stored as YYYY-MM-DD). Kept for every month that actually has
    # records, oldest first, so the graph is always purely DB-driven.
    this_month_key = now().strftime("%Y-%m")
    monthly = {}
    this_present = this_absent = 0
    for r in records:
        d = str(r.get("date") or "")
        mk = d[:7] if len(d) >= 7 else ""
        if not mk:
            continue
        bucket = monthly.setdefault(mk, {"present": 0, "absent": 0, "total": 0})
        if r.get("status") == "present":
            bucket["present"] += 1
        elif r.get("status") == "absent":
            bucket["absent"] += 1
        bucket["total"] = bucket["present"] + bucket["absent"]
        if mk == this_month_key:
            if r.get("status") == "present":
                this_present += 1
            elif r.get("status") == "absent":
                this_absent += 1

    monthly_list = [{
        "month": mk,
        "present": b["present"],
        "absent": b["absent"],
        "total": b["total"],
    } for mk, b in sorted(monthly.items())]
    this_total = this_present + this_absent
    this_month = {
        "month": this_month_key,
        "present": this_present,
        "absent": this_absent,
        "total": this_total,
        "percentage": round(this_present / this_total * 100, 2) if this_total else 0,
    }
    return {
        "total": total,
        "present": present,
        "absent": absent,
        "percentage": percentage,
        "lastUpdated": iso_utc(last_ts),
        "history": history,
        "thisMonth": this_month,
        "monthly": monthly_list,
    }


# ------------------------------------------------------------
# Placement Rules — SINGLE SOURCE OF TRUTH for cohort thresholds.
# Configured by the Super Admin, stored in db.placement_rules, and
# read fresh on every calculation by every module (student, trainer,
# super_admin). Nothing in this codebase hardcodes a score range.
# ------------------------------------------------------------
def get_placement_rules(db):
    """Fetch the active Placement Rules doc, seeding a sane default
    exactly once if the collection has never been configured."""
    rules = db.placement_rules.find_one({"active": True}, sort=[("updatedAt", -1)])
    if rules:
        return rules
    seed = dict(_DEFAULT_PLACEMENT_RULES_SEED)
    seed["active"] = True
    seed["createdAt"] = now()
    seed["updatedAt"] = now()
    seed["updatedBy"] = "system_default_seed"
    result = db.placement_rules.insert_one(seed)
    seed["_id"] = result.inserted_id
    return seed


def cohort_from_score(db, final_score):
    """Final Employability Score -> cohort, using whatever thresholds are
    CURRENTLY configured in db.placement_rules (fetched fresh, every call).

    Score >= placementReadyThreshold          -> Cohort A (Placement Ready)
    Score >= nearReadyThreshold (and < above) -> Cohort B (Near Ready)
    Score <  nearReadyThreshold                -> Cohort C (High Risk)

    There are no hardcoded ranges: both boundaries come straight from the
    Super Admin's saved Placement Rules, every single call."""
    rules = get_placement_rules(db)
    placement_ready = rules.get("placementReadyThreshold", 75)
    near_ready = rules.get("nearReadyThreshold", 50)
    if final_score >= placement_ready:
        return "A"
    if final_score >= near_ready:
        return "B"
    return "C"


def compute_final_employability_score(db, assessment_percentage, interview_percentage):
    """Weighted combination of assessment score + interview score, using
    the weights (percentages, 0-100, summing to 100) currently configured
    in db.placement_rules."""
    rules = get_placement_rules(db)
    a_weight = rules.get("assessmentWeight", 50)
    i_weight = rules.get("interviewWeight", 50)
    total_weight = (a_weight + i_weight) or 100
    final_score = (
        (assessment_percentage * a_weight) + (interview_percentage * i_weight)
    ) / total_weight
    return round(final_score, 2)


def record_assessment_score_for_cohort(db, student_id, assessment_type, percentage):
    """
    Step 1 of cohort generation. Called after ANY assessment is scored.
    Stores the score on the student record when it's a baseline assessment,
    but does NOT assign a cohort yet — cohort generation additionally
    requires the manual interview score (see check_and_generate_cohort).
    """
    if assessment_type != "baseline":
        return
    db.users.update_one(
        {"_id": student_id},
        {"$set": {
            "baselineAssessmentScore": percentage,
            "baselineAssessmentScoredAt": now(),
        }},
    )
    check_and_generate_cohort(db, student_id)


def record_interview_score_for_cohort(db, student_id, interview_percentage):
    """Step 2 of cohort generation. Called when a manual interview is
    scored. Stores the score, then checks whether cohort generation
    can now proceed."""
    db.users.update_one(
        {"_id": student_id},
        {"$set": {
            "interviewScore": interview_percentage,
            "interviewScoredAt": now(),
        }},
    )
    check_and_generate_cohort(db, student_id)


def check_and_generate_cohort(db, student_id):
    """
    Cohort is generated ONLY when BOTH the assessment score and the
    manual interview score exist. Until then the student stays Entry
    Level, sees no cohort-specific assessments, and has no cohort.

    Workflow: assessment score + interview score -> weighted Final
    Employability Score (weights from Placement Rules) -> cohort
    (ranges from Placement Rules) -> assigned to student.

    Returns the new cohort string, or None if generation didn't happen
    (either already assigned, or one of the two inputs is still missing).
    """
    student = db.users.find_one({"_id": student_id})
    if not student or student.get("cohort") in VALID_COHORTS:
        return None  # already has a real cohort — rule doesn't re-fire

    assessment_score = student.get("baselineAssessmentScore")
    interview_score = student.get("interviewScore")
    if assessment_score is None or interview_score is None:
        return None  # still missing one of the two required inputs

    final_score = compute_final_employability_score(db, assessment_score, interview_score)
    new_cohort = cohort_from_score(db, final_score)

    db.users.update_one(
        {"_id": student_id},
        {"$set": {
            "cohort": new_cohort,
            "finalEmployabilityScore": final_score,
            "cohortAssignedAt": now(),
            "cohortAssignedFrom": "assessment_plus_interview",
        }},
    )
    sync_student_cohort_record(db, student_id, new_cohort, final_score, "assessment_plus_interview")
    return new_cohort


RESULT_STATUS_PENDING = "Interview Pending"
RESULT_STATUS_INTERVIEW_DONE = "Interview Completed"
RESULT_STATUS_VALIDATED = "Validated"


# ------------------------------------------------------------
# Part 5 — dedicated StudentCohort collection (db.student_cohort).
#
# db.users.cohort remains the single field every eligibility check in
# this codebase already reads (quiz_module, student.py, cohort_counts,
# etc.) — this collection is a deliberate ADDITIVE mirror of it, not a
# replacement, so nothing that already depends on db.users.cohort needs
# to change or risks reading a second, possibly-out-of-sync source.
# It exists so a caller that only cares about "current cohort + the
# score that produced it" (e.g. a reporting/export job) doesn't need to
# reconstruct that from db.users' broader profile fields, and so cohort
# history is queryable/indexable on its own without touching db.users.
# ------------------------------------------------------------
def sync_student_cohort_record(db, student_id, cohort, average_score, source):
    """Upserts db.student_cohort with this student's current cohort —
    always called immediately after db.users.cohort is written, from the
    same function, so the two can never drift apart."""
    student = db.users.find_one({"_id": student_id}) or {}
    db.student_cohort.update_one(
        {"studentId": student_id},
        {"$set": {
            "studentId": student_id,
            "rollNumber": student.get("rollNumber"),
            "averageScore": average_score,
            "cohort": cohort,
            "lastUpdated": now(),
            "source": source,
        }},
        upsert=True,
    )


def recompute_cohort_from_quiz_results(db, student_id):
    """
    ROOT-CAUSE FIX: set_quiz_interview_marks()/validate_quiz_result()
    (the Create-Quiz Interview/Validator Verification workflow) computed
    and stored `assignedCohort` on each individual quiz_attempts document
    — but never wrote it back to db.users.cohort, the ONE field every
    quiz-eligibility check in this codebase actually reads
    (student_matches_cohort_target). A student could be fully validated
    into Cohort B here and still never see a Cohort-B-targeted quiz,
    because their actual account-level cohort never moved off Entry
    Level (or whatever the older baseline-assessment/interview engine
    had last set it to).

    Called every time a Create-Quiz result is validated. Recomputes this
    student's cohort from the AVERAGE Final Average across every one of
    their Validated results (spec: "cohort determination depends upon
    average assessment score") — not just the single most-recently
    validated quiz — then commits it as their authoritative cohort and
    mirrors it into db.student_cohort (Part 5).

    Precedence: if this student's cohort was already assigned by the
    older, dedicated baseline-assessment+interview engine
    (`cohortAssignedFrom == "assessment_plus_interview"`), that engine's
    result is left alone — it's the more deliberate, single-purpose
    signal, and silently overwriting it every time an unrelated
    Create-Quiz result gets validated would be its own bug. This function
    only ever sets/updates a cohort that either doesn't exist yet or was
    itself last set by this same Create-Quiz path.

    Returns the new cohort string, or None if the student has no
    Validated Create-Quiz results yet, or their cohort is "owned" by the
    other engine.
    """
    student = db.users.find_one({"_id": student_id}, {"cohortAssignedFrom": 1}) or {}
    if student.get("cohortAssignedFrom") == "assessment_plus_interview":
        return None
    validated = list(db.quiz_attempts.find({
        "studentId": student_id,
        "resultStatus": RESULT_STATUS_VALIDATED,
        "finalAverage": {"$ne": None},
    }))
    if not validated:
        return None
    average_score = round(sum(a["finalAverage"] for a in validated) / len(validated), 2)
    cohort = cohort_from_score(db, average_score)
    db.users.update_one(
        {"_id": student_id},
        {"$set": {
            "cohort": cohort,
            "finalEmployabilityScore": average_score,
            "cohortAssignedAt": now(),
            "cohortAssignedFrom": "quiz_validation_average",
        }},
    )
    sync_student_cohort_record(db, student_id, cohort, average_score, "quiz_validation_average")
    return cohort


def compute_cohort_recalculation_ops(db, rules):
    """
    Batch equivalent of check_and_generate_cohort() + recompute_cohort_from_quiz_results()
    run across EVERY student, built for the "Save Rules recalculates all
    existing students" workflow (Placement Rules spec). Deliberately does
    NOT loop find_one()/update_one() per student — at 5000+ students that
    is 10,000+ round trips. Instead:

      - one find() to pull every "engine 1" (baseline assessment +
        interview) candidate's scores,
      - one aggregation to pull every "engine 2" (validated Create-Quiz
        results) candidate's average score,
      - one more find() (with $in) to check precedence/ownership for
        engine-2 candidates,
      - cohort math done in Python against the *already-fetched* rules
        (no per-student re-read of db.placement_rules),

    then returns ready-to-run pymongo.UpdateOne lists for db.users and
    db.student_cohort so the caller can execute them with bulk_write
    (ideally inside the same transaction as the placement_rules save).

    This has no side effects itself — callers own persistence, which is
    what makes it safe to use inside a transaction with rollback.

    Returns (users_ops, student_cohort_ops, recalculated_count).
    """
    placement_ready = rules.get("placementReadyThreshold", 75)
    near_ready = rules.get("nearReadyThreshold", 50)
    a_weight = rules.get("assessmentWeight", 50)
    i_weight = rules.get("interviewWeight", 50)
    total_weight = (a_weight + i_weight) or 100

    def cohort_for(score):
        if score >= placement_ready:
            return "A"
        if score >= near_ready:
            return "B"
        return "C"

    users_ops = []
    student_cohort_ops = []
    recalculated = 0
    ts = now()

    # ---- Engine 1: baseline assessment + manual interview -------------
    # NOTE: login.py seeds every new student with baselineAssessmentScore
    # and interviewScore explicitly set to None (not yet scored) — so
    # those fields always "$exists". Filtering on $exists alone pulls in
    # every unscored student too, and None * a_weight below then blows up
    # with "unsupported operand type(s) for *: 'NoneType' and 'float'".
    # Matching check_and_generate_cohort()'s single-student equivalent,
    # require an actual (non-null) value on both fields.
    engine1_ids = set()
    for student in db.users.find(
        {"role": "student", "baselineAssessmentScore": {"$ne": None},
         "interviewScore": {"$ne": None}},
        {"_id": 1, "cohort": 1, "baselineAssessmentScore": 1,
         "interviewScore": 1, "rollNumber": 1},
    ):
        engine1_ids.add(student["_id"])
        final_score = round(
            (student["baselineAssessmentScore"] * a_weight
             + student["interviewScore"] * i_weight) / total_weight, 2,
        )
        new_cohort = cohort_for(final_score)
        if student.get("cohort") != new_cohort:
            recalculated += 1
        users_ops.append(UpdateOne(
            {"_id": student["_id"]},
            {"$set": {
                "cohort": new_cohort,
                "finalEmployabilityScore": final_score,
                "cohortAssignedAt": ts,
                "cohortAssignedFrom": "assessment_plus_interview",
            }},
        ))
        student_cohort_ops.append(UpdateOne(
            {"studentId": student["_id"]},
            {"$set": {
                "studentId": student["_id"],
                "rollNumber": student.get("rollNumber"),
                "averageScore": final_score,
                "cohort": new_cohort,
                "lastUpdated": ts,
                "source": "assessment_plus_interview",
            }},
            upsert=True,
        ))

    # ---- Engine 2: validated Create-Quiz results (average finalAverage)
    quiz_averages = {
        row["_id"]: round(row["avgScore"], 2)
        for row in db.quiz_attempts.aggregate([
            {"$match": {"resultStatus": RESULT_STATUS_VALIDATED, "finalAverage": {"$ne": None}}},
            {"$group": {"_id": "$studentId", "avgScore": {"$avg": "$finalAverage"}}},
        ])
    }
    candidate_ids = [sid for sid in quiz_averages if sid not in engine1_ids]
    if candidate_ids:
        # Same precedence guard as recompute_cohort_from_quiz_results(): a
        # cohort "owned" by the assessment+interview engine is left alone.
        for student in db.users.find(
            {"_id": {"$in": candidate_ids}},
            {"_id": 1, "cohort": 1, "cohortAssignedFrom": 1, "rollNumber": 1},
        ):
            if student.get("cohortAssignedFrom") == "assessment_plus_interview":
                continue
            avg_score = quiz_averages[student["_id"]]
            new_cohort = cohort_for(avg_score)
            if student.get("cohort") != new_cohort:
                recalculated += 1
            users_ops.append(UpdateOne(
                {"_id": student["_id"]},
                {"$set": {
                    "cohort": new_cohort,
                    "finalEmployabilityScore": avg_score,
                    "cohortAssignedAt": ts,
                    "cohortAssignedFrom": "quiz_validation_average",
                }},
            ))
            student_cohort_ops.append(UpdateOne(
                {"studentId": student["_id"]},
                {"$set": {
                    "studentId": student["_id"],
                    "rollNumber": student.get("rollNumber"),
                    "averageScore": avg_score,
                    "cohort": new_cohort,
                    "lastUpdated": ts,
                    "source": "quiz_validation_average",
                }},
                upsert=True,
            ))

    return users_ops, student_cohort_ops, recalculated


def backfill_student_cohorts(db):
    """
    SYNCHRONIZE EXISTING DATA (spec Part 5): a retroactive, idempotent
    fix for every student whose data predates this bug fix.

    Two groups of students needed this:
    1. Anyone with a Validated Create-Quiz result recorded BEFORE
       recompute_cohort_from_quiz_results() existed — their
       quiz_attempts document has an `assignedCohort`, but it was never
       propagated to db.users.cohort, so they were (and, until this
       backfill runs, still are) invisible to cohort-targeted quizzes
       despite their own Quiz History page correctly showing "Cohort B".
       Re-running recompute_cohort_from_quiz_results() for each of them
       fixes exactly that.
    2. Anyone whose cohort was already correct in db.users (from either
       engine) but has no db.student_cohort record yet, because that
       collection didn't exist when they were assigned — mirrored in
       without touching db.users at all.

    Safe to run any number of times: every write here is either a no-op
    (a student who's already fully in sync) or a strict correction (a
    stale/missing value getting fixed) — never a regression, thanks to
    recompute_cohort_from_quiz_results()'s own precedence guard, which
    this function relies on rather than duplicates.

    Call this once, manually, right after deploying this fix (see
    POST /admin/student-cohort/backfill) — not automatically on every
    app startup, so it stays visible/auditable rather than silently
    re-running against a growing dataset on every deploy.

    Returns a small summary dict: {"promoted", "mirroredOnly", "checked"}.
    """
    promoted = 0    # db.users.cohort actually changed — these are the
                    # students the original bug was hiding quizzes from
    checked_ids = set(db.quiz_attempts.distinct(
        "studentId", {"resultStatus": RESULT_STATUS_VALIDATED, "finalAverage": {"$ne": None}},
    ))
    for sid in checked_ids:
        before = db.users.find_one({"_id": sid}, {"cohort": 1})
        before_cohort = (before or {}).get("cohort")
        after_cohort = recompute_cohort_from_quiz_results(db, sid)
        if after_cohort and after_cohort != before_cohort:
            promoted += 1

    mirrored = 0
    for student in db.users.find({"role": "student", "cohort": {"$in": list(VALID_COHORTS)}}):
        if db.student_cohort.find_one({"studentId": student["_id"]}):
            continue
        sync_student_cohort_record(
            db, student["_id"], student["cohort"],
            student.get("finalEmployabilityScore"),
            student.get("cohortAssignedFrom") or "backfill",
        )
        mirrored += 1

    return {"promoted": promoted, "mirroredOnly": mirrored, "checked": len(checked_ids)}


def init_quiz_result_fields():
    """Fields merged onto a quiz_attempts doc the moment a quiz is
    submitted, so every submitted attempt is immediately a well-formed
    'marks' record — never a partial/undefined shape."""
    return {
        "interviewMarks": None,
        "finalAverage": None,
        "assignedCohort": None,
        "resultStatus": RESULT_STATUS_PENDING,
        "interviewEnteredBy": None,
        "interviewEnteredAt": None,
        "validatedBy": None,
        "validatedAt": None,
    }


# ------------------------------------------------------------
# Overall Score / Overall Cohort / Placement Readiness — SINGLE SOURCE
# OF TRUTH, shared by every route (Dashboard Home, My Cohort, Placement
# Readiness) that needs to show these numbers. Nothing about this is
# ever computed on the frontend — every caller gets the same already-
# computed dict back from compute_overall_performance() below.
#
# Definition (matches "Overall Cohort Calculation" spec exactly):
#   - A per-quiz "Final Percentage Score" + "Cohort" only exist once
#     interview marks have been entered for that quiz_attempts doc —
#     see set_quiz_interview_marks() above, which sets finalAverage +
#     assignedCohort together, and flips resultStatus away from
#     RESULT_STATUS_PENDING ("Interview Pending"). So "completed /
#     validated" here means resultStatus in (Interview Completed,
#     Validated); "Interview Pending" (finalAverage still None) is
#     exactly the "pending" case the spec says to ignore.
#   - Overall Score = plain average of finalAverage across every
#     completed result for this student (no weighting — "(62+95)/2"
#     in the spec, not a re-weighted formula).
#   - Overall Cohort = cohort_from_score() of that Overall Score,
#     using the CURRENT db.placement_rules bands — i.e. the exact same
#     mapping rule already used everywhere else, never a second set of
#     thresholds invented just for this feature.
# ------------------------------------------------------------
def top_cohort_label(db):
    """The platform's top / readiest cohort. With the three-threshold
    Placement Rules model (placementReadyThreshold is, by validation,
    always >= nearReadyThreshold — see superadmin.py's Save Rules
    validation), Cohort A / "Placement Ready" is always the top band."""
    return "A"


def compute_placement_readiness(db, overall_score, overall_cohort):
    """Placement Readiness = Overall Score + Overall Cohort, judged
    against the Super Admin's own placement-readiness rules (the same
    db.placement_rules cohort bands used everywhere else) — a student
    is 'Ready' once their Overall Cohort is the platform's current top
    band. No separate readiness threshold is invented here; the bands
    a Super Admin already configures ARE the placement-readiness rule."""
    top = top_cohort_label(db)
    ready = overall_cohort is not None and overall_cohort == top
    if overall_score is None:
        summary = ("Complete at least one fully validated assessment "
                    "(quiz score + interview) to see your Placement Readiness.")
    elif ready:
        summary = f"You're in Cohort {overall_cohort} — the platform's top readiness band. You're Placement Ready."
    else:
        cohort_text = f"Cohort {overall_cohort}" if overall_cohort else "Entry Level"
        summary = f"You're currently in {cohort_text}. Reach Cohort {top} to become Placement Ready."
    return {
        "score": overall_score if overall_score is not None else 0,
        "ready": ready,
        "statusLabel": "Ready" if ready else "Not Ready",
        "summary": summary,
    }


def compute_overall_performance(db, student_id):
    """Overall Score, Overall Cohort, average Interview Score,
    Assessments Completed count, and Placement Readiness — all
    recomputed live from db.quiz_attempts on every call, so they are
    always in sync with the latest interview-mark entry or validation.
    The result is also written onto the student's user doc as a
    snapshot (overallPerformanceSnapshot) purely so other modules —
    reports, analytics, future recommendation engines — can read a
    stored value without recomputing; that stored copy is never the
    source of truth, this function always is.
    """
    student = db.users.find_one({"_id": student_id})
    if not student:
        return None

    # v0.0.7 fix: this MUST use the exact same "completed" definition as
    # recompute_cohort_from_quiz_results() (Validated only) — that function
    # is what actually writes db.users.cohort, the field the dashboard
    # header / quiz-eligibility checks read. Previously this also counted
    # RESULT_STATUS_INTERVIEW_DONE ("Interview Completed"), so a result
    # with trainer-entered-but-not-yet-validated marks could push the
    # Overall Score/Overall Cohort shown here into e.g. "Cohort B" while
    # db.users.cohort correctly stayed "Entry Level" pending validation —
    # two different cohort values on the same dashboard for the same
    # student (header vs snapshot), from two different "completed" rules.
    # Restricting both to Validated-only makes them structurally unable
    # to diverge again, rather than just happening to agree today.
    completed = list(db.quiz_attempts.find({
        "studentId": student_id,
        "status": "submitted",
        "resultStatus": RESULT_STATUS_VALIDATED,
        "finalAverage": {"$ne": None},
    }))

    scores = [c["finalAverage"] for c in completed if c.get("finalAverage") is not None]
    overall_score = round(sum(scores) / len(scores), 2) if scores else None
    overall_cohort = cohort_from_score(db, overall_score) if overall_score is not None else None

    interview_scores = [c["interviewMarks"] for c in completed if c.get("interviewMarks") is not None]
    average_interview_score = round(sum(interview_scores) / len(interview_scores), 2) if interview_scores else None

    readiness = compute_placement_readiness(db, overall_score, overall_cohort)

    result = {
        "overallScore": overall_score,
        "overallCohort": overall_cohort,
        "overallCohortLabel": (f"Cohort {overall_cohort}" if overall_cohort else "Entry Level"),
        "averageInterviewScore": average_interview_score,
        "assessmentsCompleted": len(completed),
        "placementReadiness": readiness,
    }

    db.users.update_one(
        {"_id": student_id},
        {"$set": {"overallPerformanceSnapshot": dict(result, computedAt=now())}},
    )
    return result


def serialize_quiz_result(db, attempt, quiz=None):
    """One row of Marks Management data — shape shared by Student 'My
    Results', Trainer/Super Admin 'Assessment Responses', 'Interview
    Verification' and 'Validation Verification'. Every field comes
    straight from the stored attempt document; nothing is recomputed
    on the fly except quiz title fallback."""
    overall = attempt.get("overall") or {}
    if quiz is None:
        quiz = db.quizzes.find_one({"_id": attempt.get("quizId")}) or {}
    submitted_at = attempt.get("submittedAt")
    entered_at = attempt.get("interviewEnteredAt")
    validated_at = attempt.get("validatedAt")
    return {
        "id": str(attempt["_id"]),
        "attemptId": str(attempt["_id"]),
        "studentId": str(attempt["studentId"]),
        "studentName": attempt.get("studentName"),
        "rollNumber": attempt.get("studentRollNumber"),
        "department": attempt.get("department"),
        "college": attempt.get("college"),
        "quizId": str(attempt["quizId"]),
        "quizTitle": attempt.get("quizTitle") or quiz.get("title"),
        "obtainedMarks": overall.get("marksObtained"),
        "totalMarks": overall.get("totalMarks"),
        "percentageMarks": overall.get("percentage"),
        "submittedAt": submitted_at.isoformat() if submitted_at else None,
        "interviewMarks": attempt.get("interviewMarks"),
        "finalAverage": attempt.get("finalAverage"),
        "cohort": attempt.get("assignedCohort"),
        "status": attempt.get("resultStatus") or RESULT_STATUS_PENDING,
        "interviewEnteredAt": entered_at.isoformat() if entered_at else None,
        "validatedAt": validated_at.isoformat() if validated_at else None,
    }


def compute_quiz_analytics(db, college=None, cohort="all", quiz_id="all", department="all", search=""):
    """Dynamic analytics for the Quiz Responses page — Student
    Participation, Performance Metrics, Leaderboard (top 10) and
    Section-wise Performance — computed live from db.quiz_attempts (and,
    for section-wise, each quiz's own question pool). No hardcoded
    values; recomputed on every call rather than cached, so it always
    reflects the latest submissions.

    college=None means unscoped (Super Admin, sees every college); pass
    a college name to scope everything to it (Trainer). cohort/quiz_id/
    department of "all" mean unfiltered on that dimension — same
    semantics either caller uses, so Super Admin and Trainer can never
    compute this differently from each other. `search` matches Student
    Name or Roll Number, same as list_quiz_responses()/list_quiz_results().
    """
    query = {"status": "submitted"}
    if college:
        query["college"] = college
    if cohort and cohort != "all":
        query["cohort"] = cohort
    if quiz_id and quiz_id != "all":
        oid = to_object_id(quiz_id)
        if oid is not None:
            query["quizId"] = oid
    if department and department != "all":
        query["department"] = department
    search_clause = _student_search_query(search)
    if search_clause:
        query.update(search_clause)

    submitted = list(db.quiz_attempts.find(query))

    # Students Attempted / Yet To Attempt — scoped to the same
    # college/cohort filters, over every registered student. (Course and
    # search are per-attempt refinements, not part of the "who's yet to
    # attempt" denominator, which is about the whole eligible population.)
    student_query = {"role": "student"}
    if college:
        student_query["college"] = college
    if cohort and cohort != "all":
        if cohort == ENTRY_LEVEL:
            student_query["$or"] = [{"cohort": None}, {"cohort": {"$exists": False}}]
        else:
            student_query["cohort"] = cohort
    if department and department != "all":
        student_query["department"] = department
    total_students = db.users.count_documents(student_query)
    attempted_ids = {str(a["studentId"]) for a in submitted if a.get("studentId")}
    students_attempted = len(attempted_ids)
    students_yet_to_attempt = max(0, total_students - students_attempted)

    # Performance Metrics
    scores = [a.get("overall", {}).get("percentage") for a in submitted
              if a.get("overall", {}).get("percentage") is not None]
    average_score = round(sum(scores) / len(scores), 2) if scores else 0
    highest_score = max(scores) if scores else 0
    lowest_score = min(scores) if scores else 0

    # Leaderboard — top 10 by score
    ranked = sorted(submitted, key=lambda a: a.get("overall", {}).get("percentage") or 0, reverse=True)[:10]
    leaderboard = [{
        "studentName": a.get("studentName"),
        "college": a.get("college"),
        "score": a.get("overall", {}).get("percentage"),
    } for a in ranked]

    # Section-wise Performance — recomputed read-only from each quiz's
    # question pool + the stored answers, without touching the student
    # quiz-submission flow itself.
    section_totals, section_correct = {}, {}
    quizzes_cache = {}
    for a in submitted:
        qid = a.get("quizId")
        if qid not in quizzes_cache:
            quizzes_cache[qid] = db.quizzes.find_one({"_id": qid}) or {}
        pool = quizzes_cache[qid].get("questions") or []
        answers = a.get("answers") or {}
        for pq in (a.get("pooledQuestions") or []):
            idx = pq.get("poolIndex")
            if idx is None or idx >= len(pool):
                continue
            q = pool[idx]
            section = q.get("section") or "General"
            section_totals[section] = section_totals.get(section, 0) + 1
            raw = answers.get(str(idx))
            given = raw if isinstance(raw, list) else ([raw] if raw not in (None, "") else [])
            try:
                given_idx = sorted({int(x) for x in given})
            except (TypeError, ValueError):
                given_idx = []
            correct_idx = sorted(set(q.get("correct") or []))
            if given_idx and given_idx == correct_idx:
                section_correct[section] = section_correct.get(section, 0) + 1

    section_performance = []
    for section, total in sorted(section_totals.items()):
        correct = section_correct.get(section, 0)
        pct = round((correct / total) * 100, 2) if total else 0
        section_performance.append({"section": section, "percentage": pct})

    return {
        "studentsAttempted": students_attempted,
        "studentsYetToAttempt": students_yet_to_attempt,
        "averageScore": average_score,
        "highestScore": highest_score,
        "lowestScore": lowest_score,
        "leaderboard": leaderboard,
        "sectionPerformance": section_performance,
    }


def list_quiz_responses(db, college=None, cohort="all", quiz_id="all", department="all", search="", limit=200):
    """Raw per-attempt rows for the Quiz Responses table — the same
    db.quiz_attempts query compute_quiz_analytics() aggregates, just
    returned as individual rows instead of summarized. Shared so the
    table and the analytics above can never disagree about which
    attempts are in scope. `department` filters on Course/Department,
    `search` matches Student Name or Roll Number — both resolved here,
    in the database query itself, never in the browser.
    """
    query = {"status": "submitted"}
    if college:
        query["college"] = college
    if cohort and cohort != "all":
        query["cohort"] = cohort
    if quiz_id and quiz_id != "all":
        oid = to_object_id(quiz_id)
        if oid is not None:
            query["quizId"] = oid
    if department and department != "all":
        query["department"] = department
    search_clause = _student_search_query(search)
    if search_clause:
        query.update(search_clause)

    cursor = db.quiz_attempts.find(query).sort("submittedAt", -1).limit(int(limit))
    rows = []
    for a in cursor:
        rows.append({
            "attemptId": str(a["_id"]),
            "studentName": a.get("studentName"),
            "rollNumber": a.get("studentRollNumber"),
            "college": a.get("college"),
            "department": a.get("department"),
            "cohort": a.get("cohort") or ENTRY_LEVEL,
            "quizId": str(a["quizId"]) if a.get("quizId") else None,
            "assessmentName": a.get("quizTitle"),
            "overallPercentage": a.get("overall", {}).get("percentage"),
            "status": "Completed",
            "submittedAt": a["submittedAt"].isoformat() if a.get("submittedAt") else None,
        })
    return rows


def _student_search_query(search):
    """Builds a Mongo `$or` clause matching `search` against Student Name
    OR Roll Number, case-insensitive, partial match — used by every quiz
    Responses/Results search box across both dashboards so they can never
    disagree about what "search" means.

    NOTE on "Register Number": this platform has exactly one student
    identifier field (`rollNumber` on `db.users`, copied onto attempts as
    `studentRollNumber`/`rollNumber`) — there is no separate Register
    Number field anywhere in the schema. Searching Roll Number already
    covers "Register Number (if available)" from the spec; if a distinct
    Register Number field is added to the platform later, add it to the
    `$or` below and nowhere else needs to change.
    """
    search = (search or "").strip()
    if not search:
        return None
    pattern = re.escape(search)
    return {"$or": [
        {"studentName": {"$regex": pattern, "$options": "i"}},
        {"studentRollNumber": {"$regex": pattern, "$options": "i"}},
        {"rollNumber": {"$regex": pattern, "$options": "i"}},
    ]}


def _verification_search_query(search):
    """Same idea as _student_search_query, but also matches Assessment
    Name / College / Department — used by Manual (Interview) Verification
    and Validator Verification's search boxes, which the spec explicitly
    asks to search across all five of those fields at once. quizTitle,
    college and department are all stored directly on the quiz_attempts
    document itself (see student.py's start_quiz / _finalize_attempt), so
    this is still a single flat query, no join required.
    """
    search = (search or "").strip()
    if not search:
        return None
    pattern = re.escape(search)
    return {"$or": [
        {"studentName": {"$regex": pattern, "$options": "i"}},
        {"studentRollNumber": {"$regex": pattern, "$options": "i"}},
        {"quizTitle": {"$regex": pattern, "$options": "i"}},
        {"college": {"$regex": pattern, "$options": "i"}},
        {"department": {"$regex": pattern, "$options": "i"}},
    ]}


def list_distinct_departments(db, college=None):
    """Distinct, non-empty Course/Department names actually in use by
    students — never hardcoded, always read live from db.users. Backs the
    "Course" filter dropdown on both dashboards' Quiz Responses pages
    (GET .../quiz-responses/filters)."""
    query = {"role": "student", "department": {"$nin": [None, ""]}}
    if college:
        query["college"] = college
    return sorted(d for d in db.users.distinct("department", query) if d)


def build_quiz_responses_workbook(rows):
    """Part: Export — builds an in-memory .xlsx (openpyxl, same library
    already used everywhere else in this codebase for Excel I/O) from
    exactly the rows a Quiz Responses table is currently showing —
    whatever `list_quiz_responses()` returned for the caller's current
    filters/search. Never re-queries or re-filters; the export is always
    the same backend-generated data the page is already displaying.
    Returns a BytesIO positioned at 0, ready for flask.send_file.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Responses"
    headers = [
        "Student Name", "Roll Number", "College", "Course/Department",
        "Cohort", "Assessment", "Score (%)", "Status", "Submitted At",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([
            r.get("studentName") or "",
            r.get("rollNumber") or "",
            r.get("college") or "",
            r.get("department") or "",
            r.get("cohort") or "",
            r.get("assessmentName") or "",
            r.get("overallPercentage") if r.get("overallPercentage") is not None else "",
            r.get("status") or "",
            r.get("submittedAt") or "",
        ])
    for col, width in zip("ABCDEFGHI", (22, 16, 22, 20, 12, 26, 12, 12, 22)):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def list_quiz_results(db, college=None, student_id=None, statuses=None, search="", broad_search=False, limit=500):
    """Submitted Create-Quiz attempts, newest first, optionally scoped to
    one college (Trainer) or left unscoped (Super Admin sees everyone).
    Never returns an attempt that hasn't actually been submitted — a
    student who never finished a quiz never appears anywhere here.
    `search` matches Student Name / Roll Number (see
    _student_search_query) by default; pass broad_search=True to also
    match Assessment Name / College / Department (see
    _verification_search_query) — used by Manual (Interview) Verification
    and Validator Verification, which search across all five fields.
    """
    query = {"status": "submitted"}
    if college:
        query["college"] = college
    if student_id is not None:
        query["studentId"] = student_id
    if statuses:
        query["resultStatus"] = {"$in": list(statuses)}
    search_clause = _verification_search_query(search) if broad_search else _student_search_query(search)
    if search_clause:
        query.update(search_clause)
    cursor = db.quiz_attempts.find(query).sort("submittedAt", -1).limit(int(limit))
    quizzes_cache = {}
    rows = []
    for attempt in cursor:
        qid = attempt.get("quizId")
        if qid not in quizzes_cache:
            quizzes_cache[qid] = db.quizzes.find_one({"_id": qid}) or {}
        rows.append(serialize_quiz_result(db, attempt, quiz=quizzes_cache[qid]))
    return rows


def get_quiz_result_or_none(db, attempt_id, college=None):
    """Look up a single submitted quiz_attempts doc, optionally enforcing
    that it belongs to a given college (trainer scoping)."""
    query = {"_id": attempt_id, "status": "submitted"}
    if college:
        query["college"] = college
    return db.quiz_attempts.find_one(query)


def set_quiz_interview_marks(db, attempt_id, marks, entered_by, college=None):
    """
    Trainer/Super Admin action: enter (or update) interview marks (0-100)
    for one quiz result. Immediately recomputes:
        Final Average = (Quiz Percentage + Interview Marks) / 2
    and assigns a cohort using the CURRENT db.placement_rules ranges
    (never hardcoded). Returns (updated_attempt_doc, error_message).
    Re-calling this (Super Admin "update interview marks") simply
    overwrites the previous marks/average/cohort with the new values.
    """
    attempt = get_quiz_result_or_none(db, attempt_id, college=college)
    if not attempt:
        return None, "Quiz result not found, or you do not have access to it."
    if marks is None or isinstance(marks, bool) or not isinstance(marks, (int, float)):
        return None, "Interview marks are required and must be a number."
    if marks < 0 or marks > 100:
        return None, "Interview marks must be between 0 and 100."

    quiz_percentage = (attempt.get("overall") or {}).get("percentage") or 0.0
    final_average = round((float(quiz_percentage) + float(marks)) / 2, 2)
    cohort = cohort_from_score(db, final_average)

    db.quiz_attempts.update_one(
        {"_id": attempt_id},
        {"$set": {
            "interviewMarks": round(float(marks), 2),
            "finalAverage": final_average,
            "assignedCohort": cohort,
            "resultStatus": RESULT_STATUS_INTERVIEW_DONE,
            "interviewEnteredBy": str(entered_by) if entered_by is not None else None,
            "interviewEnteredAt": now(),
        }},
    )
    return db.quiz_attempts.find_one({"_id": attempt_id}), None


def validate_quiz_result(db, attempt_id, validated_by, college=None):
    """Trainer/Super Admin confirms a fully-scored result (Interview
    Completed -> Validated). Cannot validate a result that hasn't had
    interview marks entered yet."""
    attempt = get_quiz_result_or_none(db, attempt_id, college=college)
    if not attempt:
        return None, "Quiz result not found, or you do not have access to it."
    if attempt.get("resultStatus") not in (RESULT_STATUS_INTERVIEW_DONE, RESULT_STATUS_VALIDATED):
        return None, "Interview marks must be entered before this result can be validated."

    db.quiz_attempts.update_one(
        {"_id": attempt_id},
        {"$set": {
            "resultStatus": RESULT_STATUS_VALIDATED,
            "validatedBy": str(validated_by) if validated_by is not None else None,
            "validatedAt": now(),
        }},
    )
    # Root-cause fix (see recompute_cohort_from_quiz_results docstring):
    # validating a result is the point this workflow considers a cohort
    # assignment "final" — propagate it to the student's actual account
    # cohort (db.users.cohort), the field every quiz-eligibility check
    # reads, instead of leaving it stranded on this one attempt document.
    recompute_cohort_from_quiz_results(db, attempt["studentId"])
    return db.quiz_attempts.find_one({"_id": attempt_id}), None


def cohort_counts(db, college=None):
    """Shared by Super Admin + Trainer dashboards: counts of students in
    Cohort A / B / C / Entry Level. `college` optionally scopes to one college
    (trainers only see their own college's students)."""
    match = {"role": "student"}
    if college:
        match["college"] = college
    pipeline = [
        {"$match": match},
        {"$group": {"_id": {"$ifNull": ["$cohort", ENTRY_LEVEL]}, "count": {"$sum": 1}}},
    ]
    counts = {"A": 0, "B": 0, "C": 0, ENTRY_LEVEL: 0}
    for row in db.users.aggregate(pipeline):
        key = row["_id"] if row["_id"] in VALID_COHORTS else ENTRY_LEVEL
        counts[key] = counts.get(key, 0) + row["count"]
    return counts


def cohort_query_for_target(cohort_target):
    """Turn an assessment's cohortTarget into a Mongo filter fragment for
    matching eligible students in db.users."""
    if cohort_target == "all":
        return {}
    if cohort_target == ENTRY_LEVEL:
        return {"$or": [{"cohort": None}, {"cohort": {"$exists": False}}]}
    if cohort_target in VALID_COHORTS:
        return {"cohort": cohort_target}
    return {"_id": None}  # invalid target matches nobody


def student_matches_cohort_target(student_doc, cohort_target):
    if cohort_target == "all":
        return True
    if cohort_target == ENTRY_LEVEL:
        return student_cohort_label(student_doc) == ENTRY_LEVEL
    return student_doc.get("cohort") == cohort_target


# ------------------------------------------------------------
# Master Excel workbook parsing (ONE upload -> many sheets -> question bank)
# ------------------------------------------------------------
QUESTION_COL_ALIASES = {
    "question": {"question", "questions", "question text", "questiontext"},
    "option_a": {"option a", "optiona", "a", "option1"},
    "option_b": {"option b", "optionb", "b", "option2"},
    "option_c": {"option c", "optionc", "c", "option3"},
    "option_d": {"option d", "optiond", "d", "option4"},
    "answer": {"answer", "correct answer", "correctanswer", "correct option", "correct"},
    "difficulty": {"difficulty", "level"},
}


def _match_column(header_cells):
    """Map a header row (list of cell values) to column-index-by-role."""
    mapping = {}
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        for role, aliases in QUESTION_COL_ALIASES.items():
            if key in aliases:
                mapping[role] = idx
                break
    return mapping


def parse_master_workbook(file_stream, uploaded_by):
    """
    Parse ONE master Excel workbook (openpyxl). Every sheet is a section
    (e.g. Communication, Programming, Reasoning, Professionalism,
    Career Readiness). Each sheet has ~100 rows of questions.

    Returns (questions: list[dict ready for insert_many], errors: list[str]).
    Nothing is written to the DB here — the caller decides whether to persist.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    questions = []
    errors = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section = sheet_name.strip()
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        col_map = _match_column(list(header))
        if "question" not in col_map or "answer" not in col_map:
            errors.append(
                f"Sheet '{sheet_name}': could not find required 'Question' and "
                f"'Answer' columns — sheet skipped."
            )
            continue

        for row_idx, row in enumerate(rows, start=2):
            if row is None or all(c is None for c in row):
                continue
            question_text = row[col_map["question"]] if col_map.get("question") is not None else None
            if not question_text or not str(question_text).strip():
                continue

            options = []
            for opt_key in ("option_a", "option_b", "option_c", "option_d"):
                if opt_key in col_map and col_map[opt_key] < len(row):
                    val = row[col_map[opt_key]]
                    if val is not None and str(val).strip():
                        options.append(str(val).strip())

            answer = row[col_map["answer"]] if col_map.get("answer") is not None else None
            if answer is None or not str(answer).strip():
                errors.append(f"Sheet '{sheet_name}' row {row_idx}: missing answer — skipped.")
                continue

            difficulty = None
            if "difficulty" in col_map and col_map["difficulty"] < len(row):
                d_val = row[col_map["difficulty"]]
                if d_val is not None and str(d_val).strip():
                    difficulty = str(d_val).strip().lower()

            questions.append({
                "section": section,
                "questionText": str(question_text).strip(),
                "options": options,
                "correctAnswer": str(answer).strip(),
                "difficulty": difficulty,
                "createdAt": now(),
                "createdBy": uploaded_by,
                "active": True,
            })

        if not any(q["section"] == section for q in questions) and not errors:
            errors.append(f"Sheet '{sheet_name}': no valid questions found.")

    return questions, errors


# ------------------------------------------------------------
# Random Question Engine
# ------------------------------------------------------------
def select_random_questions(db, section_counts):
    """
    section_counts: {"Communication": 10, "Programming": 15, ...}
    Returns (selected_questions: list[dict], shortage: dict section->deficit)
    Uses MongoDB's $sample for true random, non-sequential selection —
    a fresh random draw every time this is called (i.e. every student
    attempt gets its own random set).
    """
    selected = []
    shortage = {}
    for section, count in section_counts.items():
        if not count or count <= 0:
            continue
        pipeline = [
            {"$match": {"section": section, "active": True}},
            {"$sample": {"size": int(count)}},
        ]
        docs = list(db.questions.aggregate(pipeline))
        if len(docs) < count:
            shortage[section] = count - len(docs)
        selected.extend(docs)
    random.shuffle(selected)
    return selected, shortage


def select_random_questions_v2(db, assessment):
    """
    Preferred selection entry point for student.start_assessment().

    If the assessment document defines a difficultyDistribution, e.g.
    {"easy": 10, "medium": 10, "hard": 5}, this draws exactly that many
    questions per difficulty level — independently shuffled per level,
    then the merged set shuffled again — mirroring quiz_module's
    difficulty-based draw for the separate manually-authored-quiz
    feature, so both systems behave identically for this kind of config.

    Today, the Trainer/Super Admin "Create Assessment" wizard has no
    difficultyDistribution field (only sectionCounts), so every existing
    assessment falls through to the original select_random_questions()
    above completely unchanged. This function only activates the
    difficulty-based path the moment such a field is added to an
    assessment document — no other code needs to change when that
    happens.
    """
    dist = assessment.get("difficultyDistribution")
    if not dist:
        return select_random_questions(db, assessment.get("sectionCounts", {}))

    selected = []
    shortage = {}
    for level in ("easy", "medium", "hard"):
        count = int(dist.get(level) or 0)
        if count <= 0:
            continue
        pipeline = [
            {"$match": {"difficulty": level, "active": True}},
            {"$sample": {"size": count}},
        ]
        docs = list(db.questions.aggregate(pipeline))
        if len(docs) < count:
            shortage[level] = count - len(docs)
        selected.extend(docs)
    random.shuffle(selected)
    return selected, shortage


def strip_answers(questions):
    """Student-facing question payload: never leak correctAnswer to the client."""
    safe = []
    for q in questions:
        safe.append({
            "id": str(q["_id"]),
            "section": q["section"],
            "questionText": q["questionText"],
            "options": q.get("options", []),
            "difficulty": q.get("difficulty"),
        })
    return safe


# ------------------------------------------------------------
# Scoring
# ------------------------------------------------------------
def calculate_score(db, question_ids, answers):
    """
    question_ids: list[ObjectId] — the exact questions this attempt used.
    answers: {question_id_str: submitted_answer_str}
    Returns dict: {
        sectionScores: {section: {"correct": n, "total": n, "percentage": p}},
        overall: {"correct": n, "total": n, "percentage": p, "attempted": n,
                  "wrong": n, "unanswered": n}
    }
    """
    questions = list(db.questions.find({"_id": {"$in": question_ids}}))
    section_totals = {}
    section_correct = {}
    total = 0
    correct = 0
    attempted = 0
    unanswered = 0

    for q in questions:
        section = q["section"]
        qid = str(q["_id"])
        section_totals[section] = section_totals.get(section, 0) + 1
        total += 1
        submitted = answers.get(qid)
        if submitted is None or str(submitted).strip() == "":
            unanswered += 1
            continue
        attempted += 1
        is_correct = str(submitted).strip().lower() == str(q["correctAnswer"]).strip().lower()
        if is_correct:
            section_correct[section] = section_correct.get(section, 0) + 1
            correct += 1

    section_scores = {}
    for section, sec_total in section_totals.items():
        sec_correct = section_correct.get(section, 0)
        pct = round((sec_correct / sec_total) * 100, 2) if sec_total else 0.0
        section_scores[section] = {"correct": sec_correct, "total": sec_total, "percentage": pct}

    overall_pct = round((correct / total) * 100, 2) if total else 0.0
    return {
        "sectionScores": section_scores,
        "overall": {
            "correct": correct,
            "total": total,
            "percentage": overall_pct,
            "attempted": attempted,
            "wrong": attempted - correct,
            "unanswered": unanswered,
        },
    }


# ------------------------------------------------------------
# Activity Logging — single source of truth for every "Recent
# Activity" panel across Student / Trainer / Super Admin. Call
# log_activity() right after any state-changing action succeeds;
# read it back with get_recent_activity().
# ------------------------------------------------------------
def log_activity(db, actor_id, actor_role, action, description, college=None, student_id=None, meta=None):
    """
    actor_id:    the ObjectId/str of whoever performed the action (trainer,
                 student, or super admin). Stored as string for easy querying.
    actor_role:  'student' | 'trainer' | 'super_admin'
    action:      short machine-readable code, e.g. 'assessment_created',
                 'assessment_published', 'assessment_submitted',
                 'manual_interview_scheduled', 'manual_interview_scored',
                 'report_downloaded', 'profile_updated', 'password_changed'.
    description: human-readable sentence for display, e.g.
                 'Created assessment "Communication Baseline"'.
    college:     used to scope trainer/admin "Recent Activity" queries.
    student_id:  if this activity is about a specific student (e.g. a
                 trainer scoring their interview), stored so the STUDENT's
                 own Recent Activity can also show it.
    meta:        free-form dict with extra ids (assessmentId, interviewId...).
    """
    db.activity_log.insert_one({
        "actorId": str(actor_id) if actor_id is not None else None,
        "actorRole": actor_role,
        "action": action,
        "description": description,
        "college": college,
        "studentId": str(student_id) if student_id is not None else None,
        "meta": meta or {},
        "createdAt": now(),
    })


def get_recent_activity(db, query=None, limit=20):
    """Returns newest-first activity rows, serialized for direct JSON use.
    createdAtIST is the same timestamp rendered in Asia/Kolkata with a
    24-hour clock (fmt_ist) — the frontend only ever displays this field,
    never doing its own timezone arithmetic."""
    cursor = db.activity_log.find(query or {}).sort("createdAt", -1).limit(limit)
    rows = []
    for d in cursor:
        rows.append({
            "id": str(d["_id"]),
            "action": d.get("action"),
            "description": d.get("description"),
            "actorRole": d.get("actorRole"),
            "meta": d.get("meta", {}),
            "createdAt": d["createdAt"].isoformat() if d.get("createdAt") else None,
            "createdAtIST": fmt_ist(d.get("createdAt"), "%d %b %Y %H:%M"),
        })
    return rows
