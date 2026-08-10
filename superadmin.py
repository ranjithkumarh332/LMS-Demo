"""
============================================================
 superadmin.py — Super Admin Dashboard backend
============================================================
Registered at /api/admin. Covers everything the Super Admin
Dashboard needs, all backed by real MongoDB data (no hardcoded
sample values):

  - Master Excel question-bank upload (ONE workbook, many sheets)
  - Assessment creation (incl. the new "Entry Level" cohort option
    and the random-question-count-per-section selector)
  - Assessment list (feeds the Assessment Results page dropdown)
  - Quiz Management / Quiz Responses stats, with dynamic Quiz Responses
    analytics (participation, performance, leaderboard, section-wise)
  - Cohort breakdown (A / B / C / Entry Level) — uses the SAME
    cohort helpers as trainer.py so the two dashboards can never
    disagree about who is in which cohort.
  - Dynamic charts: Skill Radar, Score by Category, Overall Score,
    Category Percentage, Performance Trends — all computed live
    from db.assessment_attempts.

Only Super Admin (role="super_admin") can call these routes.
"""

import hashlib
import io
import logging
import re
import urllib.parse
from datetime import datetime, timedelta

from flask import Blueprint, request, send_file
from flask_jwt_extended import get_jwt, get_jwt_identity
from pymongo.errors import PyMongoError, OperationFailure, ConfigurationError

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Dedicated logger for this module, same pattern as quiz_module.py /
# question_bank.py — goes through whatever handler/level app.py
# configures, so Save Rules failures are always visible server-side
# even when the HTTP response has to stay generic for the client.
logger = logging.getLogger("superadmin")

from quiz_common import (
    ok, error, role_required, now, to_object_id, serialize, iso_utc,
    fmt_ist,
    VALID_COHORT_TARGETS, VALID_COHORTS, ENTRY_LEVEL,
    parse_master_workbook, cohort_counts,
    get_placement_rules, top_cohort_label, student_cohort_label,
    attendance_summary,
    compute_cohort_recalculation_ops,
    list_quiz_results, set_quiz_interview_marks, validate_quiz_result,
    serialize_quiz_result, RESULT_STATUS_INTERVIEW_DONE, RESULT_STATUS_VALIDATED,
    compute_quiz_analytics, list_quiz_responses,
    list_distinct_departments, build_quiz_responses_workbook,
    backfill_student_cohorts,
    log_activity,
)
from colleges import resolve_active_college, resolve_active_department, _generate_temp_password

# Roles managed from the Super Admin "User Management" page. Super Admin
# itself is the single hardcoded account (see login.py) and never appears
# here — nothing to activate/deactivate/delete for it.
MANAGED_USER_ROLES = ["student", "trainer", "college_admin"]

ROLE_ALIASES = {
    "student": "student",
    "students": "student",
    "trainer": "trainer",
    "trainers": "trainer",
    "college_admin": "college_admin",
    "collegeadmin": "college_admin",
    "college admin": "college_admin",
    "college_admins": "college_admin",
}


# ==============================================================
# SCHEDULE SESSION — Super Admin "Schedule Session" module.
# Backing collection: db.workshop_sessions. Every dropdown (College,
# Department, Trainer) is resolved server-side against the real
# db.colleges / db.departments / db.users collections — nothing here
# is ever accepted as a free-text/hardcoded value from the client, only
# ids that are re-validated on every write. Colleges/Departments/
# Trainers are all stored as arrays of ids (+ a denormalized array of
# display names, snapshotted at schedule time so the session's own
# history never changes retroactively if a college is later renamed).
# Designed so Student/Trainer/College Admin/Management Office session
# views can be added later purely as new read endpoints against this
# same collection — no schema change required (see spec item 12).
# ==============================================================
SESSION_COHORT_OPTIONS = ("All Cohorts", "Cohort A", "Cohort B", "Cohort C")
SESSION_ATTENDANCE_OPTIONS = ("mandatory", "optional")
SESSION_STATUS_OPTIONS = ("scheduled", "completed", "cancelled")
# Mode of Conduct — how a session is delivered. Offline sessions need a
# venue; online sessions carry no venue and never take a meeting link at
# creation time (the link is added later via the Update Meeting Link flow,
# only while the session is live).
SESSION_MODE_OPTIONS = ("online", "offline")
_TIME_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")

# ------------------------------------------------------------
# STUDENT PROFILE MODULE — Super Admin "Edit Profile" + "Bulk CSV
# Import". Students are stored with `mobile` (see login.py register),
# so these validators match the registration rules exactly: email is a
# plain format check, mobile is exactly 10 digits.
# ==============================================================
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_MOBILE_RE = re.compile(r"^\d{10}$")


# ------------------------------------------------------------
# ATTENDANCE MODULE — Super Admin "Mark Attendance" / "View Records".
# Backing collection: db.attendance. One document per student per date
# (enforced by a unique index on (studentId, date) created in app.py),
# so re-marking the same student on the same date always updates the
# existing record instead of creating a duplicate. Every filter value
# (college / department / cohort / trainer) is resolved server-side
# against db.colleges / db.departments / db.users / db.attendance —
# nothing is hardcoded, nothing is accepted as free text.
# ==============================================================
ATTENDANCE_STATUSES = ("present", "absent")
ATTENDANCE_STATUS_LABELS = {"present": "Present", "absent": "Absent", "not_marked": "Not Yet Marked"}
# Cohort labels — derived from the platform's single cohort system
# (VALID_COHORTS + ENTRY_LEVEL), NOT a separate hardcoded list.
ATTENDANCE_COHORT_LABELS = {
    "A": "Cohort A – Placement Ready",
    "B": "Cohort B – Near Ready",
    "C": "Cohort C",
    ENTRY_LEVEL: "Entry Level",
}
_ATTENDANCE_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _parse_optional_id_list(raw):
    """Like _parse_id_list but a filter list that may be empty/absent
    (an empty selection means 'all'). Returns (object_ids, error)."""
    if raw is None or raw == "":
        return [], None
    if isinstance(raw, str):
        parts = [p.strip() for p in raw.split(",") if p.strip()]
    elif isinstance(raw, list):
        parts = [str(p) for p in raw]
    else:
        return None, "Invalid id list."
    seen = set()
    oids = []
    for item in parts:
        oid = to_object_id(item)
        if not oid:
            return None, "One of the selected ids is invalid."
        if oid not in seen:
            seen.add(oid)
            oids.append(oid)
    return oids, None


def _attendance_student_query(db, college_oids, department_oids, cohorts):
    """Mongo filter for students matching the Mark/View filters. An empty
    list on any dimension means 'no restriction' (i.e. all). Cohorts is a
    list of 'A'/'B'/'C'/'entry_level' values; 'all' disables the filter."""
    query = {
        "role": "student",
        "approvalStatus": {"$in": ["approved", "suspended"]},
        "isDeleted": {"$ne": True},
    }
    if college_oids:
        query["collegeId"] = {"$in": college_oids}
    if department_oids:
        query["departmentId"] = {"$in": department_oids}
    if cohorts and "all" not in cohorts:
        branches = []
        if ENTRY_LEVEL in cohorts:
            branches.append({"cohort": {"$nin": list(VALID_COHORTS)}})
        for c in cohorts:
            if c in VALID_COHORTS:
                branches.append({"cohort": c})
        if branches:
            query["$or"] = branches
    return query


def _session_target_students(db, doc):
    """The exact audience a workshop session was scheduled for (spec §28).
    If the session stored specific student ids, those are returned. Otherwise
    students are matched by the session's collegeIds + departmentIds + cohort
    — the exact filters chosen at schedule time, never a blanket college
    dump. Used by both Mark Attendance and the Student Workshops feed."""
    specific = doc.get("studentIds")
    if specific:
        oids = [o for o in (to_object_id(s) for s in specific) if o]
        if oids:
            return list(db.users.find(
                {"_id": {"$in": oids}, "role": "student", "isDeleted": {"$ne": True}}
            ).sort("fullName", 1))

    query = {"role": "student", "isDeleted": {"$ne": True}}
    college_oids = [o for o in (to_object_id(x) for x in doc.get("collegeIds", [])) if o]
    if college_oids:
        query["collegeId"] = {"$in": college_oids}
    department_oids = [o for o in (to_object_id(x) for x in doc.get("departmentIds", [])) if o]
    if department_oids:
        query["departmentId"] = {"$in": department_oids}
    cohort = doc.get("cohort")
    if cohort in ("Cohort A", "Cohort B", "Cohort C"):
        query["cohort"] = cohort[-1]
    return list(db.users.find(query).sort("fullName", 1))


def _attendance_student_public(doc):
    """Serialize a student user doc for the attendance mark/list views.
    cohortLabel uses the same labels the rest of the platform uses."""
    cohort = student_cohort_label(doc)
    return {
        "id": str(doc["_id"]),
        "name": doc.get("fullName") or doc.get("email") or "—",
        "rollNumber": doc.get("rollNumber") or "—",
        "college": doc.get("college") or "—",
        "collegeId": str(doc["collegeId"]) if doc.get("collegeId") else None,
        "department": doc.get("department") or "—",
        "departmentId": str(doc["departmentId"]) if doc.get("departmentId") else None,
        "cohort": cohort,
        "cohortLabel": ATTENDANCE_COHORT_LABELS.get(cohort, cohort),
    }


def _safe_filename(name):
    """Sanitize a name for use in a download filename (drops any
    characters that would be invalid in a Windows/Linux filename)."""
    cleaned = re.sub(r'[^A-Za-z0-9 _\-]+', '', str(name or 'report')).strip().replace(' ', '_')
    return cleaned[:60] or 'report'


def _parse_id_list(raw, field_label):
    """Normalize a request field that must be a non-empty list of id
    strings. Returns (object_ids, error_message) — object_ids is a
    deduped list preserving order; error_message is None on success."""
    if raw is None:
        raw = []
    if not isinstance(raw, list):
        return None, f"{field_label} must be a list of ids."
    if not raw:
        return None, f"Select at least one {field_label.lower()}."
    seen = set()
    oids = []
    for item in raw:
        oid = to_object_id(item)
        if not oid:
            return None, f"One of the selected {field_label.lower()} is invalid."
        if oid not in seen:
            seen.add(oid)
            oids.append(oid)
    return oids, None


def _resolve_colleges(db, college_ids):
    """Every id must correspond to an existing, active college. Returns
    (docs_by_id, error) — docs_by_id preserves the request order."""
    docs = {d["_id"]: d for d in db.colleges.find({"_id": {"$in": college_ids}, "status": "active"})}
    missing = [str(oid) for oid in college_ids if oid not in docs]
    if missing:
        return None, "One or more selected colleges is invalid or inactive."
    return [docs[oid] for oid in college_ids], None


def _resolve_departments(db, department_ids):
    """Returns (docs_in_request_order, college_names_by_id, error)."""
    docs = {d["_id"]: d for d in db.departments.find({"_id": {"$in": department_ids}, "status": "active"})}
    missing = [str(oid) for oid in department_ids if oid not in docs]
    if missing:
        return None, None, "One or more selected departments is invalid or inactive."
    college_names = {c["_id"]: c.get("college_name") for c in db.colleges.find({}, {"college_name": 1})}
    ordered = [docs[oid] for oid in department_ids]
    return ordered, college_names, None


def _resolve_trainers(db, trainer_ids):
    docs = {
        d["_id"]: d for d in db.users.find(
            {"_id": {"$in": trainer_ids}, "role": "trainer", "approvalStatus": {"$in": ["approved", "suspended"]}}
        )
    }
    missing = [str(oid) for oid in trainer_ids if oid not in docs]
    if missing:
        return None, "One or more selected trainers is invalid."
    return [docs[oid] for oid in trainer_ids], None


def _attendance_window_open(doc):
    """True only for a Mandatory session, and only while 'now' falls
    within that session's own Session Date + Start/End Time (spec §10).
    Optional sessions never expose an attendance window/button."""
    return _attendance_window_state(doc) == "open"


def _attendance_window_state(doc):
    """Three-state attendance availability for a session card, computed
    entirely from the stored session schedule against the server clock
    (the client only mirrors it):
      - "not_applicable" — attendance not required (Optional)
      - "upcoming"       — before the session start, attendance locked
      - "open"           — inside the session window (start <= now <= end)
      - "closed"         — after the session end time, attendance locked
    The session date/start/end are stored as plain wall-clock strings the
    admin entered, so they are compared to the server's local wall clock —
    the same convention every other session-time rule in this module uses
    (_session_started / _meeting_link_window_open), keeping them consistent
    and keeping the browser out of the time decision entirely."""
    if doc.get("attendanceRequirement") != "mandatory":
        return "not_applicable"
    date_str, start, end = doc.get("date"), doc.get("startTime"), doc.get("endTime")
    if not (date_str and start and end):
        return "not_applicable"
    try:
        start_dt = datetime.strptime(f"{date_str} {start}", "%Y-%m-%d %H:%M")
        end_dt = datetime.strptime(f"{date_str} {end}", "%Y-%m-%d %H:%M")
    except ValueError:
        return "not_applicable"
    now_dt = datetime.now()
    if now_dt < start_dt:
        return "upcoming"
    if now_dt > end_dt:
        return "closed"
    return "open"


def workshop_session_public(doc):
    if not doc:
        return None
    colleges = doc.get("collegeNames", [])
    departments = doc.get("departmentNames", [])
    trainers = doc.get("trainerNames", [])
    return {
        "id": str(doc["_id"]),
        "name": doc.get("name"),
        "collegeIds": [str(x) for x in doc.get("collegeIds", [])],
        "colleges": colleges,
        "college": ", ".join(colleges) or "—",
        "departmentIds": [str(x) for x in doc.get("departmentIds", [])],
        "departments": departments,
        "department": ", ".join(departments) or "—",
        "trainerIds": [str(x) for x in doc.get("trainerIds", [])],
        "trainers": trainers,
        "trainer": ", ".join(trainers) or "—",
        "cohort": doc.get("cohort"),
        "date": doc.get("date"),
        "startTime": doc.get("startTime"),
        "endTime": doc.get("endTime"),
        "time": f"{doc.get('startTime', '')} \u2013 {doc.get('endTime', '')}",
        "mode": doc.get("mode") or "offline",
        "venue": doc.get("venue"),
        "meetingLink": doc.get("meetingLink"),
        "meetingLinkUpdatedAt": iso_utc(doc.get("meetingLinkUpdatedAt")),
        "meetingLinkUpdatedAtIST": fmt_ist(doc.get("meetingLinkUpdatedAt")),
        "meetingLinkUpdatedBy": doc.get("meetingLinkUpdatedBy"),
        "meetingLinkWindowOpen": _meeting_link_window_open(doc),
        "attendanceRequirement": doc.get("attendanceRequirement"),
        "attendanceWindowOpen": _attendance_window_open(doc),
        "attendanceWindowState": _attendance_window_state(doc),
        "status": doc.get("status", "scheduled"),
        "approvalStatus": doc.get("approvalStatus", "pending"),
        "createdBy": doc.get("createdBy"),
        "createdAt": iso_utc(doc.get("createdAt")),
        "createdAtIST": fmt_ist(doc.get("createdAt")),
        "updatedAt": iso_utc(doc.get("updatedAt")),
        "updatedAtIST": fmt_ist(doc.get("updatedAt")),
    }


def _session_start_dt(doc):
    """Naive datetime for a session's start (date + startTime), or None if
    the stored schedule can't be parsed. Server-local time throughout —
    the same convention every other session-time rule in this module uses
    (_session_started / _attendance_window_open), so all timing decisions
    agree with each other and the browser is never the time source."""
    date_str, start = doc.get("date"), doc.get("startTime")
    if not (date_str and start):
        return None
    try:
        return datetime.strptime(f"{date_str} {start}", "%Y-%m-%d %H:%M")
    except ValueError:
        return None


def _session_started(doc):
    """A workshop session counts as started once its start datetime has
    passed. Business rule: Edit / Delete are only allowed before that
    moment — enforced server-side here, and mirrored in the frontend by
    hiding the Edit/Delete buttons on started sessions."""
    start_dt = _session_start_dt(doc)
    return start_dt is not None and datetime.now() >= start_dt


def _auto_complete_workshop_sessions(db):
    """Auto-complete every scheduled workshop session whose end datetime
    has already passed. Stored date + endTime are plain wall-clock strings
    (the same convention _attendance_window_state / _session_started use),
    so they are compared to the server's local wall clock — the browser is
    never the time source. The status change is persisted to
    db.workshop_sessions, so the Super Admin intervention list (and every
    other dashboard reading this collection) shows Completed without a
    manual status change or a frontend refresh. Cancelled sessions are
    left untouched."""
    now_dt = datetime.now()
    updated = 0
    for doc in db.workshop_sessions.find(
        {"status": "scheduled"}, {"date": 1, "endTime": 1}
    ):
        date_str, end = doc.get("date"), doc.get("endTime")
        if not (date_str and end):
            continue
        try:
            end_dt = datetime.strptime(f"{date_str} {end}", "%Y-%m-%d %H:%M")
        except ValueError:
            continue
        if now_dt >= end_dt:
            db.workshop_sessions.update_one(
                {"_id": doc["_id"]},
                {"$set": {"status": "completed", "updatedAt": now()}},
            )
            updated += 1
    return updated


def _meeting_link_window_open(doc):
    """Online sessions may take a meeting link only while the session is
    live: Start Time <= now <= End Time. Before the start time, and once
    the session has reached its end time, the meeting link can no longer
    be updated. Computed entirely from the server clock against the stored
    date/start/end wall-clock schedule (same convention as
    _attendance_window_state) — the client only mirrors it."""
    if doc.get("mode") != "online":
        return False
    date_str, start, end = doc.get("date"), doc.get("startTime"), doc.get("endTime")
    if not (date_str and start and end):
        return False
    try:
        start_dt = datetime.strptime(f"{date_str} {start}", "%Y-%m-%d %H:%M")
        end_dt = datetime.strptime(f"{date_str} {end}", "%Y-%m-%d %H:%M")
    except ValueError:
        return False
    now_dt = datetime.now()
    return start_dt <= now_dt <= end_dt


def _valid_meeting_link(url):
    """A meeting link must be an absolute http/https URL. Stored as-is and
    surfaced to students as a join link — anything else is rejected."""
    candidate = (url or "").strip()
    if not candidate or len(candidate) > 2000:
        return None
    parsed = urllib.parse.urlparse(candidate)
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        return None
    return candidate




def _user_public_row(doc):
    """Serialize a db.users document for the User Management table. Every
    field here is read straight off the real MongoDB document — no mock
    values. `college` is intentionally omitted for trainers: trainers are
    not directly associated with a single college on this platform, so the
    Trainer table should never show one (see colleges.py assign_trainer_college,
    which lets a trainer be linked to more than one college over time)."""
    role = doc.get("role")
    last_login = doc.get("lastLoginAt")
    created_at = doc.get("createdAt")
    is_active = doc.get("approvalStatus") == "approved"
    return {
        "id": str(doc["_id"]),
        "name": doc.get("fullName") or "—",
        "email": doc.get("email") or "—",
        "role": role,
        "college": doc.get("college") if role != "trainer" else None,
        "department": doc.get("department"),
        "rollNumber": doc.get("rollNumber"),
        "employeeId": doc.get("employeeId"),
        "collegeAdminId": doc.get("tneaCode") if role == "college_admin" else None,
        "status": "active" if is_active else "inactive",
        # Both keys populated: `lastLogin` is what the frontend table reads,
        # `lastLoginAt` (ISO, or null) is kept for anything that wants the
        # raw value. Never a dummy date — null means "never logged in".
        "lastLogin": iso_utc(last_login),
        "lastLoginAt": iso_utc(last_login),
        "createdAt": iso_utc(created_at),
    }


def _month_bounds(months_back=6):
    """Ascending list of (start, end, label) tuples for the last
    `months_back` calendar months (label is the 3-letter month name),
    most recent month last. `end` is exclusive (start of next month).
    Naive UTC datetimes throughout — matches what PyMongo hands back
    for every timestamp field in this database."""
    today = datetime.utcnow()
    y, m = today.year, today.month
    bounds = []
    for i in range(months_back - 1, -1, -1):
        mm, yy = m - i, y
        while mm <= 0:
            mm += 12
            yy -= 1
        start = datetime(yy, mm, 1)
        end = datetime(yy + 1, 1, 1) if mm == 12 else datetime(yy, mm + 1, 1)
        bounds.append((start, end, start.strftime("%b")))
    return bounds


def _bucket_dates(dates, bounds):
    """Count how many `dates` (datetimes, Nones filtered out) fall inside
    each (start, end) window in `bounds`. Not cumulative — one month's
    count only counts that month's events."""
    counts = [0] * len(bounds)
    for d in dates:
        if not d:
            continue
        for idx, (start, end, _label) in enumerate(bounds):
            if start <= d < end:
                counts[idx] += 1
                break
    return counts


def _pct_trend(series):
    """% change of the last bucket vs the one before it. 0 if there's
    nothing to compare against (avoids a misleading +inf/-100%)."""
    if len(series) < 2 or not series[-2]:
        return 0
    return round(((series[-1] - series[-2]) / series[-2]) * 100, 1)


def _notification_id(kind, source_id):
    """Deterministic id for a notification derived from live data (no
    notification is ever stored — it's recomputed from the source
    collections on every request). Stable across requests as long as the
    underlying source document doesn't change, which is what lets
    read/unread state persist in db.notification_reads."""
    raw = f"{kind}:{source_id}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:24]


def init_superadmin(db, bcrypt=None):
    bp = Blueprint("superadmin", __name__)

    questions = db.questions
    assessments = db.assessments
    attempts = db.assessment_attempts
    users = db.users

    # ==========================================================
    # 1. MASTER EXCEL UPLOAD — ONE workbook, every sheet = a section
    # ==========================================================
    @bp.route("/questions/upload", methods=["POST"])
    @role_required("super_admin", "trainer")
    def upload_master_workbook():
        if "file" not in request.files:
            return error("No file uploaded. Attach the master Excel workbook as 'file'.")
        f = request.files["file"]
        if not f.filename:
            return error("Empty file upload.")
        if not f.filename.lower().endswith((".xlsx", ".xlsm")):
            return error("Only .xlsx / .xlsm workbooks are supported.")

        try:
            parsed_questions, parse_errors = parse_master_workbook(f.stream, uploaded_by="super_admin")
        except Exception as exc:  # noqa: BLE001 — surface parse failures to the admin
            return error(f"Could not read workbook: {exc}")

        if not parsed_questions:
            return error(
                "No valid questions were found in this workbook. "
                + " ".join(parse_errors) if parse_errors else
                "No valid questions were found in this workbook."
            )

        # Persist every parsed question — never keep in memory only.
        result = questions.insert_many(parsed_questions)

        section_counts = {}
        for q in parsed_questions:
            section_counts[q["section"]] = section_counts.get(q["section"], 0) + 1

        return ok({
            "insertedCount": len(result.inserted_ids),
            "sectionCounts": section_counts,
            "warnings": parse_errors,
        }, message=f"{len(result.inserted_ids)} questions stored across {len(section_counts)} sections.")

    @bp.route("/questions/stats", methods=["GET"])
    @role_required("super_admin", "trainer")
    def question_bank_stats():
        pipeline = [
            {"$match": {"active": True}},
            {"$group": {"_id": "$section", "count": {"$sum": 1}}},
        ]
        stats = {row["_id"]: row["count"] for row in questions.aggregate(pipeline)}
        return ok({"sectionCounts": stats, "totalQuestions": sum(stats.values())})

    # ==========================================================
    # 2. ASSESSMENT CREATION — includes the Entry Level cohort option
    #    and per-section random-question-count selection.
    # ==========================================================
    @bp.route("/assessments", methods=["POST"])
    @role_required("super_admin")
    def create_assessment():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        assessment_type = (data.get("type") or "custom").strip().lower()
        cohort_target = (data.get("cohortTarget") or "").strip()
        section_counts = data.get("sectionCounts") or {}
        scheduled_at = data.get("scheduledAt")
        available_from = data.get("availableFrom")
        available_to = data.get("availableTo")

        if not name:
            return error("Assessment name is required.")
        if cohort_target not in VALID_COHORT_TARGETS:
            return error(
                "cohortTarget must be one of: Cohort A, Cohort B, Cohort C, "
                "All Cohorts, or Entry Level (send as 'A' / 'B' / 'C' / 'all' / 'entry_level')."
            )
        if not isinstance(section_counts, dict) or not section_counts:
            return error("sectionCounts is required, e.g. {'Communication': 10, 'Programming': 15}.")
        for section, count in section_counts.items():
            if not isinstance(count, int) or count <= 0:
                return error(f"Question count for section '{section}' must be a positive integer.")

        # Optional college/department scoping — both dropdowns read from the
        # SAME colleges/departments collections as Student Registration and
        # Super Admin's College Management (one source of truth, no
        # duplicate/hardcoded data). Leaving collegeId blank targets every
        # college, matching the existing "All Cohorts" style behaviour.
        college_id = data.get("collegeId")
        department_id = data.get("departmentId")
        college_name = None
        department_name = None
        if college_id:
            college_doc = resolve_active_college(db, college_id)
            if not college_doc:
                return error("Selected college is invalid or inactive.")
            college_name = college_doc["college_name"]
            if department_id:
                department_doc = resolve_active_department(db, department_id, college_id)
                if not department_doc:
                    return error("Selected department is invalid or inactive for this college.")
                department_name = department_doc["department_name"]

        # Warn (not block) if the question bank can't fully satisfy the request yet —
        # the random engine will surface the actual shortage at attempt-start time too.
        doc = {
            "name": name,
            "type": assessment_type,  # baseline | mid | final | custom
            "cohortTarget": cohort_target,  # A | B | C | all | entry_level
            "sectionCounts": {k: int(v) for k, v in section_counts.items()},
            "totalQuestions": sum(int(v) for v in section_counts.values()),
            "scheduledAt": scheduled_at,
            "availableFrom": available_from,
            "availableTo": available_to,
            "college": college_name,
            "department": department_name,
            "createdBy": "super_admin",
            "createdAt": now(),
            "status": "active",
        }
        result = assessments.insert_one(doc)
        doc["_id"] = result.inserted_id
        return ok({"assessment": serialize(doc)}, message="Assessment created.", status=201)

    @bp.route("/assessments", methods=["GET"])
    @role_required("super_admin", "trainer", "student")
    def list_assessments():
        """Feeds the Assessment Results page dropdown — every new assessment
        appears here automatically, no code change required."""
        cursor = assessments.find({}).sort("createdAt", -1)
        return ok({"assessments": [serialize(a) for a in cursor]})

    @bp.route("/assessments/<assessment_id>", methods=["GET"])
    @role_required("super_admin", "trainer")
    def get_assessment(assessment_id):
        oid = to_object_id(assessment_id)
        if not oid:
            return error("Invalid assessment id.", 404)
        doc = assessments.find_one({"_id": oid})
        if not doc:
            return error("Assessment not found.", 404)
        return ok({"assessment": serialize(doc)})

    @bp.route("/assessments/<assessment_id>", methods=["DELETE"])
    @role_required("super_admin")
    def delete_assessment(assessment_id):
        oid = to_object_id(assessment_id)
        if not oid:
            return error("Invalid assessment id.", 404)
        result = assessments.delete_one({"_id": oid})
        if result.deleted_count == 0:
            return error("Assessment not found.", 404)
        return ok(message="Assessment deleted.")

    # ==========================================================
    # 3. COHORTS — Entry Level + A/B/C, shared logic with trainer.py
    # ==========================================================
    @bp.route("/cohorts/counts", methods=["GET"])
    @role_required("super_admin")
    def get_cohort_counts():
        return ok({"cohortCounts": cohort_counts(db)})

    @bp.route("/cohorts/students", methods=["GET"])
    @role_required("super_admin")
    def list_cohort_students():
        cohort = request.args.get("cohort", "").strip()
        query = {"role": "student"}
        if cohort == ENTRY_LEVEL:
            query["$or"] = [{"cohort": None}, {"cohort": {"$exists": False}}]
        elif cohort in {"A", "B", "C"}:
            query["cohort"] = cohort
        docs = users.find(query, {"passwordHash": 0}).sort("createdAt", -1)
        return ok({"students": [serialize(d) for d in docs]})

    # ==========================================================
    # 4. QUIZ MANAGEMENT / QUIZ RESPONSES — live stats + dynamic analytics
    # ==========================================================
    @bp.route("/quiz-management/summary", methods=["GET"])
    @role_required("super_admin")
    def quiz_management_summary():
        total_questions = questions.count_documents({"active": True})
        total_assessments = assessments.count_documents({})
        total_attempts = attempts.count_documents({})
        submitted_attempts = attempts.count_documents({"status": "submitted"})
        in_progress = attempts.count_documents({"status": "in_progress"})
        return ok({
            "totalQuestions": total_questions,
            "totalAssessments": total_assessments,
            "totalAttempts": total_attempts,
            "submittedAttempts": submitted_attempts,
            "inProgressAttempts": in_progress,
            "cohortCounts": cohort_counts(db),
        })

    @bp.route("/quiz-responses", methods=["GET"])
    @role_required("super_admin")
    def quiz_responses():
        """Raw per-attempt rows for the Super Admin Quiz Responses table.

        ROOT-CAUSE FIX: this used to run its own hand-rolled aggregation
        against db.assessment_attempts/db.assessments (the older, largely
        unused cohort-placement engine) with no College/Course/Cohort/
        search filters at all — which is also why the frontend's filter
        bar was doing its own (incorrect, unrequested) client-side
        filtering on top of it. Now shares the exact same
        list_quiz_responses() helper Trainer's identically-named endpoint
        uses, just unscoped by college unless one is explicitly requested,
        so the two dashboards can never disagree about what a "response"
        is or how it's filtered.
        """
        college = request.args.get("college") or "all"
        cohort = request.args.get("cohort") or "all"
        quiz_id = request.args.get("quizId") or "all"
        department = request.args.get("department") or "all"
        search = request.args.get("search") or ""
        return ok({"responses": list_quiz_responses(
            db,
            college=None if college == "all" else college,
            cohort=cohort, quiz_id=quiz_id, department=department, search=search,
        )})

    @bp.route("/quiz-responses/filters", methods=["GET"])
    @role_required("super_admin")
    def quiz_responses_filters():
        """Course/Department dropdown options for the Quiz Responses
        filter bar — distinct values actually in use across every
        college (or one college, if ?college= is given), read live from
        the database, never hardcoded."""
        college = request.args.get("college") or "all"
        return ok({"departments": list_distinct_departments(
            db, college=None if college == "all" else college,
        )})

    @bp.route("/quiz-responses/export", methods=["GET"])
    @role_required("super_admin")
    def quiz_responses_export():
        """Backend-generated .xlsx of exactly what the Quiz Responses
        table is currently showing — same filters, same query, same rows."""
        college = request.args.get("college") or "all"
        cohort = request.args.get("cohort") or "all"
        quiz_id = request.args.get("quizId") or "all"
        department = request.args.get("department") or "all"
        search = request.args.get("search") or ""
        rows = list_quiz_responses(
            db,
            college=None if college == "all" else college,
            cohort=cohort, quiz_id=quiz_id, department=department, search=search,
            limit=100000,
        )
        buf = build_quiz_responses_workbook(rows)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="quiz_responses.xlsx",
        )

    @bp.route("/quiz-analytics", methods=["GET"])
    @role_required("super_admin")
    def quiz_analytics():
        """Dynamic analytics for the Quiz Responses page — Student
        Participation, Performance Metrics, Leaderboard (top 10) and
        Section-wise Performance. See quiz_common.compute_quiz_analytics
        for the shared computation Trainer's equivalent endpoint also
        uses, so the two dashboards can never disagree.
        """
        college = request.args.get("college") or "all"
        cohort = request.args.get("cohort") or "all"
        quiz_id = request.args.get("quizId") or "all"
        department = request.args.get("department") or "all"
        search = request.args.get("search") or ""
        return ok(compute_quiz_analytics(
            db,
            college=None if college == "all" else college,
            cohort=cohort,
            quiz_id=quiz_id,
            department=department,
            search=search,
        ))

    @bp.route("/assessments/responses", methods=["GET"])
    @role_required("super_admin")
    def assessment_responses_all():
        """Deprecated alias of /quiz-responses (kept for backward
        compatibility — nothing in this codebase calls it anymore, the
        frontend now calls /quiz-responses directly, matching Trainer's
        endpoint name). Shares the exact same filtered helper, so even if
        something external still calls this, it can never disagree with
        the main table.
        """
        college = request.args.get("college") or "all"
        cohort = request.args.get("cohort") or "all"
        quiz_id = request.args.get("quizId") or "all"
        department = request.args.get("department") or "all"
        search = request.args.get("search") or ""
        return ok({"responses": list_quiz_responses(
            db,
            college=None if college == "all" else college,
            cohort=cohort, quiz_id=quiz_id, department=department, search=search,
        )})

    # ==========================================================
    # MARKS MANAGEMENT — Create-Quiz results, Interview Verification,
    # and Validation Verification, platform-wide (no college filter).
    # Mirrors trainer.py's routes exactly so both dashboards can never
    # disagree; see the header comment in trainer.py for the full
    # rationale on why this is separate from /quiz-responses above.
    # ==========================================================
    @bp.route("/quiz-results", methods=["GET"])
    @role_required("super_admin")
    def quiz_results_all():
        """`search` matches Student Name / Roll Number / Register Number,
        resolved entirely by the database (see
        quiz_common._student_search_query) — same matching rules Trainer's
        identically-named endpoint uses."""
        search = request.args.get("search") or ""
        return ok({"results": list_quiz_results(db, search=search)})

    @bp.route("/quiz-interview-verification", methods=["GET"])
    @role_required("super_admin")
    def quiz_interview_verification_all():
        """`search` matches Assessment Name / Student Name / Roll Number /
        College / Department, resolved entirely by the database (see
        quiz_common._verification_search_query)."""
        search = request.args.get("search") or ""
        return ok({"results": list_quiz_results(db, search=search, broad_search=True)})

    @bp.route("/quiz-interview-verification/<attempt_id>/marks", methods=["POST"])
    @role_required("super_admin")
    def enter_quiz_interview_marks_admin(attempt_id):
        oid = to_object_id(attempt_id)
        if not oid:
            return error("Invalid quiz result id.", 404)
        data = request.get_json(silent=True) or {}
        updated, err = set_quiz_interview_marks(db, oid, data.get("marks"), get_jwt_identity())
        if err:
            return error(err)
        log_activity(
            db, get_jwt_identity(), "super_admin", "quiz_interview_scored",
            f'Entered interview marks for {updated.get("studentName", "a student")} '
            f'({updated.get("quizTitle", "a quiz")}) — {updated.get("interviewMarks")}%',
            college=updated.get("college"), student_id=updated.get("studentId"),
            meta={"attemptId": attempt_id, "finalAverage": updated.get("finalAverage"),
                  "cohort": updated.get("assignedCohort")},
        )
        return ok({"result": serialize_quiz_result(db, updated)}, message="Interview marks saved.")

    @bp.route("/quiz-validation-verification", methods=["GET"])
    @role_required("super_admin")
    def quiz_validation_verification_all():
        search = request.args.get("search") or ""
        return ok({"results": list_quiz_results(
            db, statuses=[RESULT_STATUS_INTERVIEW_DONE, RESULT_STATUS_VALIDATED],
            search=search, broad_search=True,
        )})

    @bp.route("/quiz-validation-verification/<attempt_id>/validate", methods=["POST"])
    @role_required("super_admin")
    def validate_quiz_result_route_admin(attempt_id):
        oid = to_object_id(attempt_id)
        if not oid:
            return error("Invalid quiz result id.", 404)
        updated, err = validate_quiz_result(db, oid, get_jwt_identity())
        if err:
            return error(err)
        log_activity(
            db, get_jwt_identity(), "super_admin", "quiz_result_validated",
            f'Validated result for {updated.get("studentName", "a student")} '
            f'({updated.get("quizTitle", "a quiz")}) — Cohort {updated.get("assignedCohort")}',
            college=updated.get("college"), student_id=updated.get("studentId"),
            meta={"attemptId": attempt_id},
        )
        return ok({"result": serialize_quiz_result(db, updated)}, message="Result validated.")

    # ==========================================================
    # PLACEMENT RULES — single source of truth for cohort thresholds.
    # Every module (student, trainer, super_admin) reads this same
    # collection via quiz_common.get_placement_rules(); nothing anywhere
    # else in the codebase hardcodes a score range or a weight.
    # ==========================================================
    @bp.route("/placement-rules", methods=["GET"])
    @role_required("super_admin", "trainer")
    def get_rules():
        return ok({"placementRules": serialize(get_placement_rules(db))})

    @bp.route("/placement-rules", methods=["PUT"])
    @role_required("super_admin")
    def update_rules():
        """
        Update the three cohort thresholds and/or the assessment/interview
        weighting. Every save is fully automatic: validate -> persist ->
        immediately recalculate every existing student's Final Employability
        Score and cohort against the brand-new rules -> those cohorts are
        visible everywhere else in the platform on their very next read,
        since every module reads db.placement_rules fresh via
        quiz_common.get_placement_rules() / cohort_from_score(). There is no
        opt-in "recalculate" flag and no background job — recalculation
        always happens as part of Save Rules.

        Body: {
          "placementReadyThreshold": 80,   // Marks >= this -> Cohort A
          "nearReadyThreshold": 50,        // Marks >= this -> Cohort B
          "highRiskThreshold": 50,         // Marks <  this -> Cohort C
          "assessmentWeight": 70,          // percentage, 0-100
          "interviewWeight": 30            // percentage, 0-100 — must sum to 100 with assessmentWeight
        }
        """
        data = request.get_json(silent=True) or {}

        def as_number(key):
            val = data.get(key)
            if val is None:
                return None
            try:
                return float(val)
            except (TypeError, ValueError):
                raise ValueError(key)

        try:
            placement_ready = as_number("placementReadyThreshold")
            near_ready = as_number("nearReadyThreshold")
            high_risk = as_number("highRiskThreshold")
            assessment_weight = as_number("assessmentWeight")
            interview_weight = as_number("interviewWeight")
        except ValueError as bad_key:
            return error(f"'{bad_key}' must be a number.")

        current = get_placement_rules(db)

        # Merge onto the currently-saved rules so a partial update (e.g.
        # weights only) still validates against real, current values.
        effective_placement_ready = placement_ready if placement_ready is not None else current.get("placementReadyThreshold", 75)
        effective_near_ready = near_ready if near_ready is not None else current.get("nearReadyThreshold", 50)
        effective_high_risk = high_risk if high_risk is not None else current.get("highRiskThreshold", 50)
        effective_assessment_weight = assessment_weight if assessment_weight is not None else current.get("assessmentWeight", 50)
        effective_interview_weight = interview_weight if interview_weight is not None else current.get("interviewWeight", 50)

        # --- Validation -------------------------------------------------
        for label, val in (
            ("Placement Ready threshold", effective_placement_ready),
            ("Near Ready threshold", effective_near_ready),
            ("High Risk threshold", effective_high_risk),
        ):
            if val < 0 or val > 100:
                return error(f"{label} must be between 0 and 100.")

        if effective_placement_ready < effective_near_ready:
            return error("Placement Ready threshold must be greater than or equal to Near Ready threshold.")

        if round(effective_assessment_weight + effective_interview_weight, 2) != 100:
            return error("Assessment Weight and Interview Weight must add up to exactly 100.")

        # High Risk and Near Ready describe the same B/C boundary from
        # opposite sides ("Near Ready starts at X" == "High Risk ends at
        # X") — keep them in sync regardless of which field the Super
        # Admin actually edited.
        if high_risk is not None and near_ready is None:
            effective_near_ready = effective_high_risk
        else:
            effective_high_risk = effective_near_ready

        update_doc = {
            "updatedAt": now(),
            "updatedBy": "super_admin",
            "placementReadyThreshold": effective_placement_ready,
            "nearReadyThreshold": effective_near_ready,
            "highRiskThreshold": effective_high_risk,
            "assessmentWeight": effective_assessment_weight,
            "interviewWeight": effective_interview_weight,
        }
        # Snapshot of every field we're about to overwrite, so a failed
        # recalculation can be rolled back to exactly what was there before
        # (used only on the no-transaction-support fallback path below —
        # the transaction path rolls back automatically on any error).
        previous_rules_snapshot = {
            k: current.get(k) for k in (
                "placementReadyThreshold", "nearReadyThreshold", "highRiskThreshold",
                "assessmentWeight", "interviewWeight", "updatedAt", "updatedBy",
            )
        }

        def apply_rules_and_recalculate(session=None):
            """Save the new rules, then recalculate EVERY existing student
            against them in a handful of bulk_write() calls (no per-student
            find_one()/update_one() loop — see compute_cohort_recalculation_ops
            docstring). Runs either inside a Mongo transaction (preferred,
            atomic) or, on deployments without transaction support, directly
            — see the try/except below for how each path is made safe."""
            db.placement_rules.update_one(
                {"_id": current["_id"]}, {"$set": update_doc}, session=session,
            )
            updated_rules = db.placement_rules.find_one({"_id": current["_id"]}, session=session)
            users_ops, student_cohort_ops, recalculated_count = compute_cohort_recalculation_ops(
                db, updated_rules,
            )
            if users_ops:
                db.users.bulk_write(users_ops, ordered=False, session=session)
            if student_cohort_ops:
                db.student_cohort.bulk_write(student_cohort_ops, ordered=False, session=session)
            return updated_rules, recalculated_count

        def rollback_rules(reason):
            """Best-effort rollback: undo the rules write so the database
            is never left with new rules but stale, un-recalculated
            cohorts. Failure to roll back is logged but never masks the
            original error returned to the client."""
            try:
                db.placement_rules.update_one(
                    {"_id": current["_id"]}, {"$set": previous_rules_snapshot},
                )
            except Exception:
                logger.exception(
                    "Placement rules rollback itself failed after: %s", reason,
                )

        # --- Save + recalculate, atomically where the deployment supports it.
        # Any failure anywhere in this flow is logged with a full traceback
        # server-side, and always turned into a clear, specific JSON error
        # response — Save Rules must never surface a bare/empty 500.
        updated = recalculated = None
        try:
            try:
                with db.client.start_session() as txn_session:
                    with txn_session.start_transaction():
                        updated, recalculated = apply_rules_and_recalculate(session=txn_session)
            except (OperationFailure, ConfigurationError, NotImplementedError) as exc:
                # Transactions require a replica set / mongos. A standalone
                # MongoDB (no replica set) rejects them outright, and some
                # drivers/test doubles don't implement sessions at all —
                # in every one of these cases, fall back to a direct,
                # non-transactional apply. Atlas / any replica-set
                # deployment never hits this branch; the transaction above
                # already gives full atomic rollback there.
                logger.warning(
                    "Placement rules: transactions unavailable on this "
                    "deployment (%s: %s); falling back to a direct, "
                    "non-transactional save + recalculate.",
                    type(exc).__name__, exc,
                )
                updated, recalculated = apply_rules_and_recalculate(session=None)
        except PyMongoError as exc:
            logger.exception("Placement rules save/recalculate failed (database error).")
            rollback_rules(exc)
            return error(
                f"Could not save placement rules: a database error occurred ({exc}). "
                "Any partial changes were rolled back — please try again.", 500,
            )
        except Exception as exc:  # noqa: BLE001 - deliberate catch-all
            # Anything unexpected (bad data shape, driver incompatibility,
            # etc.) still must not leak as a bare, message-less 500 — the
            # frontend has nothing useful to show the Super Admin otherwise.
            logger.exception("Placement rules save/recalculate failed (unexpected error).")
            rollback_rules(exc)
            return error(
                f"Could not save placement rules due to an unexpected server error: {exc}. "
                "Any partial changes were rolled back — please check server logs and try again.",
                500,
            )

        logger.info(
            "Placement rules updated by %s — %s student(s) recalculated.",
            get_jwt_identity(), recalculated,
        )
        log_activity(
            db, "super_admin", "super_admin", "placement_rules_updated",
            f"Updated placement rules — {recalculated} student(s) recalculated.",
        )
        return ok({"placementRules": serialize(updated), "recalculatedStudents": recalculated},
                   message="Placement rules updated.")

    @bp.route("/student-cohort/backfill", methods=["POST"])
    @role_required("super_admin")
    def student_cohort_backfill():
        """
        Cohort-mapping bug fix, Part 5 ("Synchronize Existing Data"): run
        this once after deploying the cohort-mapping fix to retroactively
        correct any student whose Create-Quiz result was Validated
        BEFORE that fix existed — their db.users.cohort (the field every
        quiz-eligibility check reads) never moved, even though their own
        Quiz History page correctly showed their assigned cohort all
        along. Also backfills db.student_cohort for anyone whose cohort
        was already correct but predates that collection.

        Safe to re-run any number of times — see
        quiz_common.backfill_student_cohorts()'s docstring for exactly
        why nothing here can regress an already-correct student.
        """
        result = backfill_student_cohorts(db)
        log_activity(
            db, get_jwt_identity(), "super_admin", "student_cohort_backfill",
            f"Ran cohort backfill — {result['promoted']} student(s) corrected, "
            f"{result['mirroredOnly']} mirrored into StudentCohort, "
            f"{result['checked']} candidate(s) checked.",
        )
        return ok(result, message=(
            f"Backfill complete — {result['promoted']} student(s) had their cohort corrected, "
            f"{result['mirroredOnly']} more were mirrored into the StudentCohort collection."
        ))

    # ==========================================================
    # 5. DYNAMIC CHARTS — Skill Radar, Category %, Overall, Trends
    # ==========================================================
    @bp.route("/dashboard/charts", methods=["GET"])
    @role_required("super_admin")
    def dashboard_charts():
        cohort = request.args.get("cohort")
        match = {"status": "submitted"}
        if cohort:
            pipeline_students = list(users.find(
                {"role": "student", **({"cohort": cohort} if cohort != ENTRY_LEVEL else
                                        {"$or": [{"cohort": None}, {"cohort": {"$exists": False}}]})},
                {"_id": 1},
            ))
            match["studentId"] = {"$in": [s["_id"] for s in pipeline_students]}

        # Skill Radar / Score by Category: average % per section across all submitted attempts
        section_agg = {}
        overall_scores = []
        trend_points = []
        for att in attempts.find(match).sort("submittedAt", 1):
            overall = att.get("overall", {})
            if "percentage" in overall:
                overall_scores.append(overall["percentage"])
                trend_points.append({
                    "date": iso_utc(att.get("submittedAt")),
                    "percentage": overall["percentage"],
                })
            for section, s in att.get("sectionScores", {}).items():
                bucket = section_agg.setdefault(section, {"sum": 0.0, "n": 0})
                bucket["sum"] += s.get("percentage", 0)
                bucket["n"] += 1

        skill_radar = {
            section: round(b["sum"] / b["n"], 2) if b["n"] else 0
            for section, b in section_agg.items()
        }
        overall_avg = round(sum(overall_scores) / len(overall_scores), 2) if overall_scores else 0

        return ok({
            "skillRadar": skill_radar,
            "scoreByCategory": skill_radar,
            "overallScore": overall_avg,
            "categoryPercentage": skill_radar,
            "performanceTrend": trend_points,
        })

    # ==========================================================
    # 5b. EXECUTIVE DASHBOARD — summary cards, trends, department
    #     distribution, institution ranking, active users, search
    #     and notifications. Every value below is computed live from
    #     db.users / db.colleges / db.departments / db.assessment_attempts /
    #     db.quizzes / db.activity_log — nothing here is hardcoded, and
    #     nothing is cached beyond the lifetime of a single request.
    # ==========================================================
    @bp.route("/dashboard/kpis", methods=["GET"])
    @role_required("super_admin")
    def dashboard_kpis():
        rules = get_placement_rules(db)
        ready_threshold = rules.get("placementReadyThreshold", 75)
        risk_threshold = rules.get("nearReadyThreshold", 50)
        bounds = _month_bounds(6)

        counts = cohort_counts(db)
        total_students = users.count_documents({"role": "student"})
        total_colleges = db.colleges.count_documents({"status": "active"})

        # Total Students — new-registration spark (per-month, not
        # cumulative) over the trailing 6 months.
        student_dates = [d.get("createdAt") for d in users.find(
            {"role": "student"}, {"createdAt": 1})]
        student_spark = _bucket_dates(student_dates, bounds)

        # Total Colleges — new-college spark.
        college_dates = [d.get("created_at") for d in db.colleges.find(
            {}, {"created_at": 1})]
        college_spark = _bucket_dates(college_dates, bounds)

        # Placement Ready / At-Risk — monthly count of DISTINCT students
        # whose latest submitted attempt that month cleared/missed the
        # current threshold. Real signal, not a snapshot replay of the
        # live cohort count (which has no history to chart).
        ready_spark = [0] * 6
        risk_spark = [0] * 6
        for idx, (start, end, _label) in enumerate(bounds):
            month_attempts = attempts.find(
                {"status": "submitted", "submittedAt": {"$gte": start, "$lt": end}},
                {"studentId": 1, "overall.percentage": 1},
            )
            best_pct = {}
            for a in month_attempts:
                sid = a.get("studentId")
                pct = (a.get("overall") or {}).get("percentage")
                if sid is None or pct is None:
                    continue
                if sid not in best_pct or pct > best_pct[sid]:
                    best_pct[sid] = pct
            ready_spark[idx] = sum(1 for p in best_pct.values() if p >= ready_threshold)
            risk_spark[idx] = sum(1 for p in best_pct.values() if p < risk_threshold)

        kpis = [
            {
                "label": "Total Students", "value": str(total_students), "icon": "users",
                "tone": "teal", "trend": _pct_trend(student_spark), "spark": student_spark,
            },
            {
                "label": "Placement Ready Students", "value": str(counts.get("A", 0)),
                "icon": "shieldCheck", "tone": "amber",
                "trend": _pct_trend(ready_spark), "spark": ready_spark,
            },
            {
                "label": "At-Risk Students", "value": str(counts.get("C", 0)),
                "icon": "alert", "tone": "rose",
                "trend": _pct_trend(risk_spark), "spark": risk_spark,
            },
            {
                "label": "Total Colleges", "value": str(total_colleges), "icon": "building",
                "tone": "ink", "trend": _pct_trend(college_spark), "spark": college_spark,
            },
        ]
        return ok({"kpis": kpis})

    @bp.route("/dashboard/department-distribution", methods=["GET"])
    @role_required("super_admin")
    def dashboard_department_distribution():
        """Live student headcount per department, for the redesigned
        'Student Distribution by Department' bar chart."""
        pipeline = [
            {"$match": {"role": "student", "department": {"$nin": [None, ""]}}},
            {"$group": {"_id": "$department", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10},
        ]
        rows = list(users.aggregate(pipeline))
        distribution = [{"label": r["_id"], "value": r["count"]} for r in rows]
        return ok({"distribution": distribution})

    @bp.route("/dashboard/placement-trend", methods=["GET"])
    @role_required("super_admin")
    def dashboard_placement_trend():
        """Month-wise placement-readiness trend: % of that month's
        submitted attempts (best attempt per student) that met the
        current Placement Ready threshold."""
        rules = get_placement_rules(db)
        ready_threshold = rules.get("placementReadyThreshold", 75)
        bounds = _month_bounds(6)
        readiness = []
        for (start, end, _label) in bounds:
            month_attempts = attempts.find(
                {"status": "submitted", "submittedAt": {"$gte": start, "$lt": end}},
                {"studentId": 1, "overall.percentage": 1},
            )
            best_pct = {}
            for a in month_attempts:
                sid = a.get("studentId")
                pct = (a.get("overall") or {}).get("percentage")
                if sid is None or pct is None:
                    continue
                if sid not in best_pct or pct > best_pct[sid]:
                    best_pct[sid] = pct
            if best_pct:
                ready = sum(1 for p in best_pct.values() if p >= ready_threshold)
                readiness.append(round(ready / len(best_pct) * 100, 1))
            else:
                readiness.append(0)
        months = [label for (_s, _e, label) in bounds]
        return ok({"months": months, "trend": {"baseline": readiness}})

    @bp.route("/dashboard/assessment-progress", methods=["GET"])
    @role_required("super_admin")
    def dashboard_assessment_progress():
        """Monthly count of submitted assessment attempts (completions),
        for the Assessment Progress chart."""
        bounds = _month_bounds(6)
        dates = [d.get("submittedAt") for d in attempts.find(
            {"status": "submitted"}, {"submittedAt": 1})]
        counts = _bucket_dates(dates, bounds)
        months = [label for (_s, _e, label) in bounds]
        return ok({"months": months, "progress": counts})

    @bp.route("/dashboard/institution-ranking", methods=["GET"])
    @role_required("super_admin")
    def dashboard_institution_ranking():
        """Top 5 institutions ranked by average student assessment score
        (average of each student's best submitted attempt, averaged
        across the college)."""
        pipeline = [
            {"$match": {"status": "submitted", "college": {"$nin": [None, ""]}}},
            {"$group": {
                "_id": {"college": "$college", "studentId": "$studentId"},
                "best": {"$max": "$overall.percentage"},
            }},
            {"$group": {
                "_id": "$_id.college",
                "avgScore": {"$avg": "$best"},
                "students": {"$sum": 1},
            }},
            {"$sort": {"avgScore": -1}},
            {"$limit": 5},
        ]
        rows = list(attempts.aggregate(pipeline))
        ranking = [
            {
                "rank": i + 1,
                "institution": r["_id"],
                "averageScore": round(r["avgScore"], 2) if r["avgScore"] is not None else 0,
                "students": r["students"],
            }
            for i, r in enumerate(rows)
        ]
        return ok({"ranking": ranking})

    @bp.route("/dashboard/live-stats", methods=["GET"])
    @role_required("super_admin")
    def dashboard_live_stats():
        """Currently-active users. Two conditions, both enforced by login.py:
          1. The account holds a live, un-superseded session
             (db.users.currentSessionId is non-null — cleared on logout and
             replaced on login elsewhere), AND
          2. It has made an authenticated request within the last 5 minutes
             (db.users.lastActiveAt, touched on every valid token check).

        The second condition is what makes users drop off automatically after
        session expiry / idle timeout instead of lingering forever, while a
        clean logout removes them instantly via condition 1."""
        cutoff = datetime.utcnow() - timedelta(minutes=5)

        def active_count(role):
            return users.count_documents({
                "role": role,
                "currentSessionId": {"$ne": None},
                "lastActiveAt": {"$gte": cutoff},
            })

        return ok({
            "studentsOnline": active_count("student"),
            "trainersOnline": active_count("trainer"),
            "adminsOnline": active_count("college_admin"),
        })

    # ==========================================================
    # 5c. COLLEGE ANALYTICS + STUDENT DIRECTORY — real, DB-backed datasets
    #     for the College Management grid and the College Details /
    #     Student Profile pages (which previously pointed at nonexistent
    #     /colleges and /students endpoints and always rendered empty).
    # ==========================================================
    @bp.route("/college-analytics", methods=["GET"])
    @role_required("super_admin")
    def college_analytics():
        """Per-college aggregate: department count, live student headcount,
        trainer count, placement % (share of college students whose best
        submitted attempt met the current Placement Ready threshold) and
        assessment completion % (share of college students with at least
        one submitted attempt), plus the college admin's contact details."""
        rules = get_placement_rules(db)
        ready_threshold = rules.get("placementReadyThreshold", 75)

        dept_counts = {}
        for d in db.departments.find({}, {"college_id": 1}):
            cid = d.get("college_id")
            if cid is not None:
                dept_counts[cid] = dept_counts.get(cid, 0) + 1

        best_by_student = {}
        # db.quiz_attempts (NOT the legacy db.assessment_attempts engine) is
        # where quiz submissions actually land — see trainer.py's Quiz
        # Responses fix. Reading here keeps College Management's Placement %
        # and Assessment Completion in sync with every other dashboard.
        for a in db.quiz_attempts.find(
            {"status": "submitted", "studentId": {"$exists": True}},
            {"studentId": 1, "overall.percentage": 1},
        ):
            sid = a.get("studentId")
            pct = (a.get("overall") or {}).get("percentage")
            if sid is None or pct is None:
                continue
            if sid not in best_by_student or pct > best_by_student[sid]:
                best_by_student[sid] = pct

        college_students = {}
        college_attempted = {}
        college_ready = {}
        for s in users.find(
            {"role": "student", "collegeId": {"$nin": [None, ""]}},
            {"_id": 1, "collegeId": 1},
        ):
            cid = s["collegeId"]
            college_students[cid] = college_students.get(cid, 0) + 1
            pct = best_by_student.get(s["_id"])
            if pct is not None:
                college_attempted[cid] = college_attempted.get(cid, 0) + 1
                if pct >= ready_threshold:
                    college_ready[cid] = college_ready.get(cid, 0) + 1

        trainer_counts = {}
        for t in users.find(
            {"role": "trainer", "collegeId": {"$nin": [None, ""]}},
            {"collegeId": 1},
        ):
            cid = t.get("collegeId")
            if cid is not None:
                trainer_counts[cid] = trainer_counts.get(cid, 0) + 1

        admin_by_college = {}
        for a in users.find(
            {"role": "college_admin", "collegeId": {"$nin": [None, ""]}},
            {"collegeId": 1, "fullName": 1, "email": 1, "phone": 1, "lastLoginAt": 1},
        ):
            admin_by_college[a["collegeId"]] = a

        rows = []
        for c in db.colleges.find({}).sort("college_name", 1):
            cid = c["_id"]
            n = college_students.get(cid, 0)
            attempted = college_attempted.get(cid, 0)
            ready = college_ready.get(cid, 0)
            admin = admin_by_college.get(cid)
            rows.append({
                "id": str(cid),
                "name": c.get("college_name"),
                "status": c.get("status", "active"),
                "createdAt": c["created_at"].isoformat() if c.get("created_at") else None,
                "depts": dept_counts.get(cid, 0),
                "students": n,
                "trainers": trainer_counts.get(cid, 0),
                "placement": round(ready / n * 100, 1) if n else 0,
                "assessment": round(attempted / n * 100, 1) if n else 0,
                "adminName": (admin.get("fullName") if admin else None) or "—",
                "email": admin.get("email") if admin else None,
                "phone": admin.get("phone") if admin else None,
                "lastLogin": iso_utc(admin.get("lastLoginAt")) if admin and admin.get("lastLoginAt") else None,
            })
        return ok({"colleges": rows, "total": len(rows)})

    @bp.route("/colleges/export", methods=["GET"])
    @role_required("super_admin")
    def admin_college_directory_export():
        """College Management → Export. Real college directory data: name,
        status, department count, student headcount and the college admin's
        contact details. Served as PDF or Excel (mirrors the other report
        exports)."""
        rules = get_placement_rules(db)
        ready_threshold = rules.get("placementReadyThreshold", 75)

        dept_counts = {}
        for d in db.departments.find({}, {"college_id": 1}):
            cid = d.get("college_id")
            if cid is not None:
                dept_counts[cid] = dept_counts.get(cid, 0) + 1

        student_counts = {}
        for s in users.find(
            {"role": "student", "collegeId": {"$nin": [None, ""]}}, {"collegeId": 1}
        ):
            cid = s.get("collegeId")
            if cid is not None:
                student_counts[cid] = student_counts.get(cid, 0) + 1

        admin_by_college = {}
        for a in users.find(
            {"role": "college_admin", "collegeId": {"$nin": [None, ""]}},
            {"collegeId": 1, "fullName": 1, "email": 1, "phone": 1},
        ):
            cid = a.get("collegeId")
            if cid is not None:
                admin_by_college[cid] = a

        rows = []
        for c in db.colleges.find({}).sort("college_name", 1):
            cid = c["_id"]
            admin = admin_by_college.get(cid)
            rows.append([
                c.get("college_name") or "—",
                (c.get("status") or "active").capitalize(),
                dept_counts.get(cid, 0),
                student_counts.get(cid, 0),
                (admin.get("fullName") if admin else None) or "—",
                (admin.get("email") if admin else None) or "—",
                (admin.get("phone") if admin else None) or "—",
            ])

        columns = ["College", "Status", "Departments", "Students", "Admin", "Email", "Phone"]
        fmt = (request.args.get("format") or "pdf").lower()
        subtitle = f"{len(rows)} college(s) on the platform · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}"
        if fmt == "excel":
            return send_file(
                _excel_bytes(columns, rows, "College Directory"),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True, download_name="college_directory.xlsx",
            )
        return send_file(
            _pdf_bytes("College Directory", subtitle, columns, rows),
            mimetype="application/pdf", as_attachment=True,
            download_name="college_directory.pdf",
        )

    @bp.route("/students", methods=["GET"])
    @role_required("super_admin")
    def list_all_students():
        """Every non-deleted student, platform-wide, with their best
        submitted attempt % (overall). Powers the College Details student
        table and the Student Profile page."""
        best = {}
        # Same live-collection fix as college_analytics(): quiz submissions
        # are stored in db.quiz_attempts, so the Student Directory's Overall
        # column must read from there or it always renders blank/0.
        for a in db.quiz_attempts.find(
            {"status": "submitted", "studentId": {"$exists": True}},
            {"studentId": 1, "overall.percentage": 1},
        ):
            sid = a.get("studentId")
            pct = (a.get("overall") or {}).get("percentage")
            if sid is None or pct is None:
                continue
            if sid not in best or pct > best[sid]:
                best[sid] = pct

        docs = list(users.find(
            {"role": "student", "isDeleted": {"$ne": True}},
            {
                "fullName": 1, "email": 1, "mobile": 1, "rollNumber": 1,
                "tneaCode": 1, "college": 1, "department": 1,
                "cohort": 1, "approvalStatus": 1,
            },
        ).sort("fullName", 1))

        rows = []
        for s in docs:
            cohort = student_cohort_label(s)
            overall = best.get(s["_id"])
            rows.append({
                "id": str(s["_id"]),
                "name": s.get("fullName") or "—",
                "email": s.get("email"),
                "phone": s.get("mobile"),
                "reg": s.get("rollNumber") or s.get("tneaCode") or "—",
                "dept": s.get("department") or "—",
                "college": s.get("college") or "—",
                "overall": round(overall, 1) if overall is not None else None,
                "cohort": cohort,
                "cohortLabel": f"Cohort {cohort}" if cohort in ("A", "B", "C") else "Entry Level",
                "status": "active" if s.get("approvalStatus") == "approved" else (s.get("approvalStatus") or "pending"),
            })
        return ok({"students": rows, "total": len(rows)})

    # ==========================================================
    # 5d. EDIT STUDENT PROFILE — the Super Admin's "Edit Profile"
    #     button on the Student Profile page (spec §6). Edits the
    #     student's account-level contact details (name/email/mobile).
    #     Email/mobile are validated with the same rules registration
    #     uses (login.py); email is unique across the platform and the
    #     updated value is lower-cased so login lookups stay consistent.
    # ==========================================================
    @bp.route("/students/<student_id>", methods=["PUT"])
    @role_required("super_admin")
    def admin_update_student(student_id):
        oid = to_object_id(student_id)
        if not oid:
            return error("Invalid student id.")
        student = users.find_one({"_id": oid, "role": "student"})
        if not student:
            return error("Student not found.", 404)
        if student.get("isDeleted"):
            return error("This account has been deleted and can no longer be modified.")

        data = request.get_json(silent=True) or {}
        full_name = (data.get("fullName") or "").strip()
        email = (data.get("email") or "").strip().lower()
        mobile = (data.get("mobile") or "").strip()

        if full_name and len(full_name) > 200:
            return error("Name is too long.")
        if not _EMAIL_RE.match(email):
            return error("A valid email address is required.")
        if not _MOBILE_RE.match(mobile):
            return error("Mobile number must be exactly 10 digits.")

        other = users.find_one({"email": email, "_id": {"$ne": oid}})
        if other:
            return error("An account with this email already exists.")

        users.update_one({"_id": oid}, {"$set": {
            "fullName": full_name or student.get("fullName"),
            "email": email,
            "mobile": mobile,
            "updatedAt": now(),
        }})
        log_activity(db, "super_admin", "super_admin", "student_profile_updated",
                     f"Updated profile for {student.get('fullName') or student.get('email')}")
        updated = users.find_one({"_id": oid})
        return ok({
            "student": {
                "id": str(updated["_id"]),
                "name": updated.get("fullName") or "—",
                "email": updated.get("email"),
                "phone": updated.get("mobile"),
                "rollNumber": updated.get("rollNumber") or updated.get("tneaCode") or "—",
                "college": updated.get("college") or "—",
                "department": updated.get("department") or "—",
            },
        }, message="Student profile updated.")

    # ==========================================================
    # 5d. BULK CSV STUDENT IMPORT — the "Bulk CSV" button on the
    #     Student Profile page (spec §6). Accepts a JSON array of rows
    #     (parsed client-side from a CSV file) with columns
    #     fullName / email / mobile / rollNumber / college / department.
    #     College and department are resolved by name against the real
    #     db.colleges / db.departments collections (active only); email
    #     and roll number must be unique platform-wide. New students are
    #     created approved with a generated temporary password, then the
    #     admin is shown the temporary passwords for the created rows.
    # ==========================================================
    @bp.route("/students/bulk-import", methods=["POST"])
    @role_required("super_admin")
    def bulk_import_students():
        data = request.get_json(silent=True) or {}
        rows = data.get("rows") or []
        if not isinstance(rows, list) or not rows:
            return error("No student rows provided.")
        if len(rows) > 500:
            return error("Maximum 500 students per import.")

        colleges_by_name = {}
        departments_by_college = {}
        for c in db.colleges.find({"status": "active"}, {"college_name": 1}):
            colleges_by_name[str(c.get("college_name") or "").strip().lower()] = c
        for d in db.departments.find({"status": "active"}, {"department_name": 1, "collegeId": 1}):
            key = (d.get("collegeId"), str(d.get("department_name") or "").strip().lower())
            departments_by_college[key] = d

        existing_emails = {
            u.get("email") for u in users.find(
                {"email": {"$exists": True}}, {"email": 1})
        }
        existing_rolls = {
            u.get("rollNumber") for u in users.find(
                {"role": "student", "rollNumber": {"$exists": True, "$ne": None}},
                {"rollNumber": 1})
        }

        if bcrypt is None:
            return error("Bulk import is not available right now.", 503)

        created, failed = [], []
        for i, row in enumerate(rows, start=2):
            if not isinstance(row, dict):
                failed.append({"row": i, "reason": "Row is not a valid object."})
                continue
            full_name = str(row.get("fullName") or row.get("name") or "").strip()
            email = str(row.get("email") or "").strip().lower()
            mobile = str(row.get("mobile") or row.get("phone") or "").strip()
            roll_number = str(row.get("rollNumber") or row.get("reg") or "").strip()
            college_name = str(row.get("college") or "").strip()
            department_name = str(row.get("department") or "").strip()

            reason = None
            if not full_name:
                reason = "Name is required."
            elif not _EMAIL_RE.match(email):
                reason = "Invalid email address."
            elif not _MOBILE_RE.match(mobile):
                reason = "Mobile must be exactly 10 digits."
            elif not roll_number:
                reason = "Roll number is required."
            elif email in existing_emails:
                reason = "Email already exists."
            elif roll_number in existing_rolls:
                reason = "Roll number already exists."
            elif not college_name:
                reason = "College is required."
            else:
                college_doc = colleges_by_name.get(college_name.lower())
                if not college_doc:
                    reason = f"Unknown college: {college_name}"
                else:
                    department_doc = None
                    if department_name:
                        department_doc = departments_by_college.get(
                            (college_doc["_id"], department_name.lower()))
                        if not department_doc:
                            reason = f"Unknown department in this college: {department_name}"
            if reason:
                failed.append({"row": i, "reason": reason})
                continue

            temp_password = _generate_temp_password()
            password_hash = bcrypt.generate_password_hash(temp_password).decode("utf-8")
            user_doc = {
                "fullName": full_name,
                "email": email,
                "mobile": mobile,
                "role": "student",
                "passwordHash": password_hash,
                "approvalStatus": "approved",
                "approvedBy": "super_admin",
                "approvedDate": now(),
                "googleLogin": False,
                "isDeleted": False,
                "cohort": None,
                "baselineAssessmentScore": None,
                "interviewScore": None,
                "finalEmployabilityScore": None,
                "cohortAssignedAt": None,
                "rollNumber": roll_number,
                "tneaCode": None,
                "college": college_doc["college_name"],
                "collegeId": college_doc["_id"],
                "department": department_doc["department_name"] if department_doc else None,
                "departmentId": department_doc["_id"] if department_doc else None,
                "district": None,
                "employeeId": None,
                "createdAt": now(),
                "updatedAt": now(),
            }
            try:
                result = users.insert_one(user_doc)
            except PyMongoError:
                failed.append({"row": i, "reason": "Could not create this account."})
                continue
            existing_emails.add(email)
            existing_rolls.add(roll_number)
            created.append({
                "id": str(result.inserted_id),
                "fullName": full_name,
                "email": email,
                "mobile": mobile,
                "rollNumber": roll_number,
                "college": college_doc["college_name"],
                "department": department_doc["department_name"] if department_doc else None,
                "temporaryPassword": temp_password,
            })

        log_activity(db, "super_admin", "super_admin", "students_bulk_imported",
                     f"Bulk imported {len(created)} student(s), {len(failed)} failed")
        return ok({
            "created": created,
            "failed": failed,
            "createdCount": len(created),
            "failedCount": len(failed),
        }, message=f"Imported {len(created)} student(s), {len(failed)} failed.")

    # ==========================================================
    # 5e. STUDENT DETAIL — attendance, assessment summary and the
    #     overall report for ONE student, platform-wide. These mirror
    #     the College Admin student-profile endpoints (collegeadmin.py)
    #     WITHOUT the college-scope restriction, so the Super Admin sees
    #     every submitted attempt / attendance record for that student
    #     no matter which college they belong to. All three read live
    #     from db.quiz_attempts / db.attendance / db.quiz_sections —
    #     nothing hardcoded.
    # ==========================================================
    @bp.route("/students/<student_id>/attendance", methods=["GET"])
    @role_required("super_admin")
    def admin_student_attendance(student_id):
        """Full live attendance for one student — summary counts +
        newest-first history, both from the shared attendance_summary()
        helper (same data the student sees on their own dashboard and
        the College Admin sees on their student profile)."""
        oid = to_object_id(student_id)
        if not oid:
            return error("Invalid student id.")
        student = users.find_one({"_id": oid, "role": "student"})
        if not student:
            return error("Student not found.", 404)
        return ok({
            "student": {
                "id": str(student["_id"]),
                "name": student.get("fullName") or student.get("name"),
                "rollNumber": student.get("rollNumber") or student.get("tneaCode"),
                "cohort": student_cohort_label(student),
            },
            "attendance": attendance_summary(db, oid),
        })

    @bp.route("/students/<student_id>/assessment-summary", methods=["GET"])
    @role_required("super_admin")
    def admin_student_assessment_summary(student_id):
        """The six assessment parameters for one student — the five
        DB-backed quiz sections (db.quiz_sections, same source the Create
        Quiz wizard reads) plus the Manual Interview score. Every value is
        computed live from the student's submitted quiz attempts in
        db.quiz_attempts, exactly like the College Admin version."""
        oid = to_object_id(student_id)
        if not oid:
            return error("Invalid student id.")
        student = users.find_one({"_id": oid, "role": "student"})
        if not student:
            return error("Student not found.", 404)

        sections_docs = list(db.quiz_sections.find({}).sort("name", 1))
        attempts = list(db.quiz_attempts.find({
            "studentId": oid,
            "status": "submitted",
        }).sort("submittedAt", -1))

        # Per-section average % across this student's attempts. Pooled
        # questions resolve against the quiz's own question list (same
        # logic as compute_quiz_analytics) so the numbers come straight
        # from the DB every time.
        quizzes_cache = {}
        sec_sums = {}
        sec_counts = {}
        interview_scores = []
        overall_scores = []
        for attempt in attempts:
            interview = attempt.get("interviewMarks")
            if isinstance(interview, (int, float)) and not isinstance(interview, bool):
                interview_scores.append(float(interview))
            overall_pct = (attempt.get("overall") or {}).get("percentage")
            if isinstance(overall_pct, (int, float)) and not isinstance(overall_pct, bool):
                overall_scores.append(float(overall_pct))
            qid = attempt.get("quizId")
            if qid not in quizzes_cache:
                quizzes_cache[qid] = db.quizzes.find_one({"_id": qid}) or {}
            quiz = quizzes_cache[qid]
            pool = quiz.get("questions") or []
            answers = attempt.get("answers") or {}
            totals = {}
            correct = {}
            for pq in (attempt.get("pooledQuestions") or []):
                idx = pq.get("poolIndex")
                if idx is None or idx >= len(pool):
                    continue
                q = pool[idx]
                section = q.get("section") or "General"
                totals[section] = totals.get(section, 0) + 1
                raw = answers.get(str(idx))
                given = raw if isinstance(raw, list) else ([raw] if raw not in (None, "") else [])
                try:
                    given_idx = sorted({int(x) for x in given})
                except (TypeError, ValueError):
                    given_idx = []
                correct_idx = sorted(set(q.get("correct") or []))
                if given_idx and given_idx == correct_idx:
                    correct[section] = correct.get(section, 0) + 1
            for section, n in totals.items():
                pct = round((correct.get(section, 0) / n) * 100, 2) if n else 0.0
                sec_sums[section] = sec_sums.get(section, 0) + pct
                sec_counts[section] = sec_counts.get(section, 0) + 1

        section_rows = []
        for sec in sections_docs:
            name = sec.get("name") or "—"
            n = sec_counts.get(name, 0)
            score = round(sec_sums.get(name, 0) / n, 2) if n else None
            section_rows.append({"name": name, "score": score, "attempts": n})

        interview_score = round(sum(interview_scores) / len(interview_scores), 2) if interview_scores else None
        overall_quiz = round(sum(overall_scores) / len(overall_scores), 2) if overall_scores else None

        return ok({
            "student": {
                "id": str(student["_id"]),
                "name": student.get("fullName") or student.get("name"),
                "rollNumber": student.get("rollNumber") or student.get("tneaCode"),
                "cohort": student_cohort_label(student),
            },
            "sections": section_rows,
            "manualInterview": {
                "score": interview_score,
                "attempts": len(interview_scores),
            },
            "overallQuizScore": overall_quiz,
            "attemptsCount": len(attempts),
        })

    def _overall_report_rows(student_oid=None, quiz_ids=None, college=None, colleges=None):
        """Shared row builder for the Overall Report JSON + all of
        its PDF/Excel exports. Every row is computed live from
        db.quiz_attempts + db.users.finalEmployabilityScore.
        Optional quiz_ids (list of assessment ObjectIds), college
        (exact college name) and colleges (list of exact college names)
        scope the query so the Assessment Report export never has to
        load the whole platform attempt set."""
        query = {"status": "submitted"}
        if student_oid:
            query["studentId"] = student_oid
        if quiz_ids:
            query["quizId"] = {"$in": list(quiz_ids)}
        if college:
            query["college"] = college
        if colleges:
            query["college"] = {"$in": list(colleges)}
        attempts = list(db.quiz_attempts.find(query).sort("submittedAt", -1))
        user_scores = {}
        if student_oid:
            stu = users.find_one({"_id": student_oid}, {"finalEmployabilityScore": 1})
            if stu:
                user_scores[student_oid] = stu.get("finalEmployabilityScore")
        else:
            ids = {a.get("studentId") for a in attempts if a.get("studentId")}
            for s in users.find({"_id": {"$in": list(ids)}}, {"finalEmployabilityScore": 1}):
                user_scores[s["_id"]] = s.get("finalEmployabilityScore")
        rows = []
        for a in attempts:
            overall = a.get("overall") or {}
            rows.append({
                "attemptId": str(a["_id"]),
                "studentId": str(a["studentId"]) if a.get("studentId") else None,
                "studentName": a.get("studentName"),
                "rollNumber": a.get("studentRollNumber"),
                "department": a.get("department"),
                "college": a.get("college"),
                "assessmentId": str(a["quizId"]) if a.get("quizId") else None,
                "assessmentName": a.get("quizTitle") or "Quiz",
                "quizMarks": overall.get("percentage"),
                "quizMarksObtained": overall.get("marksObtained") or overall.get("obtainedMarks"),
                "quizTotalMarks": overall.get("totalMarks") or overall.get("total"),
                "passFail": overall.get("passFail"),
                "interviewMarks": a.get("interviewMarks"),
                "averageMarks": a.get("finalAverage"),
                "finalOverallMarks": user_scores.get(a.get("studentId")) if a.get("studentId") else None,
                "submittedAt": iso_utc(a.get("submittedAt")),
            })
        return rows

    @bp.route("/reports/overall", methods=["GET"])
    @role_required("super_admin")
    def admin_overall_report():
        """The single Overall Report for the Super Admin — every submitted
        quiz attempt, platform-wide, or for one student when studentId is
        passed (Student Profile → Reports tab). Shows Assessment Name, Quiz
        Marks, Interview Marks, Average Marks and Final Overall Marks."""
        student_id = (request.args.get("studentId") or "").strip()
        oid = to_object_id(student_id) if student_id else None
        rows = _overall_report_rows(oid)
        return ok({"report": rows, "studentId": student_id})

    # ==========================================================
    # REPORT EXPORTS — PDF (reportlab) and Excel (openpyxl), both
    # generated server-side from the same live queries the JSON
    # endpoints use, so the downloaded files always contain real
    # database data and proper MIME types/filenames.
    # ==========================================================
    def _excel_bytes(columns, rows, sheet_name):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        ws.append(columns)
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F9E93")
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(["" if v is None else v for v in r])
        for i, col_name in enumerate(columns, start=1):
            longest = len(str(col_name))
            for r in rows:
                longest = max(longest, len(str(r[i - 1])) if r[i - 1] is not None else 0)
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(longest + 2, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _pdf_col_widths(doc, columns, data):
        """Allocate the table's full horizontal width (landscape page)
        to columns in proportion to their content, so text-heavy columns
        get more room while narrow ones stay compact. Guarantees the
        widths sum to exactly the usable width; long values are wrapped
        by the Paragraph cells, never clipped or overlapped."""
        usable = doc.width
        n = len(columns)
        if not n:
            return []
        raw = []
        for ci in range(n):
            longest = len(str(columns[ci]))
            for row in data[1:]:
                val = row[ci] if ci < len(row) else ""
                longest = max(longest, len(str(val)))
            raw.append(max(6, min(longest, 70)))
        total = sum(raw) or 1
        return [usable * (w / total) for w in raw]

    def _pdf_bytes(title, subtitle, columns, rows):
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("RTitle", parent=styles["Title"], fontSize=15, spaceAfter=4)
        sub_style = ParagraphStyle(
            "RSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=12)
        cell_style = ParagraphStyle(
            "RCell", parent=styles["Normal"], fontSize=7.5, leading=9.5)
        head_style = ParagraphStyle(
            "RHead", parent=cell_style, fontName="Helvetica-Bold",
            textColor=colors.white, alignment=TA_CENTER)
        data = [columns] + [["" if v is None else str(v) for v in row] for row in rows]
        col_widths = _pdf_col_widths(doc, columns, data)
        body = [[Paragraph(c, head_style) for c in data[0]]]
        body.extend([Paragraph(c, cell_style) for c in row] for row in data[1:])
        table = Table(body, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F9E93")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEE2")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8CDB8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story = [Paragraph(title, title_style), Paragraph(subtitle, sub_style), table]
        doc.build(story)
        buf.seek(0)
        return buf

    def _fmt(v):
        return round(v, 1) if isinstance(v, (int, float)) else v

    def _pdf_sections(title, subtitle, sections):
        """PDF with multiple labelled table sections (profile + performance +
        interventions + activity in the student report). Each section is a
        (heading, columns, rows) tuple rendered as its own table, on a
        landscape page so wide tables use the full horizontal width. Cells
        are Paragraphs so long text wraps instead of colliding; headings
        keep with their table; headers repeat across page breaks."""
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("RTitle", parent=styles["Title"], fontSize=15, spaceAfter=4)
        sub_style = ParagraphStyle(
            "RSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=10)
        head_style = ParagraphStyle(
            "RSec", parent=styles["Heading2"], fontSize=11, spaceBefore=12, spaceAfter=6,
            textColor=colors.HexColor("#1F9E93"), keepWithNext=1)
        cell_style = ParagraphStyle(
            "RCell", parent=styles["Normal"], fontSize=7.5, leading=9.5)
        cell_head_style = ParagraphStyle(
            "RHead", parent=cell_style, fontName="Helvetica-Bold",
            textColor=colors.white, alignment=TA_CENTER)
        story = [Paragraph(title, title_style), Paragraph(subtitle, sub_style)]
        for heading, columns, rows in sections:
            story.append(Paragraph(heading, head_style))
            data = [columns] + [["" if v is None else str(v) for v in row] for row in rows]
            col_widths = _pdf_col_widths(doc, columns, data)
            body = [[Paragraph(c, cell_head_style) for c in data[0]]]
            body.extend([Paragraph(c, cell_style) for c in row] for row in data[1:])
            table = Table(body, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F9E93")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 7.5),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEE2")]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8CDB8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(table)
        doc.build(story)
        buf.seek(0)
        return buf

    def _attendance_counts_for_date(date_str):
        """Present / absent / total attendance records for one date from
        db.attendance — used to size participation on intervention sessions."""
        if not date_str:
            return {"total": 0, "present": 0, "absent": 0}
        present = db.attendance.count_documents({"date": date_str, "status": "present"})
        absent = db.attendance.count_documents({"date": date_str, "status": "absent"})
        return {"total": present + absent, "present": present, "absent": absent}

    @bp.route("/reports/overall/export", methods=["GET"])
    @role_required("super_admin")
    def admin_overall_report_export():
        fmt = (request.args.get("format") or "pdf").lower()
        rows = _overall_report_rows()
        columns = ["Student", "Register No", "Department", "College", "Assessment",
                   "Quiz %", "Interview %", "Average %", "Final Overall %", "Submitted"]
        data = [[r["studentName"] or "—", r["rollNumber"] or "—", r["department"] or "—",
                 r["college"] or "—", r["assessmentName"] or "—", _fmt(r["quizMarks"]),
                 _fmt(r["interviewMarks"]), _fmt(r["averageMarks"]), _fmt(r["finalOverallMarks"]),
                 (r["submittedAt"] or "")[:16].replace("T", " ")] for r in rows]
        subtitle = f"{len(rows)} submitted attempt(s) across the platform · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}"
        if fmt == "excel":
            return send_file(
                _excel_bytes(columns, data, "Overall Report"),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True, download_name="overall_platform_report.xlsx",
            )
        return send_file(
            _pdf_bytes("Overall Platform Report", subtitle, columns, data),
            mimetype="application/pdf", as_attachment=True,
            download_name="overall_platform_report.pdf",
        )

    def _college_report_payload(college_doc):
        """Shared builder for the College Report and the multi-college
        report export. Real data: the college's own record, its department
        list, and its students with best Overall score and cohort."""
        oid = college_doc["_id"]
        college_name = college_doc.get("college_name") or "College"
        departments = list(db.departments.find(
            {"college_id": oid, "status": "active"}).sort("department_name", 1))
        dept_names = [d.get("department_name") for d in departments if d.get("department_name")]
        dept_line = f"Departments ({len(departments)}):  {', '.join(dept_names) or '—'}"
        college_columns = ["Student", "Register No", "Department", "Cohort", "Overall %"]

        best = {}
        for a in db.quiz_attempts.find(
            {"status": "submitted", "college": college_name, "studentId": {"$exists": True}},
            {"studentId": 1, "overall.percentage": 1},
        ):
            sid = a.get("studentId")
            pct = (a.get("overall") or {}).get("percentage")
            if sid is None or pct is None:
                continue
            if sid not in best or pct > best[sid]:
                best[sid] = pct

        students = list(users.find(
            {"role": "student", "college": college_name, "isDeleted": {"$ne": True}},
            {"fullName": 1, "rollNumber": 1, "tneaCode": 1, "department": 1, "cohort": 1},
        ).sort("fullName", 1))
        student_rows = [[s.get("fullName") or "—",
                         s.get("rollNumber") or s.get("tneaCode") or "—",
                         s.get("department") or "—",
                         student_cohort_label(s),
                         _fmt(best.get(s["_id"])) if best.get(s["_id"]) is not None else "—"]
                        for s in students]

        subtitle = (f"{college_name} · {len(departments)} department(s) · "
                    f"{len(students)} student(s) · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")
        return {
            "name": college_name,
            "filename": _safe_filename(college_name),
            "dept_line": dept_line,
            "num_departments": len(departments),
            "num_students": len(students),
            "columns": college_columns,
            "rows": student_rows,
            "subtitle": subtitle,
        }

    @bp.route("/reports/college/<college_id>/export", methods=["GET"])
    @role_required("super_admin")
    def admin_college_report_export(college_id):
        """College Profile → Download Report (PDF/Excel)."""
        oid = to_object_id(college_id)
        if not oid:
            return error("Invalid college id.", 404)
        college = db.colleges.find_one({"_id": oid})
        if not college:
            return error("College not found.", 404)

        p = _college_report_payload(college)
        fmt = (request.args.get("format") or "pdf").lower()
        if fmt == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "College Report"
            ws.append(["College Report", p["name"]])
            ws.append(["Departments", p["dept_line"]])
            ws.append(["Students", p["num_students"]])
            ws.append([])
            ws.append(p["columns"])
            for r in p["rows"]:
                ws.append(r)
            for c in range(1, len(p["columns"]) + 1):
                cell = ws.cell(row=5, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F9E93")
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"college_report_{p['filename']}.xlsx",
            )
        return send_file(
            _pdf_bytes(f"College Report — {p['name']}", p["subtitle"], p["columns"], p["rows"]),
            mimetype="application/pdf", as_attachment=True,
            download_name=f"college_report_{p['filename']}.pdf",
        )

    @bp.route("/reports/colleges/export", methods=["GET"])
    @role_required("super_admin")
    def admin_colleges_report_export():
        """Multi-college report — the College Report for every college
        picked in the Reports modal. One sheet per college in Excel; one
        section per college in the PDF. Each college's students never mix
        with another college's."""
        fmt = (request.args.get("format") or "pdf").lower()
        college_oids, err = _parse_optional_id_list(request.args.get("colleges"))
        if err:
            return error(err)
        if not college_oids:
            return error("Select at least one college to export.")
        college_docs = list(db.colleges.find({"_id": {"$in": college_oids}}))
        if not college_docs:
            return error("No colleges match the selected filters.")
        payloads = [_college_report_payload(c) for c in college_docs]
        total_students = sum(p["num_students"] for p in payloads)
        subtitle = (f"{len(payloads)} college report(s) · {total_students} student(s)"
                    f" · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")

        if fmt == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Overview"
            ws.append(["College Reports", len(payloads)])
            ws.append(["Total Students", total_students])
            ws.append(["Generated", datetime.utcnow().strftime('%d %b %Y %H:%M UTC')])
            for p in payloads:
                sheet = wb.create_sheet(p["filename"][:31])
                sheet.append(["College Report", p["name"]])
                sheet.append([p["dept_line"]])
                sheet.append(["Students", p["num_students"]])
                sheet.append([])
                sheet.append(p["columns"])
                for r in p["rows"]:
                    sheet.append(r)
                for c in range(1, len(p["columns"]) + 1):
                    cell = sheet.cell(row=5, column=c)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F9E93")
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="colleges_report.xlsx",
            )

        sections = []
        for p in payloads:
            sections.append((p["name"], ["Detail"],
                             [[f"Departments: {p['num_departments']}"],
                              [f"Students: {p['num_students']}"]]))
            sections.append((f"{p['name']} — Students", p["columns"], p["rows"]))
        return send_file(
            _pdf_sections("College Report", subtitle, sections),
            mimetype="application/pdf", as_attachment=True,
            download_name="colleges_report.pdf",
        )

    def _student_report_payload(student):
        """Shared builder for the Single Student Report and the multi-student
        report export. Returns the summary lines, the three data sections
        (Performance / Interventions / Activity) and their columns — all
        computed live from the same sources the Student Profile page shows."""
        oid = student["_id"]
        name = student.get("fullName") or student.get("name") or "Student"
        reg = student.get("rollNumber") or student.get("tneaCode") or "—"
        college = student.get("college") or "—"
        department = student.get("department") or "—"
        cohort = student_cohort_label(student)
        att = attendance_summary(db, oid)
        att_pct = att.get("percentage") if att and isinstance(att, dict) else None

        def pct_of(v):
            return _fmt(v) if v is not None else "—"

        perf_columns = ["Assessment", "Quiz %", "Interview %", "Average %", "Final Overall %", "Submitted"]
        perf_data = [[r["assessmentName"] or "—", pct_of(r["quizMarks"]), pct_of(r["interviewMarks"]),
                      pct_of(r["averageMarks"]), pct_of(r["finalOverallMarks"]),
                      (r["submittedAt"] or "")[:16].replace("T", " ")]
                     for r in _overall_report_rows(oid)]

        sess_query = {"collegeNames": {"$in": [college]}} if college and college != "—" else {}
        session_columns = ["Session", "Date", "Time", "Trainer", "Attendance Req", "Status", "Student Status"]
        session_data = []
        for sc in db.workshop_sessions.find(sess_query).sort("date", -1).limit(10):
            date_str = sc.get("date")
            marked = None
            if date_str:
                marked = db.attendance.find_one({"studentId": oid, "date": date_str})
            if marked:
                st = marked.get("status")
                student_status = "Present" if st == "present" else ("Absent" if st == "absent" else "—")
            else:
                student_status = "Not marked"
            session_data.append([
                sc.get("name") or "—",
                date_str or "—",
                f"{sc.get('startTime','')} – {sc.get('endTime','')}" if sc.get("startTime") else "—",
                ", ".join(sc.get("trainerNames", [])) or "—",
                (sc.get("attendanceRequirement") or "—").title(),
                (sc.get("status") or "—").title(),
                student_status,
            ])

        activity_columns = ["Date", "Action", "Detail"]
        activity_data = [[iso_utc(a.get("createdAt"))[:16].replace("T", " "),
                          (a.get("action") or "—").replace("_", " ").title(),
                          a.get("description") or "—"]
                         for a in db.activity_log.find(
                             {"$or": [{"studentId": str(oid)}, {"actorId": str(oid)}]},
                             {"action": 1, "description": 1, "createdAt": 1},
                         ).sort("createdAt", -1).limit(10)]

        summary = [
            f"Student: {name}  ({reg})",
            f"College: {college} · Department: {department} · Cohort: {cohort}",
            (f"Overall Employability: {pct_of(student.get('finalEmployabilityScore'))}% · "
             f"Baseline: {pct_of(student.get('baselineAssessmentScore'))}% · "
             f"Interview: {pct_of(student.get('interviewScore'))}% · Attendance: {_fmt(att_pct)}%"),
        ]
        return {
            "name": name,
            "filename": _safe_filename(name),
            "summary": summary,
            "subtitle": " · ".join(summary),
            "sections": [
                ("Assessment Performance", perf_columns, perf_data),
                ("Interventions", session_columns, session_data),
                ("Activity", activity_columns, activity_data),
            ],
            "excel_sections": [
                ("Performance", perf_columns, perf_data),
                ("Interventions", session_columns, session_data),
                ("Activity", activity_columns, activity_data),
            ],
        }

    @bp.route("/reports/student/<student_id>/export", methods=["GET"])
    @role_required("super_admin")
    def admin_student_report_export(student_id):
        """Student Profile → Export (PDF/Excel). Every section live from
        the same sources the Student Profile page shows (see
        _student_report_payload)."""
        oid = to_object_id(student_id)
        if not oid:
            return error("Invalid student id.", 404)
        student = users.find_one({"_id": oid, "role": "student"})
        if not student:
            return error("Student not found.", 404)

        p = _student_report_payload(student)
        fmt = (request.args.get("format") or "pdf").lower()
        if fmt == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            for line in p["summary"]:
                ws.append([line])
            for sheet_name, cols, data in p["excel_sections"]:
                sheet = wb.create_sheet(sheet_name[:31])
                sheet.append(cols)
                for c in range(1, len(cols) + 1):
                    cell = sheet.cell(row=1, column=c)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F9E93")
                for r in data:
                    sheet.append(r)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True, download_name=f"{p['filename']}_report.xlsx",
            )
        return send_file(
            _pdf_sections(f"Student Report — {p['name']}", p["subtitle"], p["sections"]),
            mimetype="application/pdf", as_attachment=True,
            download_name=f"{p['filename']}_report.pdf",
        )

    @bp.route("/reports/students/export", methods=["GET"])
    @role_required("super_admin")
    def admin_students_report_export():
        """Multi-student report — the Individual Student Report for a list of
        students picked in the Reports modal (optionally scoped by college /
        department ids). One summary block per student in the PDF; one sheet
        per student in Excel. Never mixes two students' data in one block."""
        fmt = (request.args.get("format") or "pdf").lower()
        student_oids, err = _parse_optional_id_list(request.args.get("students"))
        if err:
            return error(err)
        college_oids, err = _parse_optional_id_list(request.args.get("colleges"))
        if err:
            return error(err)
        department_oids, err = _parse_optional_id_list(request.args.get("departments"))
        if err:
            return error(err)

        query = {"role": "student", "isDeleted": {"$ne": True}}
        if student_oids:
            query["_id"] = {"$in": student_oids}
        if college_oids:
            query["collegeId"] = {"$in": college_oids}
        if department_oids:
            query["departmentId"] = {"$in": department_oids}
        students = list(users.find(query).sort("fullName", 1))
        if not students:
            return error("No students match the selected filters.")

        payloads = [_student_report_payload(s) for s in students]
        total_sessions = sum(len(p["sections"][1][2]) for p in payloads)
        subtitle = (f"{len(students)} student report(s) · {total_sessions} intervention record(s)"
                    f" · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")

        if fmt == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Overview"
            ws.append(["Student Reports", len(students)])
            ws.append(["Generated", datetime.utcnow().strftime('%d %b %Y %H:%M UTC')])
            ws.append([])
            ws.append(["Student", "Register No", "College", "Department", "Cohort"])
            for p, s in zip(payloads, students):
                ws.append([
                    p["name"], s.get("rollNumber") or s.get("tneaCode") or "—",
                    s.get("college") or "—", s.get("department") or "—",
                    student_cohort_label(s),
                ])
            for p in payloads:
                sheet = wb.create_sheet(_safe_filename(p["name"])[:31])
                for line in p["summary"]:
                    sheet.append([line])
                sheet.append([])
                for sheet_name, cols, data in p["excel_sections"]:
                    sheet.append([f"— {sheet_name} —"])
                    sheet.append(cols)
                    for c in range(1, len(cols) + 1):
                        cell = sheet.cell(row=sheet.max_row - len(data), column=c)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="1F9E93")
                    for r in data:
                        sheet.append(r)
                    sheet.append([])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True, download_name="students_report.xlsx",
            )

        sections = []
        for p in payloads:
            sections.append((p["name"], ["Detail"], [[line] for line in p["summary"]]))
            sections.extend(p["sections"])
        return send_file(
            _pdf_sections("Individual Student Report", subtitle, sections),
            mimetype="application/pdf", as_attachment=True,
            download_name="students_report.pdf",
        )

    @bp.route("/reports/assessments/export", methods=["GET"])
    @role_required("super_admin")
    def admin_assessment_report_export():
        """Individual Assessment Report — one or more assessments selected
        by the admin, downloaded as PDF/Excel. For each assessment, every
        student who submitted it, with Quiz %, Interview %, Average %,
        Final Overall % and submission time (live from db.quiz_attempts)."""
        fmt = (request.args.get("format") or "pdf").lower()
        college = (request.args.get("college") or "").strip()
        college_oids, err = _parse_optional_id_list(request.args.get("colleges"))
        if err:
            return error(err)
        college_names = None
        if college_oids:
            college_names = [c.get("college_name") or c.get("name")
                             for c in db.colleges.find({"_id": {"$in": college_oids}})]
            college_names = [n for n in college_names if n]
        raw_ids = [x.strip() for x in (request.args.get("assessments") or "").split(",") if x.strip()]
        quiz_ids = []
        for qid in raw_ids:
            oid = to_object_id(qid)
            if oid:
                quiz_ids.append(oid)
        if not quiz_ids:
            return error("Select at least one assessment to export.")
        names = {str(d["_id"]): (d.get("title") or "Quiz")
                 for d in db.quizzes.find({"_id": {"$in": quiz_ids}}, {"title": 1})}

        rows = _overall_report_rows(quiz_ids=quiz_ids, colleges=college_names or ([college] if college else None))
        id_set = set(raw_ids)
        by_assessment = {qid: [] for qid in quiz_ids}
        for r in rows:
            aid = r.get("assessmentId")
            if aid and aid in id_set:
                by_assessment.get(aid, []).append(r)
        ordered = [qid for qid in quiz_ids if by_assessment[qid]]

        def marks_of(r):
            got = r.get("quizMarksObtained")
            tot = r.get("quizTotalMarks")
            if got is None or tot is None:
                return "—"
            return f"{got}/{tot}"

        def result_of(r):
            pf = r.get("passFail")
            if pf in ("Pass", "Fail"):
                return pf
            pct = r.get("quizMarks")
            if pct is None:
                return "—"
            return "Pass" if float(pct) >= 40 else "Fail"

        columns = ["Assessment", "Student", "Register No", "Department", "College",
                   "Quiz %", "Marks", "Result", "Interview %", "Average %", "Final Overall %", "Submitted"]
        def row_of(r):
            return [names.get(r.get("assessmentId"), r["assessmentName"] or "—"),
                    r["studentName"] or "—", r["rollNumber"] or "—", r["department"] or "—",
                    r["college"] or "—", _fmt(r["quizMarks"]), marks_of(r), result_of(r),
                    _fmt(r["interviewMarks"]), _fmt(r["averageMarks"]), _fmt(r["finalOverallMarks"]),
                    (r["submittedAt"] or "")[:16].replace("T", " ")]
        data = [row_of(r) for r in rows]

        subtitle = (f"{len(rows)} submitted attempt(s) across {len(ordered)} selected assessment(s)"
                    f"{(' · college: ' + college) if college else ''}"
                    f" · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")
        if fmt == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "All Assessments"
            ws.append(columns)
            for c in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F9E93")
            for r in data:
                ws.append(r)
            for qid in ordered:
                name = names.get(str(qid), "Assessment")
                sheet = wb.create_sheet(_safe_filename(name)[:31])
                per_columns = columns[1:]
                sheet.append(per_columns)
                for c in range(1, len(per_columns) + 1):
                    cell = sheet.cell(row=1, column=c)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F9E93")
                for r in by_assessment[qid]:
                    sheet.append([r["studentName"] or "—", r["rollNumber"] or "—",
                                  r["department"] or "—", r["college"] or "—",
                                  _fmt(r["quizMarks"]), marks_of(r), result_of(r),
                                  _fmt(r["interviewMarks"]), _fmt(r["averageMarks"]),
                                  _fmt(r["finalOverallMarks"]),
                                  (r["submittedAt"] or "")[:16].replace("T", " ")])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="assessment_report.xlsx",
            )
        sections = [(names.get(str(qid), "Assessment"), columns[1:],
                     [[r["studentName"] or "—", r["rollNumber"] or "—", r["department"] or "—",
                       r["college"] or "—", _fmt(r["quizMarks"]), marks_of(r), result_of(r),
                       _fmt(r["interviewMarks"]), _fmt(r["averageMarks"]),
                       _fmt(r["finalOverallMarks"]),
                       (r["submittedAt"] or "")[:16].replace("T", " ")] for r in by_assessment[qid]])
                    for qid in ordered]
        return send_file(
            _pdf_sections("Individual Assessment Report", subtitle, sections),
            mimetype="application/pdf", as_attachment=True,
            download_name="assessment_report.pdf",
        )

    @bp.route("/reports/interventions/export", methods=["GET"])
    @role_required("super_admin")
    def admin_intervention_report_export():
        """Intervention Report — every workshop session in db.workshop_sessions
        with its target colleges/departments/trainers, cohort, schedule,
        venue, attendance requirement, workflow status and approval state,
        plus real attendance participation (present/absent counts on each
        session date from db.attendance)."""
        fmt = (request.args.get("format") or "pdf").lower()
        statuses = [s for s in (request.args.get("statuses") or "").split(",") if s.strip() in SESSION_STATUS_OPTIONS]
        college_oids, err = _parse_optional_id_list(request.args.get("colleges"))
        if err:
            return error(err)
        department_oids, err = _parse_optional_id_list(request.args.get("departments"))
        if err:
            return error(err)
        trainer_oids, err = _parse_optional_id_list(request.args.get("trainers"))
        if err:
            return error(err)
        session_oids, err = _parse_optional_id_list(request.args.get("sessions"))
        if err:
            return error(err)

        # Keep report statuses consistent with the auto-complete rule applied
        # on the intervention list.
        _auto_complete_workshop_sessions(db)

        query = {}
        if statuses:
            query["status"] = {"$in": statuses}
        if session_oids:
            query["_id"] = {"$in": session_oids}
        docs = list(db.workshop_sessions.find(query).sort("date", -1))
        if college_oids:
            docs = [d for d in docs if d.get("collegeIds") and any(
                c in college_oids for c in d["collegeIds"])]
        if department_oids:
            docs = [d for d in docs if d.get("departmentIds") and any(
                d_ in department_oids for d_ in d["departmentIds"])]
        if trainer_oids:
            docs = [d for d in docs if d.get("trainerIds") and any(
                t in trainer_oids for t in d["trainerIds"])]

        columns = ["Session", "Colleges", "Departments", "Trainers", "Cohort", "Date",
                   "Time", "Venue", "Attendance Req", "Status", "Approval",
                   "Marked", "Present", "Absent", "Created"]
        data = []
        for d in docs:
            cnt = _attendance_counts_for_date(d.get("date"))
            data.append([
                d.get("name") or "—",
                ", ".join(d.get("collegeNames", [])) or "—",
                ", ".join(d.get("departmentNames", [])) or "—",
                ", ".join(d.get("trainerNames", [])) or "—",
                d.get("cohort") or "—",
                d.get("date") or "—",
                f"{d.get('startTime','')} – {d.get('endTime','')}" if d.get("startTime") else "—",
                d.get("venue") or "—",
                (d.get("attendanceRequirement") or "—").title(),
                (d.get("status") or "—").title(),
                (d.get("approvalStatus") or "—").title(),
                cnt["total"], cnt["present"], cnt["absent"],
                iso_utc(d.get("createdAt"))[:16].replace("T", " "),
            ])
        subtitle = (f"{len(docs)} intervention session(s)"
                    f"{(' · status: ' + ', '.join(statuses)) if statuses else ''}"
                    f" · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")
        if fmt == "excel":
            return send_file(
                _excel_bytes(columns, data, "Intervention Report"),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="intervention_report.xlsx",
            )
        return send_file(
            _pdf_bytes("Intervention Report", subtitle, columns, data),
            mimetype="application/pdf", as_attachment=True,
            download_name="intervention_report.pdf",
        )

    @bp.route("/reports/performance/export", methods=["GET"])
    @role_required("super_admin")
    def admin_student_performance_export():
        """Student Performance Report — a per-student roster with the exact
        performance fields the College Admin Student Management shows
        (overall employability, baseline, interview, cohort, attendance),
        computed live from db.users + db.attendance so both dashboards can
        never disagree. Optional college / cohort filters."""
        fmt = (request.args.get("format") or "pdf").lower()
        college_oids, err = _parse_optional_id_list(request.args.get("colleges"))
        if err:
            return error(err)
        department_oids, err = _parse_optional_id_list(request.args.get("departments"))
        if err:
            return error(err)
        student_oids, err = _parse_optional_id_list(request.args.get("students"))
        if err:
            return error(err)
        cohort_param = (request.args.get("cohorts") or "").strip()
        cohorts = [c for c in cohort_param.split(",") if c.strip()] if cohort_param else []

        query = {"role": "student", "isDeleted": {"$ne": True}}
        if college_oids:
            query["collegeId"] = {"$in": college_oids}
        if department_oids:
            query["departmentId"] = {"$in": department_oids}
        if student_oids:
            query["_id"] = {"$in": student_oids}
        if cohorts:
            allowed = set(VALID_COHORTS) | {ENTRY_LEVEL}
            if any(c not in allowed for c in cohorts):
                return error("Invalid cohort filter.")
            query["$or"] = []
            if ENTRY_LEVEL in cohorts:
                query["$or"].append({"cohort": None})
                query["$or"].append({"cohort": {"$exists": False}})
            for c in cohorts:
                if c in VALID_COHORTS:
                    query["$or"].append({"cohort": c})

        docs = list(users.find(
            query,
            {"fullName": 1, "rollNumber": 1, "tneaCode": 1, "college": 1,
             "department": 1, "cohort": 1, "baselineAssessmentScore": 1,
             "interviewScore": 1, "finalEmployabilityScore": 1},
        ).sort("fullName", 1))

        # One aggregation over db.attendance for the whole roster — same
        # shared source the College Admin roster and Student Dashboard use.
        oids = [d["_id"] for d in docs]
        att_stats = {}
        if oids:
            for row in db.attendance.aggregate([
                {"$match": {"studentId": {"$in": oids}}},
                {"$group": {
                    "_id": "$studentId",
                    "present": {"$sum": {"$cond": [{"$eq": ["$status", "present"]}, 1, 0]}},
                    "absent": {"$sum": {"$cond": [{"$eq": ["$status", "absent"]}, 1, 0]}},
                }},
            ]):
                att_stats[row["_id"]] = row

        columns = ["Student", "Register No", "College", "Department", "Cohort",
                   "Baseline %", "Interview %", "Overall Employability %", "Attendance %", "Assessment Status"]
        data = []
        for s in docs:
            a = att_stats.get(s["_id"], {})
            present = a.get("present", 0)
            absent = a.get("absent", 0)
            total = present + absent
            att_pct = round(present / total * 100, 1) if total else 0
            baseline = s.get("baselineAssessmentScore")
            interview = s.get("interviewScore")
            status = "Completed" if baseline is not None and interview is not None else "Pending"
            cohort_label = student_cohort_label(s)
            data.append([
                s.get("fullName") or "—",
                s.get("rollNumber") or s.get("tneaCode") or "—",
                s.get("college") or "—",
                s.get("department") or "—",
                f"Cohort {cohort_label}" if cohort_label in ("A", "B", "C") else "Entry Level",
                _fmt(baseline) if baseline is not None else "—",
                _fmt(interview) if interview is not None else "—",
                _fmt(s.get("finalEmployabilityScore")) if s.get("finalEmployabilityScore") is not None else "—",
                att_pct, status,
            ])
        subtitle = (f"{len(docs)} student(s)"
                    f"{(' · cohort: ' + ', '.join(cohorts)) if cohorts else ''}"
                    f" · generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}")
        if fmt == "excel":
            return send_file(
                _excel_bytes(columns, data, "Student Performance"),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="student_performance_report.xlsx",
            )
        return send_file(
            _pdf_bytes("Student Performance Report", subtitle, columns, data),
            mimetype="application/pdf", as_attachment=True,
            download_name="student_performance_report.pdf",
        )

    # ============================================================
    # PLATFORM SETTINGS — backed by db.platform_settings (single doc,
    # _id "platform"). Every Super Admin session loads from here and
    # saves back here, so settings persist platform-wide rather than
    # per-browser. Nothing here is enforced beyond storage yet —
    # maintenance mode / password-policy flags are stored for the
    # login/register modules to consume later.
    # ============================================================
    SETTINGS_KEY = "platform"
    SETTINGS_DEFAULTS = {
        "theme": "light",
        "maintenanceMode": False,
        "passwordPolicy": {
            "minLength": True,
            "uppercaseNumber": True,
            "special": False,
            "expiry90": False,
        },
    }

    def _read_platform_settings():
        doc = db.platform_settings.find_one({"_id": SETTINGS_KEY})
        settings = {
            "theme": SETTINGS_DEFAULTS["theme"],
            "maintenanceMode": SETTINGS_DEFAULTS["maintenanceMode"],
            "passwordPolicy": dict(SETTINGS_DEFAULTS["passwordPolicy"]),
        }
        if not doc:
            return settings
        policy = doc.get("passwordPolicy") or {}
        settings["theme"] = "dark" if doc.get("theme") == "dark" else "light"
        settings["maintenanceMode"] = bool(doc.get("maintenanceMode", False))
        for flag in ("minLength", "uppercaseNumber", "special", "expiry90"):
            if flag in policy:
                settings["passwordPolicy"][flag] = bool(policy.get(flag))
        return settings

    @bp.route("/settings", methods=["GET"])
    @role_required("super_admin")
    def get_platform_settings():
        return ok({"settings": _read_platform_settings()})

    @bp.route("/settings", methods=["PUT"])
    @role_required("super_admin")
    def save_platform_settings():
        data = request.get_json(silent=True) or {}
        policy = data.get("passwordPolicy") or {}
        saved = _read_platform_settings()
        saved["theme"] = "dark" if data.get("theme") == "dark" else "light"
        saved["maintenanceMode"] = bool(data.get("maintenanceMode", saved["maintenanceMode"]))
        for flag in ("minLength", "uppercaseNumber", "special", "expiry90"):
            if flag in policy:
                saved["passwordPolicy"][flag] = bool(policy.get(flag))
        db.platform_settings.update_one(
            {"_id": SETTINGS_KEY},
            {"$set": {
                "theme": saved["theme"],
                "maintenanceMode": saved["maintenanceMode"],
                "passwordPolicy": saved["passwordPolicy"],
                "updatedAt": now(),
            }},
            upsert=True,
        )
        log_activity(
            db, get_jwt_identity(), "super_admin", "settings_updated",
            "Platform settings saved",
        )
        return ok({"settings": saved})

    @bp.route("/audit-logs", methods=["GET"])
    @role_required("super_admin")
    def list_audit_logs():
        """Super Admin Audit Logs — reads the real db.activity_log feed
        (written by log_activity() on every meaningful action), newest
        first. No mock records. Actor display names are resolved from
        db.users so the table shows who did what; unknown actors fall
        back to their role label."""
        docs = list(db.activity_log.find({}).sort("createdAt", -1).limit(500))
        actor_ids = [d.get("actorId") for d in docs if d.get("actorId")]
        names = {}
        if actor_ids:
            for u in users.find({"_id": {"$in": [to_object_id(x) or x for x in actor_ids]}}, {"fullName": 1}):
                names[str(u["_id"])] = u.get("fullName") or "—"
        role_labels = {
            "super_admin": "Super Admin",
            "college_admin": "College Admin",
            "trainer": "Trainer",
            "student": "Student",
        }
        return ok({"logs": [{
            "id": str(d["_id"]),
            "actorName": names.get(d.get("actorId")) or None,
            "actorRole": d.get("actorRole"),
            "actorRoleLabel": role_labels.get(d.get("actorRole"), d.get("actorRole") or "—"),
            "action": d.get("action"),
            "description": d.get("description"),
            "college": d.get("college"),
            "meta": d.get("meta") or {},
            "createdAt": iso_utc(d.get("createdAt")),
            "createdAtIST": fmt_ist(d.get("createdAt")),
        } for d in docs]})

    @bp.route("/dashboard/search", methods=["GET"])
    @role_required("super_admin")
    def dashboard_search():
        """Instant search across Student Name, Register Number, College,
        Department, Trainer, Assessment. Empty/whitespace query returns an
        empty (not an error) result set."""
        q = (request.args.get("q") or "").strip()
        if not q:
            return ok({"results": {
                "students": [], "trainers": [], "colleges": [],
                "departments": [], "assessments": [],
            }})

        rx = re.compile(re.escape(q), re.IGNORECASE)
        limit = 8

        students = list(users.find(
            {"role": "student", "$or": [
                {"fullName": rx}, {"rollNumber": rx}, {"tneaCode": rx},
            ]}, {"fullName": 1, "rollNumber": 1, "college": 1, "department": 1},
        ).limit(limit))

        trainers = list(users.find(
            {"role": "trainer", "$or": [{"fullName": rx}, {"employeeId": rx}]},
            {"fullName": 1, "employeeId": 1, "college": 1},
        ).limit(limit))

        colleges_found = list(db.colleges.find(
            {"college_name": rx}, {"college_name": 1, "status": 1},
        ).limit(limit))

        departments_found = list(db.departments.find(
            {"department_name": rx}, {"department_name": 1, "college_id": 1},
        ).limit(limit))

        assessments_found = list(assessments.find(
            {"name": rx}, {"name": 1, "type": 1, "college": 1},
        ).limit(limit))
        quizzes_found = list(db.quizzes.find(
            {"title": rx}, {"title": 1, "collegeNames": 1},
        ).limit(limit))

        return ok({"results": {
            "students": [
                {"id": str(s["_id"]), "name": s.get("fullName"), "rollNumber": s.get("rollNumber"),
                 "college": s.get("college"), "department": s.get("department")}
                for s in students
            ],
            "trainers": [
                {"id": str(t["_id"]), "name": t.get("fullName"), "employeeId": t.get("employeeId"),
                 "college": t.get("college")}
                for t in trainers
            ],
            "colleges": [
                {"id": str(c["_id"]), "name": c.get("college_name"), "status": c.get("status")}
                for c in colleges_found
            ],
            "departments": [
                {"id": str(d["_id"]), "name": d.get("department_name")}
                for d in departments_found
            ],
            "assessments": (
                [{"id": str(a["_id"]), "name": a.get("name"), "type": a.get("type")} for a in assessments_found]
                + [{"id": str(z["_id"]), "name": z.get("title"), "type": "quiz"} for z in quizzes_found]
            ),
        }})

    @bp.route("/dashboard/notifications", methods=["GET"])
    @role_required("super_admin")
    def dashboard_notifications():
        """Aggregates every pending/attention-needed item across the
        platform into one feed. Nothing here is stored — each item is
        recomputed live from its source collection on every call; only
        read/unread state persists (db.notification_reads)."""
        now_dt = datetime.utcnow()
        recent_cutoff = now_dt - timedelta(days=14)
        items = []

        for doc in users.find(
            {"role": "college_admin", "approvalStatus": "pending"},
            {"fullName": 1, "createdAt": 1},
        ):
            items.append({
                "type": "college_verification",
                "title": "College verification pending",
                "description": f"{doc.get('fullName', 'A college admin')} is awaiting approval.",
                "createdAt": doc.get("createdAt"),
                "sourceId": str(doc["_id"]),
            })

        for doc in users.find(
            {"role": "trainer", "approvalStatus": "pending"},
            {"fullName": 1, "createdAt": 1},
        ):
            items.append({
                "type": "trainer_approval",
                "title": "Trainer approval pending",
                "description": f"{doc.get('fullName', 'A trainer')} is awaiting approval.",
                "createdAt": doc.get("createdAt"),
                "sourceId": str(doc["_id"]),
            })

        pending_students = users.count_documents({"role": "student", "approvalStatus": "pending"})
        if pending_students:
            items.append({
                "type": "student_request",
                "title": "Student requests pending",
                "description": f"{pending_students} student registration(s) awaiting approval.",
                "createdAt": now_dt,
                "sourceId": "student-pending-bucket",
            })

        for doc in db.quizzes.find(
            {"endDateTime": {"$lt": now_dt, "$gte": recent_cutoff}, "state": "published", "cancelled": {"$ne": True}},
            {"title": 1, "endDateTime": 1},
        ):
            items.append({
                "type": "overdue_assessment",
                "title": "Assessment window closed",
                "description": f'"{doc.get("title", "Untitled assessment")}" has ended — review results.',
                "createdAt": doc.get("endDateTime"),
                "sourceId": str(doc["_id"]),
            })

        for doc in attempts.find(
            {"status": "submitted", "submittedAt": {"$gte": now_dt - timedelta(days=2)}},
            {"studentId": 1, "submittedAt": 1},
        ).sort("submittedAt", -1).limit(20):
            items.append({
                "type": "new_submission",
                "title": "New assessment submission",
                "description": "A student submitted an assessment attempt.",
                "createdAt": doc.get("submittedAt"),
                "sourceId": str(doc["_id"]),
            })

        for doc in db.colleges.find(
            {"created_at": {"$gte": recent_cutoff}}, {"college_name": 1, "created_at": 1},
        ):
            items.append({
                "type": "newly_added_college",
                "title": "New college added",
                "description": f'"{doc.get("college_name", "A college")}" was added to the platform.',
                "createdAt": doc.get("created_at"),
                "sourceId": str(doc["_id"]),
            })

        for doc in db.activity_log.find(
            {"action": "placement_rules_updated", "createdAt": {"$gte": recent_cutoff}},
            {"description": 1, "createdAt": 1},
        ):
            items.append({
                "type": "placement_rule_update",
                "title": "Placement rules updated",
                "description": doc.get("description", "Placement rules were changed."),
                "createdAt": doc.get("createdAt"),
                "sourceId": str(doc["_id"]),
            })

        other_actions = {
            "workshop_session_scheduled", "workshop_session_updated", "workshop_session_deleted",
            "user_activated", "user_deactivated", "user_deleted", "student_cohort_backfill",
        }
        for doc in db.activity_log.find(
            {"actorRole": "super_admin", "action": {"$in": list(other_actions)},
             "createdAt": {"$gte": now_dt - timedelta(days=3)}},
            {"description": 1, "createdAt": 1, "action": 1},
        ).sort("createdAt", -1).limit(15):
            items.append({
                "type": "other_administrative",
                "title": "Administrative activity",
                "description": doc.get("description", doc.get("action", "")),
                "createdAt": doc.get("createdAt"),
                "sourceId": str(doc["_id"]),
            })

        # Priority: Meeting Link Update Required — an ONLINE session is now
        # live (inside its start→end window) and still has no meeting link.
        # Generated live from the session's own scheduled date + start/end
        # time (never hardcoded); it disappears once the link is set or the
        # session reaches its end time.
        for doc in db.workshop_sessions.find(
            {"mode": "online", "status": "scheduled"},
            {"name": 1, "date": 1, "startTime": 1, "endTime": 1, "meetingLink": 1},
        ):
            if not _meeting_link_window_open(doc):
                continue
            if doc.get("meetingLink"):
                continue
            items.append({
                "type": "meeting_link_update",
                "title": "Priority: Meeting Link Update Required",
                "description": (
                    f'The meeting link for "{doc.get("name", "the live session")}" has not been '
                    f"added yet. The session is live and can be updated until it ends on "
                    f"{doc.get('date')} at {doc.get('endTime')}."
                ),
                "createdAt": now_dt,
                "sourceId": str(doc["_id"]),
            })

        # System Alerts / Failed Background Jobs — sourced from
        # db.system_alerts if/when something writes to it (e.g. a future
        # background worker). Queried defensively so the panel is ready
        # for that day without ever showing a fake alert in the meantime.
        for doc in db.system_alerts.find(
            {"createdAt": {"$gte": recent_cutoff}},
        ).sort("createdAt", -1).limit(15):
            items.append({
                "type": doc.get("kind", "system_alert"),
                "title": doc.get("title", "System alert"),
                "description": doc.get("description", ""),
                "createdAt": doc.get("createdAt"),
                "sourceId": str(doc["_id"]),
            })

        for it in items:
            it["id"] = _notification_id(it["type"], it["sourceId"])
            it["createdAt"] = iso_utc(it["createdAt"]) if it["createdAt"] else None

        read_ids = {r["_id"] for r in db.notification_reads.find({}, {"_id": 1})}
        for it in items:
            it["unread"] = it["id"] not in read_ids

        items.sort(key=lambda x: x["createdAt"] or "", reverse=True)
        return ok({"notifications": items, "unreadCount": sum(1 for it in items if it["unread"])})

    @bp.route("/dashboard/notifications/<notification_id>/read", methods=["PATCH"])
    @role_required("super_admin")
    def mark_notification_read(notification_id):
        db.notification_reads.update_one(
            {"_id": notification_id}, {"$set": {"readAt": now()}}, upsert=True,
        )
        return ok(message="Marked as read.")

    @bp.route("/dashboard/notifications/read-all", methods=["POST"])
    @role_required("super_admin")
    def mark_all_notifications_read():
        data = request.get_json(silent=True) or {}
        ids = data.get("ids") or []
        if not isinstance(ids, list) or not ids:
            return error("ids (non-empty list) is required.")
        ops = [
            {"filter": {"_id": nid}, "update": {"$set": {"readAt": now()}}, "upsert": True}
            for nid in ids
        ]
        for op in ops:
            db.notification_reads.update_one(op["filter"], op["update"], upsert=op["upsert"])
        return ok(message="All notifications marked as read.")

    # ==========================================================
    # 6. USER MANAGEMENT — Students, Trainers, College Admins, all
    #    read live from db.users. No hardcoded records anywhere.
    # ==========================================================
    @bp.route("/users", methods=["GET"])
    @role_required("super_admin")
    def list_users():
        role_param = (request.args.get("role") or request.args.get("userType") or "all").strip().lower()
        status_param = (request.args.get("status") or "all").strip().lower()
        search = (request.args.get("search") or "").strip()

        if role_param in ("all", ""):
            roles = MANAGED_USER_ROLES
        elif role_param in ROLE_ALIASES:
            roles = [ROLE_ALIASES[role_param]]
        else:
            return error("Invalid role filter.")

        # Only real, approved-at-some-point accounts show up here — pending
        # applications live in the Verification queues, and soft-deleted
        # accounts never appear in any normal listing.
        query = {
            "role": {"$in": roles},
            "approvalStatus": {"$in": ["approved", "suspended"]},
            "isDeleted": {"$ne": True},
        }
        if status_param == "active":
            query["approvalStatus"] = "approved"
        elif status_param == "inactive":
            query["approvalStatus"] = "suspended"
        elif status_param not in ("all", ""):
            return error("status must be 'active' or 'inactive'.")

        if search:
            pattern = re.escape(search)
            query["$or"] = [
                {"fullName": {"$regex": pattern, "$options": "i"}},
                {"email": {"$regex": pattern, "$options": "i"}},
                {"rollNumber": {"$regex": pattern, "$options": "i"}},
                {"employeeId": {"$regex": pattern, "$options": "i"}},
                {"tneaCode": {"$regex": pattern, "$options": "i"}},
            ]

        docs = users.find(query).sort("fullName", 1)
        rows = [_user_public_row(d) for d in docs]
        return ok({"users": rows, "total": len(rows)})

    def _find_managed_user(user_id):
        oid = to_object_id(user_id)
        if not oid:
            return None, error("Invalid user id.", 404)
        doc = users.find_one({"_id": oid, "role": {"$in": MANAGED_USER_ROLES}})
        if not doc:
            return None, error("User not found.", 404)
        return doc, None

    # Activate / Deactivate — reuses the SAME approvalStatus field that
    # already gates login() ("approved" required to sign in; anything else
    # is rejected). Deactivating also clears currentSessionId so a user who
    # is mid-session is signed out immediately, not just blocked next time.
    @bp.route("/users/<user_id>/status", methods=["PATCH"])
    @role_required("super_admin")
    def set_user_status(user_id):
        user, err = _find_managed_user(user_id)
        if err:
            return err
        if user.get("isDeleted"):
            return error("This account has been deleted and can no longer be modified.")
        if user.get("approvalStatus") not in ("approved", "suspended"):
            return error("Only approved accounts can be activated or deactivated.")

        data = request.get_json(silent=True) or {}
        requested = (data.get("status") or "").strip().lower()
        if requested not in ("active", "inactive"):
            return error("status must be 'active' or 'inactive'.")

        is_activating = requested == "active"
        new_approval_status = "approved" if is_activating else "suspended"
        update = {"approvalStatus": new_approval_status, "updatedAt": now()}
        if not is_activating:
            update["currentSessionId"] = None
        users.update_one({"_id": user["_id"]}, {"$set": update})
        verb = "activated" if is_activating else "deactivated"
        log_activity(
            db, actor_id="super_admin", actor_role="super_admin",
            action=f"user_{verb}",
            description=f"{verb.capitalize()} {user.get('fullName') or user.get('email')}",
        )
        return ok(message=f"User {verb}.")

    # Soft delete — never a hard delete. Deleted users can never log in
    # (see login.py) and are excluded from every listing above by the
    # isDeleted:$ne:true filter, but the record itself is left in place
    # and fully recoverable (e.g. by clearing isDeleted directly in the
    # database) unless a genuine permanent-delete feature is added later.
    @bp.route("/users/<user_id>", methods=["DELETE"])
    @role_required("super_admin")
    def soft_delete_user(user_id):
        user, err = _find_managed_user(user_id)
        if err:
            return err
        if user.get("isDeleted"):
            return ok(message="User already deleted.")
        users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "isDeleted": True,
                "deletedAt": now(),
                "currentSessionId": None,
                "updatedAt": now(),
            }},
        )
        log_activity(
            db, actor_id="super_admin", actor_role="super_admin",
            action="user_deleted",
            description=f"Deleted {user.get('fullName') or user.get('email')}",
        )
        return ok(message="User deleted.")

    # Reset credentials — same random-temp-password pattern already used
    # for Trainers (see colleges.py admin_reset_trainer_password), now
    # available for every managed role from one place.
    @bp.route("/users/<user_id>/reset-password", methods=["POST"])
    @role_required("super_admin")
    def reset_user_password(user_id):
        user, err = _find_managed_user(user_id)
        if err:
            return err
        if user.get("isDeleted"):
            return error("This account has been deleted and can no longer be modified.")
        if bcrypt is None:
            return error("Password reset is not available right now.", 503)

        temp_password = _generate_temp_password()
        password_hash = bcrypt.generate_password_hash(temp_password).decode("utf-8")
        users.update_one(
            {"_id": user["_id"]},
            {"$set": {
                "passwordHash": password_hash,
                "currentSessionId": None,
                "updatedAt": now(),
            }},
        )
        return ok({"temporaryPassword": temp_password}, message="Credentials reset.")

    # ==========================================================
    # 7. DEPARTMENT SUMMARY — every department across every college,
    #    read live from db.departments/db.colleges/db.users. Powers the
    #    Department Management page and gives the Super Admin a single,
    #    real count of departments platform-wide.
    # ==========================================================
    @bp.route("/departments/summary", methods=["GET"])
    @role_required("super_admin")
    def department_summary():
        colleges_by_id = {c["_id"]: c.get("college_name") for c in db.colleges.find({}, {"college_name": 1})}
        top_cohort = top_cohort_label(db)

        rows = []
        for d in db.departments.find({}).sort("department_name", 1):
            student_ids_query = {
                "role": "student",
                "departmentId": d["_id"],
                "isDeleted": {"$ne": True},
            }
            student_count = users.count_documents(student_ids_query)
            scored = list(users.find(
                {**student_ids_query, "finalEmployabilityScore": {"$ne": None}},
                {"finalEmployabilityScore": 1, "cohort": 1},
            ))
            avg_score = (
                round(sum(s["finalEmployabilityScore"] for s in scored) / len(scored), 1)
                if scored else None
            )
            ready_count = sum(1 for s in scored if s.get("cohort") == top_cohort)
            placement_pct = round((ready_count / len(scored)) * 100, 1) if scored else None

            rows.append({
                "id": str(d["_id"]),
                "name": d.get("department_name"),
                "collegeId": str(d["college_id"]),
                "college": colleges_by_id.get(d["college_id"], "—"),
                "status": d.get("status", "active"),
                "students": student_count,
                "avgScore": avg_score,
                "placement": placement_pct,
            })

        return ok({
            "departments": rows,
            "totalDepartments": len(rows),
            "totalColleges": len(colleges_by_id),
        })

    # ==========================================================
    # SUPER ADMIN — SCHEDULE SESSION (db.workshop_sessions).
    # College/Department/Trainer are all multi-select, DB-backed, and
    # re-validated here regardless of what the client already checked.
    # Session Type and Maximum Students no longer exist anywhere in this
    # workflow; Attendance Requirement (Mandatory/Optional) replaces
    # Maximum Students. See module-level docstring above for the schema.
    # ==========================================================
    @bp.route("/interventions/sessions", methods=["GET"])
    @role_required("super_admin")
    def list_workshop_sessions():
        # Sessions whose stored end date/time has passed are auto-completed
        # (persisted) so the list/dashboard never shows a stale "Scheduled".
        _auto_complete_workshop_sessions(db)
        docs = db.workshop_sessions.find({}).sort("createdAt", -1)
        return ok({"sessions": [workshop_session_public(d) for d in docs]})

    def _validate_session_payload(data):
        """Shared by create + edit. Returns (fields_to_set, error_message)."""
        name = (data.get("name") or "").strip()
        if not name:
            return None, "Session name is required."

        college_ids, err = _parse_id_list(data.get("collegeIds"), "Colleges")
        if err:
            return None, err
        college_docs, err = _resolve_colleges(db, college_ids)
        if err:
            return None, err

        department_ids, err = _parse_id_list(data.get("departmentIds"), "Departments")
        if err:
            return None, err
        department_docs, college_names_by_id, err = _resolve_departments(db, department_ids)
        if err:
            return None, err

        trainer_ids, err = _parse_id_list(data.get("trainerIds"), "Trainers")
        if err:
            return None, err
        trainer_docs, err = _resolve_trainers(db, trainer_ids)
        if err:
            return None, err

        cohort = (data.get("cohort") or "").strip()
        if cohort not in SESSION_COHORT_OPTIONS:
            return None, "Applicable Cohort must be one of: " + ", ".join(SESSION_COHORT_OPTIONS) + "."

        date_str = (data.get("date") or "").strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
            return None, "A valid Session Date is required."
        start_time = (data.get("startTime") or "").strip()
        end_time = (data.get("endTime") or "").strip()
        if not (_TIME_RE.match(start_time) and _TIME_RE.match(end_time)):
            return None, "A valid Start Time and End Time are required."
        if end_time <= start_time:
            return None, "End Time must be after Start Time."

        mode = (data.get("mode") or "").strip().lower()
        if mode not in SESSION_MODE_OPTIONS:
            return None, "Mode of Conduct must be 'Online' or 'Offline'."

        venue = (data.get("venue") or "").strip()
        if mode == "offline" and not venue:
            return None, "Venue is required for offline sessions."
        if mode == "online":
            venue = venue or None

        attendance_requirement = (data.get("attendanceRequirement") or "").strip().lower()
        if attendance_requirement not in SESSION_ATTENDANCE_OPTIONS:
            return None, "Attendance Requirement must be 'Mandatory' or 'Optional'."

        fields = {
            "name": name,
            "mode": mode,
            "collegeIds": college_ids,
            "collegeNames": [c.get("college_name") for c in college_docs],
            "departmentIds": department_ids,
            "departmentNames": [
                f"{d.get('department_name')} ({college_names_by_id.get(d['college_id'], '—')})"
                for d in department_docs
            ],
            "trainerIds": trainer_ids,
            "trainerNames": [t.get("fullName") or "—" for t in trainer_docs],
            "cohort": cohort,
            "date": date_str,
            "startTime": start_time,
            "endTime": end_time,
            "venue": venue,
            "attendanceRequirement": attendance_requirement,
        }
        return fields, None

    @bp.route("/interventions/schedule", methods=["POST"])
    @role_required("super_admin")
    def schedule_workshop_session():
        data = request.get_json(silent=True) or {}
        fields, err = _validate_session_payload(data)
        if err:
            return error(err)

        actor_id = get_jwt_identity()
        doc = dict(fields)
        doc.update({
            "status": "scheduled",
            "approvalStatus": "pending",
            "createdBy": {"id": actor_id, "name": "Super Admin"},
            "createdAt": now(),
            "updatedAt": now(),
        })
        result = db.workshop_sessions.insert_one(doc)
        doc["_id"] = result.inserted_id

        log_activity(
            db, actor_id, "super_admin", "workshop_session_scheduled",
            f"Scheduled session \"{fields['name']}\"",
            meta={"sessionId": str(doc["_id"])},
        )
        return ok({"session": workshop_session_public(doc)}, message="Session scheduled.", status=201)

    @bp.route("/interventions/sessions/<session_id>", methods=["PATCH"])
    @role_required("super_admin")
    def update_workshop_session(session_id):
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        existing = db.workshop_sessions.find_one({"_id": oid})
        if not existing:
            return error("Session not found.", 404)

        data = request.get_json(silent=True) or {}

        # Status-only update (e.g. cancel), separate from the full edit form.
        if "status" in data and len(data) == 1:
            status = (data.get("status") or "").strip().lower()
            if status not in SESSION_STATUS_OPTIONS:
                return error("status must be one of: " + ", ".join(SESSION_STATUS_OPTIONS) + ".")
            db.workshop_sessions.update_one({"_id": oid}, {"$set": {"status": status, "updatedAt": now()}})
            updated = db.workshop_sessions.find_one({"_id": oid})
            return ok({"session": workshop_session_public(updated)}, message="Session updated.")

        # Business rule: a session can only be edited before it starts.
        if _session_started(existing):
            return error("Sessions can only be edited before they start.", 409)

        fields, err = _validate_session_payload(data)
        if err:
            return error(err)
        fields["updatedAt"] = now()
        db.workshop_sessions.update_one({"_id": oid}, {"$set": fields})
        updated = db.workshop_sessions.find_one({"_id": oid})

        log_activity(
            db, get_jwt_identity(), "super_admin", "workshop_session_updated",
            f"Updated session \"{fields['name']}\"",
            meta={"sessionId": str(oid)},
        )
        return ok({"session": workshop_session_public(updated)}, message="Session updated.")

    @bp.route("/interventions/sessions/<session_id>/meeting-link", methods=["PATCH"])
    @role_required("super_admin", "college_admin")
    def update_session_meeting_link(session_id):
        """Add/update the meeting link for an ONLINE session. Allowed only
        while the session is live — between its start time and end time
        (spec §15–17); the time check runs entirely on the server clock so
        disabling the button on the client is just cosmetic. Once the
        session reaches its end time, the update authority is removed
        immediately. The link must be a valid absolute URL
        and is stored on the existing session record — no new session is
        ever created. College Admins may only touch sessions in their own
        college; Super Admins may touch any."""
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        existing = db.workshop_sessions.find_one({"_id": oid})
        if not existing:
            return error("Session not found.", 404)

        claims = get_jwt()
        actor_role = claims.get("role") or "super_admin"
        if actor_role == "college_admin":
            actor = users.find_one({"_id": to_object_id(get_jwt_identity())})
            actor_college = actor.get("college") if actor else None
            if not actor_college or actor_college not in existing.get("collegeNames", []):
                return error("You are not authorized to update this session.", 403)

        if existing.get("mode") != "online":
            return error("Meeting links only apply to online sessions.", 409)
        if not _meeting_link_window_open(existing):
            return error("Meeting link can be updated only while the session is live (between its start and end time).", 409)

        data = request.get_json(silent=True) or {}
        link = _valid_meeting_link(data.get("meetingLink"))
        if not link:
            return error("Enter a valid meeting link (must start with http:// or https://).")

        actor_id = get_jwt_identity()
        actor_label = "Super Admin" if actor_role == "super_admin" else "College Admin"
        db.workshop_sessions.update_one({"_id": oid}, {"$set": {
            "meetingLink": link,
            "meetingLinkUpdatedAt": now(),
            "meetingLinkUpdatedBy": {"id": str(actor_id), "name": actor_label, "role": actor_role},
            "updatedAt": now(),
        }})
        updated = db.workshop_sessions.find_one({"_id": oid})
        log_activity(
            db, actor_id, actor_role, "workshop_session_meeting_link",
            f"Updated meeting link for session \"{existing.get('name', '—')}\"",
            meta={"sessionId": str(oid)},
        )
        return ok({"session": workshop_session_public(updated)}, message="Meeting link saved.")

    @bp.route("/interventions/sessions/<session_id>/attendance", methods=["GET"])
    @role_required("super_admin")
    def session_attendance_students(session_id):
        """Mark Attendance for a single session. Returns exactly the
        students that session was scheduled for (college + department +
        cohort filters, or its explicit student list), each with the
        status already saved on the session date if any — so re-opening
        Mark Attendance loads existing statuses instead of a blank slate
        (spec §22–26)."""
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        doc = db.workshop_sessions.find_one({"_id": oid})
        if not doc:
            return error("Session not found.", 404)
        if doc.get("attendanceRequirement") != "mandatory":
            return error("Attendance is optional for this session.", 409)
        if not _attendance_window_open(doc):
            state = _attendance_window_state(doc)
            if state == "upcoming":
                return error(
                    f"Attendance is locked — it opens at the session start time ({doc.get('date')} {doc.get('startTime')}).",
                    409)
            return error("Attendance is closed — the session window has ended.", 409)

        students = _session_target_students(db, doc)
        date_str = doc.get("date")
        saved = {}
        if date_str:
            for r in db.attendance.find(
                {"studentId": {"$in": [s["_id"] for s in students]}, "date": date_str},
                {"studentId": 1, "status": 1},
            ):
                saved[r["studentId"]] = r.get("status")

        rows = []
        for s in students:
            cohort = student_cohort_label(s)
            rows.append({
                "studentId": str(s["_id"]),
                "studentName": s.get("fullName") or s.get("email") or "—",
                "rollNumber": s.get("rollNumber") or "—",
                "college": s.get("college") or "—",
                "department": s.get("department") or "—",
                "cohort": cohort,
                "cohortLabel": ATTENDANCE_COHORT_LABELS.get(cohort, cohort),
                "status": saved.get(s["_id"], "not_marked"),
            })
        return ok({
            "session": workshop_session_public(doc),
            "date": date_str,
            "students": rows,
            "total": len(rows),
        })

    @bp.route("/interventions/sessions/<session_id>/attendance", methods=["POST"])
    @role_required("super_admin")
    def save_session_attendance(session_id):
        """Persist attendance for one session. Each student gets their own
        record associated with the session (sessionId/sessionName) on the
        session date — statuses are upserted per student so re-saving never
        creates duplicates and existing statuses are updated (spec §26)."""
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        doc = db.workshop_sessions.find_one({"_id": oid})
        if not doc:
            return error("Session not found.", 404)
        if doc.get("attendanceRequirement") != "mandatory":
            return error("Attendance is optional for this session.", 409)
        if not _attendance_window_open(doc):
            state = _attendance_window_state(doc)
            if state == "upcoming":
                return error(
                    f"Attendance is locked — it opens at the session start time ({doc.get('date')} {doc.get('startTime')}).",
                    409)
            return error("Attendance is closed — the session window has ended.", 409)
        date_str = doc.get("date")
        if not _ATTENDANCE_DATE_RE.match(date_str or ""):
            return error("This session has no valid date to mark attendance on.")

        data = request.get_json(silent=True) or {}
        marks = data.get("marks")
        if not isinstance(marks, list) or not marks:
            return error("No attendance marks provided.")
        if len(marks) > 2000:
            return error("Too many records in one submission.")

        target = {s["_id"] for s in _session_target_students(db, doc)}
        status_by_oid = {}
        for m in marks:
            sid = m.get("studentId")
            status = (m.get("status") or "").strip().lower()
            if status not in ATTENDANCE_STATUSES:
                return error("Attendance status must be 'present' or 'absent'.")
            student_oid = to_object_id(sid)
            if not student_oid or student_oid not in target:
                return error("One of the selected students is not in this session's audience.")
            status_by_oid[student_oid] = status

        students_by_id = {d["_id"]: d for d in users.find({"_id": {"$in": list(status_by_oid.keys())}})}
        actor_id = get_jwt_identity()
        now_ts = now()
        inserted = updated = 0
        for oid_student, status in status_by_oid.items():
            student = students_by_id[oid_student]
            cohort = student_cohort_label(student)
            payload = {
                "studentId": oid_student,
                "date": date_str,
                "sessionId": str(oid),
                "sessionName": doc.get("name"),
                "studentName": student.get("fullName") or "—",
                "rollNumber": student.get("rollNumber") or "—",
                "collegeId": student.get("collegeId"),
                "college": student.get("college") or "—",
                "departmentId": student.get("departmentId"),
                "department": student.get("department") or "—",
                "cohort": cohort,
                "status": status,
                "markedBy": {"id": str(actor_id), "name": "Super Admin", "role": "super_admin"},
                "markedAt": now_ts,
                "updatedAt": now_ts,
            }
            result = db.attendance.update_one(
                {"studentId": oid_student, "date": date_str},
                {"$set": payload},
                upsert=True,
            )
            if result.upserted_id:
                inserted += 1
            else:
                updated += 1

        present = sum(1 for s in status_by_oid.values() if s == "present")
        absent = len(status_by_oid) - present
        log_activity(
            db, actor_id, "super_admin", "session_attendance_marked",
            f"Marked attendance for session \"{doc.get('name', '—')}\" ({len(status_by_oid)} students)",
            meta={"sessionId": str(oid), "date": date_str, "present": present, "absent": absent},
        )
        return ok({
            "saved": len(status_by_oid),
            "inserted": inserted,
            "updated": updated,
            "present": present,
            "absent": absent,
        }, message="Attendance saved.")

    @bp.route("/interventions/sessions/<session_id>/approve", methods=["POST"])
    @role_required("super_admin")
    def approve_workshop_session(session_id):
        """Approval workflow: marks a scheduled session as approved so it
        is live for students. Only Super Admin can approve. Persists to
        db.workshop_sessions and is recorded in the audit log."""
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        existing = db.workshop_sessions.find_one({"_id": oid})
        if not existing:
            return error("Session not found.", 404)
        if existing.get("approvalStatus") == "approved":
            return error("This session is already approved.", 409)
        db.workshop_sessions.update_one(
            {"_id": oid},
            {"$set": {"approvalStatus": "approved", "updatedAt": now()}},
        )
        updated = db.workshop_sessions.find_one({"_id": oid})
        log_activity(
            db, get_jwt_identity(), "super_admin", "workshop_session_approved",
            f"Approved session \"{existing.get('name', '—')}\"",
            meta={"sessionId": str(oid)},
        )
        return ok({"session": workshop_session_public(updated)}, message="Session approved.")

    @bp.route("/interventions/sessions/<session_id>/reject", methods=["POST"])
    @role_required("super_admin")
    def reject_workshop_session(session_id):
        """Approval workflow: marks a scheduled session as rejected. The
        record is kept (auditable) with its approval status persisted."""
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        existing = db.workshop_sessions.find_one({"_id": oid})
        if not existing:
            return error("Session not found.", 404)
        if existing.get("approvalStatus") == "rejected":
            return error("This session is already rejected.", 409)
        db.workshop_sessions.update_one(
            {"_id": oid},
            {"$set": {"approvalStatus": "rejected", "updatedAt": now()}},
        )
        updated = db.workshop_sessions.find_one({"_id": oid})
        log_activity(
            db, get_jwt_identity(), "super_admin", "workshop_session_rejected",
            f"Rejected session \"{existing.get('name', '—')}\"",
            meta={"sessionId": str(oid)},
        )
        return ok({"session": workshop_session_public(updated)}, message="Session rejected.")

    @bp.route("/interventions/sessions/<session_id>", methods=["DELETE"])
    @role_required("super_admin")
    def delete_workshop_session(session_id):
        oid = to_object_id(session_id)
        if not oid:
            return error("Invalid session id.", 404)
        existing = db.workshop_sessions.find_one({"_id": oid})
        if not existing:
            return error("Session not found.", 404)
        # Business rule: a session can only be deleted before it starts.
        if _session_started(existing):
            return error("Sessions can only be deleted before they start.", 409)
        db.workshop_sessions.delete_one({"_id": oid})
        log_activity(
            db, get_jwt_identity(), "super_admin", "workshop_session_deleted",
            f"Deleted session \"{existing.get('name', '—')}\"",
            meta={"sessionId": str(oid)},
        )
        return ok(message="Session deleted.")

    # ==========================================================
    # SUPER ADMIN — ATTENDANCE MODULE (db.attendance).
    # All four dropdowns on Mark Attendance + View Records are DB-backed:
    #   colleges    -> db.colleges (active only)
    #   departments -> db.departments (active only, joined by college)
    #   cohorts     -> the platform's shared cohort system (A/B/C/Entry)
    #   trainers    -> db.users role=trainer (approved/suspended)
    # No student records are ever returned until "Fetch Records" is
    # clicked, and re-marking a student on an existing date upserts the
    # same (studentId, date) document rather than creating a duplicate.
    # ==========================================================
    @bp.route("/attendance/filters", methods=["GET"])
    @role_required("super_admin")
    def attendance_filters():
        college_docs = list(db.colleges.find({"status": "active"}).sort("college_name", 1))
        dept_docs = list(db.departments.find({"status": "active"}).sort("department_name", 1))
        trainer_docs = list(db.users.find(
            {"role": "trainer", "approvalStatus": {"$in": ["approved", "suspended"]}}
        ).sort("fullName", 1))
        return ok({
            "colleges": [{"id": str(c["_id"]), "name": c.get("college_name")} for c in college_docs],
            "departments": [
                {"id": str(d["_id"]), "collegeId": str(d["college_id"]), "name": d.get("department_name")}
                for d in dept_docs
            ],
            "cohorts": [{"value": "all", "label": "All Cohorts"}] + [
                {"value": c, "label": ATTENDANCE_COHORT_LABELS[c]} for c in sorted(VALID_COHORTS)
            ] + [{"value": ENTRY_LEVEL, "label": ATTENDANCE_COHORT_LABELS[ENTRY_LEVEL]}],
            "trainers": [{"id": str(t["_id"]), "name": t.get("fullName") or "—"} for t in trainer_docs],
        })

    @bp.route("/attendance/students", methods=["GET"])
    @role_required("super_admin")
    def list_attendance_students():
        college_ids, err = _parse_optional_id_list(request.args.get("collegeIds"))
        if err:
            return error(err)
        department_ids, err = _parse_optional_id_list(request.args.get("departmentIds"))
        if err:
            return error(err)
        cohort_param = (request.args.get("cohorts") or "").strip()
        cohorts = [c for c in cohort_param.split(",") if c.strip()] if cohort_param else []
        allowed = set(VALID_COHORTS) | {ENTRY_LEVEL, "all"}
        if any(c not in allowed for c in cohorts):
            return error("Invalid cohort filter.")
        query = _attendance_student_query(db, college_ids, department_ids, cohorts)
        students = list(db.users.find(query).sort("fullName", 1))
        return ok({"students": [_attendance_student_public(d) for d in students], "total": len(students)})

    @bp.route("/attendance/mark", methods=["POST"])
    @role_required("super_admin")
    def mark_attendance():
        data = request.get_json(silent=True) or {}
        date_str = (data.get("date") or "").strip()
        if not _ATTENDANCE_DATE_RE.match(date_str):
            return error("A valid date (YYYY-MM-DD) is required.")

        marks = data.get("marks")
        if not isinstance(marks, list) or not marks:
            return error("No attendance marks provided.")
        if len(marks) > 2000:
            return error("Too many records in one submission.")

        student_ids = []
        status_by_oid = {}
        for m in marks:
            sid = m.get("studentId")
            status = (m.get("status") or "").strip().lower()
            if status not in ATTENDANCE_STATUSES:
                return error("Attendance status must be 'present' or 'absent'.")
            oid = to_object_id(sid)
            if not oid:
                return error("One of the student ids is invalid.")
            if oid not in status_by_oid:
                status_by_oid[oid] = status
                student_ids.append(oid)

        students_by_id = {d["_id"]: d for d in db.users.find(
            {"_id": {"$in": student_ids}, "role": "student"})}
        missing = [str(oid) for oid in student_ids if oid not in students_by_id]
        if missing:
            return error("One or more selected students no longer exist.")

        actor_id = get_jwt_identity()
        now_ts = now()
        inserted = updated = 0
        for oid, status in status_by_oid.items():
            student = students_by_id[oid]
            cohort = student_cohort_label(student)
            payload = {
                "studentId": oid,
                "date": date_str,
                "studentName": student.get("fullName") or "—",
                "rollNumber": student.get("rollNumber") or "—",
                "collegeId": student.get("collegeId"),
                "college": student.get("college") or "—",
                "departmentId": student.get("departmentId"),
                "department": student.get("department") or "—",
                "cohort": cohort,
                "status": status,
                "markedBy": {"id": str(actor_id), "name": "Super Admin", "role": "super_admin"},
                "markedAt": now_ts,
                "updatedAt": now_ts,
            }
            result = db.attendance.update_one(
                {"studentId": oid, "date": date_str},
                {"$set": payload},
                upsert=True,
            )
            if result.upserted_id:
                inserted += 1
            else:
                updated += 1

        present = sum(1 for s in status_by_oid.values() if s == "present")
        absent = len(status_by_oid) - present
        log_activity(
            db, actor_id, "super_admin", "attendance_marked",
            f"Marked attendance for {len(status_by_oid)} students on {date_str}",
            meta={"date": date_str, "present": present, "absent": absent},
        )
        return ok({
            "saved": len(status_by_oid),
            "inserted": inserted,
            "updated": updated,
            "present": present,
            "absent": absent,
        }, message="Attendance saved.")

    @bp.route("/attendance/records", methods=["GET"])
    @role_required("super_admin")
    def list_attendance_records():
        college_ids, err = _parse_optional_id_list(request.args.get("collegeIds"))
        if err:
            return error(err)
        department_ids, err = _parse_optional_id_list(request.args.get("departmentIds"))
        if err:
            return error(err)
        trainer_ids, err = _parse_optional_id_list(request.args.get("trainerIds"))
        if err:
            return error(err)
        date_str = (request.args.get("date") or "").strip()
        if not _ATTENDANCE_DATE_RE.match(date_str):
            return error("A valid date (YYYY-MM-DD) is required.")

        # All students matching the college/department scope (cohort is not a
        # View Records filter), so "Not Yet Marked" can be derived for the
        # selected date.
        student_query = _attendance_student_query(db, college_ids, department_ids, [])
        students = list(db.users.find(student_query).sort("fullName", 1))
        student_ids = [d["_id"] for d in students]

        record_query = {"date": date_str, "studentId": {"$in": student_ids}}
        if trainer_ids:
            record_query["markedBy.id"] = {"$in": [str(t) for t in trainer_ids]}
        records = list(db.attendance.find(record_query))
        record_by_student = {r["studentId"]: r for r in records}

        rows = []
        for student in students:
            marked = record_by_student.get(student["_id"])
            cohort = student_cohort_label(student)
            rows.append({
                "studentId": str(student["_id"]),
                "studentName": student.get("fullName") or "—",
                "rollNumber": student.get("rollNumber") or "—",
                "college": student.get("college") or "—",
                "department": student.get("department") or "—",
                "cohort": cohort,
                "cohortLabel": ATTENDANCE_COHORT_LABELS.get(cohort, cohort),
                "status": marked["status"] if marked else "not_marked",
                "markedBy": marked["markedBy"] if marked else None,
                "markedAt": marked["markedAt"] if marked else None,
            })

        present = sum(1 for r in rows if r["status"] == "present")
        absent = sum(1 for r in rows if r["status"] == "absent")
        not_marked = sum(1 for r in rows if r["status"] == "not_marked")
        return ok({
            "records": rows,
            "date": date_str,
            "present": present,
            "absent": absent,
            "notMarked": not_marked,
            "total": len(rows),
        })


    return bp
